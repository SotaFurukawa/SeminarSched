"""Googleフォームの生徒・講師回答を一括検証し講習データへ反映する。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Final, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from summer_scheduler.application.project_service import ProjectService
from summer_scheduler.domain.grades import grade_from_excel, grade_to_excel
from summer_scheduler.infrastructure.db.models import (
    Assignment,
    AuditLog,
    ImportBatch,
    ImportSourceSnapshot,
    LessonRequest,
    RegularLessonProfile,
    Student,
    StudentAvailability,
    Subject,
    Teacher,
    TeacherAvailability,
    TimeSlot,
)
from summer_scheduler.infrastructure.importing import (
    CsvEncoding,
    SourceTable,
    inspect_source,
    read_source_table,
)
from summer_scheduler.infrastructure.repositories.master_repository import MasterRepository

_DATE_PATTERN: Final = re.compile(r"(20\d{2})[-/年](\d{1,2})[-/月](\d{1,2})")
_RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

IssueSeverity = Literal["error", "warning"]


class CourseSurveyError(ValueError):
    """一括アンケート取込みを安全に完了できない。"""


@dataclass(frozen=True, slots=True)
class CourseSurveyIssue:
    severity: IssueSeverity
    source: str
    row: int
    person_name: str
    message: str
    resolution: str


@dataclass(frozen=True, slots=True)
class _StudentResponse:
    row: int
    name: str
    grade: str
    enrollment_type: str
    requests: tuple[tuple[str, int], ...]
    unavailable: frozenset[tuple[date, str]]
    note: str


@dataclass(frozen=True, slots=True)
class _TeacherResponse:
    row: int
    name: str
    unavailable: frozenset[tuple[date, str]]
    note: str


@dataclass(frozen=True, slots=True)
class CourseSurveyPreview:
    project_id: int
    student_path: Path
    teacher_path: Path
    student_sha256: str
    teacher_sha256: str
    open_dates: tuple[date, ...]
    slot_codes: tuple[str, ...]
    student_table: SourceTable
    teacher_table: SourceTable
    students: tuple[_StudentResponse, ...]
    teachers: tuple[_TeacherResponse, ...]
    issues: tuple[CourseSurveyIssue, ...]

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == "error" for issue in self.issues)


@dataclass(frozen=True, slots=True)
class CourseSurveyApplyResult:
    students: int
    teachers: int
    lesson_requests: int
    trial_students: int
    warnings: int


class CourseSurveyService:
    """生成済みGoogleフォームの回答2ファイルをまとめて扱う。"""

    def __init__(self, projects: ProjectService) -> None:
        self._projects = projects

    def prepare(
        self,
        student_path: Path,
        teacher_path: Path,
        *,
        trial_student_rows: frozenset[int] = frozenset(),
    ) -> CourseSurveyPreview:
        project = self._projects.require_project()
        student_table = _read_first_table(student_path)
        teacher_table = _read_first_table(teacher_path)
        database = self._projects.require_database()
        with database.session_factory() as session:
            open_dates = {
                row.date
                for row in MasterRepository(session).list_open_dates(project_id=project.project_id)
                if row.is_open
            }
            slots = tuple(
                session.scalars(
                    select(TimeSlot)
                    .where(TimeSlot.project_id == project.project_id, TimeSlot.enabled.is_(True))
                    .order_by(TimeSlot.sort_order)
                )
            )
            students = list(session.scalars(select(Student).where(Student.active.is_(True))))
            teachers = list(session.scalars(select(Teacher).where(Teacher.active.is_(True))))
            subjects = list(session.scalars(select(Subject).where(Subject.active.is_(True))))

        issues: list[CourseSurveyIssue] = []
        student_responses = _parse_student_responses(
            student_table,
            open_dates,
            slots,
            students,
            subjects,
            issues,
            trial_student_rows,
        )
        teacher_responses = _parse_teacher_responses(
            teacher_table,
            open_dates,
            slots,
            teachers,
            issues,
        )
        return CourseSurveyPreview(
            project.project_id,
            student_table.source_path,
            teacher_table.source_path,
            _file_hash(student_table.source_path),
            _file_hash(teacher_table.source_path),
            tuple(sorted(open_dates)),
            tuple(slot.code for slot in slots),
            student_table,
            teacher_table,
            tuple(student_responses),
            tuple(teacher_responses),
            tuple(issues),
        )

    def apply(self, preview: CourseSurveyPreview) -> CourseSurveyApplyResult:
        project = self._projects.require_project()
        if preview.project_id != project.project_id:
            raise CourseSurveyError("別のプロジェクトで検証した回答は反映できません。")
        if preview.has_errors:
            raise CourseSurveyError("赤色の入力エラーを解消してから反映してください。")
        if (
            _file_hash(preview.student_path) != preview.student_sha256
            or _file_hash(preview.teacher_path) != preview.teacher_sha256
        ):
            raise CourseSurveyError("検証後に回答ファイルが変更されました。再度検証してください。")
        database = self._projects.require_database()
        combined = _combined_workbook(preview)
        student_content = preview.student_path.read_bytes()
        teacher_content = preview.teacher_path.read_bytes()
        with database.session_factory.begin() as session:
            assignment_count = session.scalar(
                select(func.count(Assignment.id)).where(Assignment.project_id == project.project_id)
            )
            if assignment_count:
                raise CourseSurveyError(
                    "時間割配置後は回答の一括置換ができません。先に時間割を未配置へ戻すか、"
                    "従来の差分取込みを使用してください。"
                )
            student_by_key = {
                (_name_key(row.name), grade_from_excel(row.grade)): row
                for row in session.scalars(select(Student).where(Student.active.is_(True)))
            }
            teacher_by_name = {
                _name_key(row.name): row
                for row in session.scalars(select(Teacher).where(Teacher.active.is_(True)))
            }
            subject_by_name = {
                _text_key(row.display_name): row
                for row in session.scalars(select(Subject).where(Subject.active.is_(True)))
            }
            slots = tuple(
                session.scalars(
                    select(TimeSlot)
                    .where(TimeSlot.project_id == project.project_id, TimeSlot.enabled.is_(True))
                    .order_by(TimeSlot.sort_order)
                )
            )
            open_dates = tuple(
                row.date
                for row in MasterRepository(session).list_open_dates(project_id=project.project_id)
                if row.is_open
            )
            profiles = {
                (row.student_id, row.subject_id): row
                for row in session.scalars(
                    select(RegularLessonProfile).where(
                        RegularLessonProfile.project_id == project.project_id
                    )
                )
            }
            trial_count = 0
            request_count = 0
            seen_student_ids: set[int] = set()
            for response in preview.students:
                student = student_by_key.get((_name_key(response.name), response.grade))
                if student is None:
                    if response.enrollment_type != "体験生":
                        raise CourseSurveyError(f"未登録の在籍生です: {response.name}")
                    student = Student(
                        external_id=_next_trial_id(session),
                        name=response.name,
                        grade=response.grade,
                        default_max_consecutive_slots=2,
                        allow_gap=False,
                        active=True,
                        note="在籍区分: 体験生（アンケート取込みで作成）",
                    )
                    session.add(student)
                    session.flush()
                    student_by_key[(_name_key(response.name), response.grade)] = student
                    trial_count += 1
                seen_student_ids.add(student.id)
                session.execute(
                    delete(StudentAvailability).where(
                        StudentAvailability.project_id == project.project_id,
                        StudentAvailability.student_id == student.id,
                    )
                )
                session.execute(
                    delete(LessonRequest).where(
                        LessonRequest.project_id == project.project_id,
                        LessonRequest.student_id == student.id,
                    )
                )
                for subject_name, required_sessions in response.requests:
                    subject = subject_by_name[_text_key(subject_name)]
                    profile = profiles.get((student.id, subject.id))
                    session.add(
                        LessonRequest(
                            project_id=project.project_id,
                            student_id=student.id,
                            subject_id=subject.id,
                            required_sessions=required_sessions,
                            regular_teacher_id_optional=(
                                profile.regular_teacher_id_optional if profile else None
                            ),
                            regular_teacher_priority=(
                                profile.regular_teacher_priority if profile else 3
                            ),
                            one_to_one_required=(profile.one_to_one_required if profile else False),
                            note=_joined_note(profile.note if profile else None, response.note),
                        )
                    )
                    request_count += 1
                session.add_all(
                    StudentAvailability(
                        project_id=project.project_id,
                        student_id=student.id,
                        date=day,
                        time_slot_id=slot.id,
                        availability_level=(0 if (day, slot.code) in response.unavailable else 1),
                    )
                    for day in open_dates
                    for slot in slots
                )

            for teacher_response in preview.teachers:
                teacher = teacher_by_name[_name_key(teacher_response.name)]
                session.execute(
                    delete(TeacherAvailability).where(
                        TeacherAvailability.project_id == project.project_id,
                        TeacherAvailability.teacher_id == teacher.id,
                    )
                )
                session.add_all(
                    TeacherAvailability(
                        project_id=project.project_id,
                        teacher_id=teacher.id,
                        date=day,
                        time_slot_id=slot.id,
                        availability_level=(
                            0 if (day, slot.code) in teacher_response.unavailable else 1
                        ),
                    )
                    for day in open_dates
                    for slot in slots
                )

            repository = MasterRepository(session)
            now = datetime.now(UTC)
            for import_type, source_path, content in (
                ("student_availability", preview.student_path, student_content),
                ("teacher_availability", preview.teacher_path, teacher_content),
                ("combined_course_survey", Path("講習アンケート統合.xlsx"), combined),
            ):
                repository.replace_import_source_snapshot(
                    ImportSourceSnapshot(
                        project_id=project.project_id,
                        import_type=import_type,
                        source_file_name=source_path.name,
                        content=content,
                        sha256=sha256(content).hexdigest(),
                        size_bytes=len(content),
                        imported_at=now,
                    )
                )
            repository.create_import_batch(
                ImportBatch(
                    project_id=project.project_id,
                    import_type="combined_course_survey",
                    source_file_name="講習アンケート統合.xlsx",
                    row_count=len(preview.students) + len(preview.teachers),
                    success_count=len(preview.students) + len(preview.teachers),
                    warning_count=sum(item.severity == "warning" for item in preview.issues),
                    error_count=0,
                    mapping_json='{"format":"generated_google_forms_v1"}',
                )
            )
            repository.create_audit_log(
                AuditLog(
                    project_id=project.project_id,
                    entity_type="course_survey",
                    entity_id=str(project.project_id),
                    action="combined_import",
                    before_json="{}",
                    after_json=(
                        f'{{"students":{len(preview.students)},'
                        f'"teachers":{len(preview.teachers)},'
                        f'"lesson_requests":{request_count}}}'
                    ),
                    reason="生徒・講師Googleフォーム回答の一括取込み",
                    source="import",
                )
            )
        return CourseSurveyApplyResult(
            len(preview.students),
            len(preview.teachers),
            request_count,
            trial_count,
            sum(item.severity == "warning" for item in preview.issues),
        )

    def export_latest_combined(self, path: Path) -> Path:
        project = self._projects.require_project()
        database = self._projects.require_database()
        with database.session_factory() as session:
            snapshot = MasterRepository(session).get_import_source_snapshot(
                project_id=project.project_id,
                import_type="combined_course_survey",
            )
            if snapshot is None:
                raise CourseSurveyError("まだ統合アンケートを取り込んでいません。")
            content = snapshot.content
        destination = path.expanduser().resolve()
        if destination.suffix.casefold() != ".xlsx":
            destination = destination.with_suffix(".xlsx")
        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary: Path | None = None
        try:
            with NamedTemporaryFile(
                prefix=f".{destination.stem}_",
                suffix=".xlsx.tmp",
                dir=destination.parent,
                delete=False,
            ) as stream:
                temporary = Path(stream.name)
                stream.write(content)
            os.replace(temporary, destination)
            temporary = None
        finally:
            if temporary is not None:
                temporary.unlink(missing_ok=True)
        return destination


def _parse_student_responses(
    table: SourceTable,
    open_dates: set[date],
    slots: tuple[TimeSlot, ...],
    master_students: list[Student],
    subjects: list[Subject],
    issues: list[CourseSurveyIssue],
    trial_student_rows: frozenset[int],
) -> list[_StudentResponse]:
    surname_header = _find_header(table, "姓（苗字）")
    given_header = _find_header(table, "名（必須）")
    grade_header = _find_header(table, "学年（必須）")
    enrollment_header = _find_header(table, "在籍区分")
    _require_headers(
        ((surname_header, "姓"), (given_header, "名"), (grade_header, "学年")),
        "生徒回答",
    )
    request_headers = [header for header in table.headers if "受講教科（" in header]
    count_headers = [header for header in table.headers if "受講回数（" in header]
    if not request_headers or len(request_headers) != len(count_headers):
        raise CourseSurveyError("生徒回答の受講教科・受講回数列を判別できません。")
    date_headers = _date_headers(table, "受講不可日時")
    _validate_dates(date_headers, open_dates, "生徒回答", issues)
    known_students = {(_name_key(row.name), grade_from_excel(row.grade)) for row in master_students}
    known_subjects = {_text_key(row.display_name) for row in subjects}
    result: list[_StudentResponse] = []
    seen: set[tuple[str, str]] = set()
    for source_row in table.rows:
        values = source_row.raw_values
        name = _full_name(values.get(surname_header), values.get(given_header))
        grade = grade_from_excel(_text(values.get(grade_header)))
        enrollment = (
            "体験生"
            if source_row.row_number in trial_student_rows
            else (_text(values.get(enrollment_header)) or "在籍生")
        )
        identity = (_name_key(name), grade)
        if identity in seen:
            issues.append(
                _issue(
                    "error",
                    "生徒回答",
                    source_row.row_number,
                    name,
                    "同じ生徒の回答が重複しています。",
                    "Googleフォーム側で正しい1回答だけを残す",
                )
            )
        seen.add(identity)
        if identity not in known_students:
            if enrollment == "体験生":
                issues.append(
                    _issue(
                        "warning",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        "基本情報にない体験生です。プロジェクト内だけに自動登録します。",
                        "体験生として登録",
                    )
                )
            else:
                issues.append(
                    _issue(
                        "error",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        "生徒・講師の基本情報に一致する在籍生がいません。",
                        "共通名簿へ追加して再反映、または回答を体験生へ修正",
                    )
                )
        requests: list[tuple[str, int]] = []
        request_keys: set[str] = set()
        for subject_header, count_header in zip(request_headers, count_headers, strict=True):
            subject_name = _text(values.get(subject_header))
            count_text = _text(values.get(count_header))
            if not subject_name and not count_text:
                continue
            if not subject_name or not count_text:
                issues.append(
                    _issue(
                        "error",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        "受講教科と受講回数は組で入力してください。",
                        "フォーム回答を修正",
                    )
                )
                continue
            try:
                count = int(count_text)
            except ValueError:
                count = 0
            if count < 1:
                issues.append(
                    _issue(
                        "error",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        "受講回数は1以上で指定してください。",
                        "フォーム回答を修正",
                    )
                )
            canonical_subject_name = _canonical_questionnaire_subject(
                subject_name,
                subject_header,
            )
            key = _text_key(canonical_subject_name)
            if key not in known_subjects:
                issues.append(
                    _issue(
                        "error",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        f"未登録の科目です: {subject_name}",
                        "フォームをアプリから再生成するか科目名を修正",
                    )
                )
            if key in request_keys:
                issues.append(
                    _issue(
                        "error",
                        "生徒回答",
                        source_row.row_number,
                        name,
                        f"同じ科目が重複しています: {subject_name}",
                        "重複する受講教科を修正",
                    )
                )
            request_keys.add(key)
            requests.append((canonical_subject_name, count))
        if not requests:
            issues.append(
                _issue(
                    "error",
                    "生徒回答",
                    source_row.row_number,
                    name,
                    "受講教科が1件もありません。",
                    "少なくとも1教科を回答",
                )
            )
        result.append(
            _StudentResponse(
                source_row.row_number,
                name,
                grade,
                enrollment,
                tuple(requests),
                _unavailable(values, date_headers, slots),
                _find_note(values, ("特記事項",)),
            )
        )
    return result


def _parse_teacher_responses(
    table: SourceTable,
    open_dates: set[date],
    slots: tuple[TimeSlot, ...],
    master_teachers: list[Teacher],
    issues: list[CourseSurveyIssue],
) -> list[_TeacherResponse]:
    surname_header = _find_header(table, "姓（苗字）")
    given_header = _find_header(table, "名（必須）")
    _require_headers(((surname_header, "姓"), (given_header, "名")), "講師回答")
    date_headers = _date_headers(table, "出勤不可日時")
    _validate_dates(date_headers, open_dates, "講師回答", issues)
    known = {_name_key(row.name) for row in master_teachers}
    result: list[_TeacherResponse] = []
    seen: set[str] = set()
    for source_row in table.rows:
        values = source_row.raw_values
        name = _full_name(values.get(surname_header), values.get(given_header))
        key = _name_key(name)
        if key in seen:
            issues.append(
                _issue(
                    "error",
                    "講師回答",
                    source_row.row_number,
                    name,
                    "同じ講師の回答が重複しています。",
                    "Googleフォーム側で正しい1回答だけを残す",
                )
            )
        seen.add(key)
        if key not in known:
            issues.append(
                _issue(
                    "error",
                    "講師回答",
                    source_row.row_number,
                    name,
                    "生徒・講師の基本情報に一致する講師がいません。",
                    "共通名簿へ追加して再反映、または氏名を修正",
                )
            )
        result.append(
            _TeacherResponse(
                source_row.row_number,
                name,
                _unavailable(values, date_headers, slots),
                _find_note(values, ("勤務に関する特記事項", "特記事項")),
            )
        )
    return result


def _combined_workbook(preview: CourseSurveyPreview) -> bytes:
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)
    for title, table in (
        ("生徒回答原本", preview.student_table),
        ("講師回答原本", preview.teacher_table),
    ):
        sheet = workbook.create_sheet(title)
        sheet.append(table.headers)
        for row in table.rows:
            sheet.append([row.raw_values.get(header) for header in table.headers])
        _style_header(sheet)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    requests = workbook.create_sheet("受講希望（正規化）")
    requests.append(("回答行", "氏名", "学年", "在籍区分", "科目", "必要回数"))
    for student in preview.students:
        for subject, count in student.requests:
            requests.append(
                (
                    student.row,
                    student.name,
                    grade_to_excel(student.grade),
                    student.enrollment_type,
                    subject,
                    count,
                )
            )
    _style_header(requests)
    availability = workbook.create_sheet("可用性（正規化）")
    availability.append(("区分", "氏名", "日付", "コマ", "可否"))
    for student in preview.students:
        for day in preview.open_dates:
            for slot_code in preview.slot_codes:
                availability.append(
                    (
                        "生徒",
                        student.name,
                        day.isoformat(),
                        slot_code,
                        "不可" if (day, slot_code) in student.unavailable else "可能",
                    )
                )
    for teacher in preview.teachers:
        for day in preview.open_dates:
            for slot_code in preview.slot_codes:
                availability.append(
                    (
                        "講師",
                        teacher.name,
                        day.isoformat(),
                        slot_code,
                        "不可" if (day, slot_code) in teacher.unavailable else "可能",
                    )
                )
    _style_header(availability)
    availability.freeze_panes = "A2"
    availability.auto_filter.ref = availability.dimensions
    issues = workbook.create_sheet("要確認")
    issues.append(("重大度", "原本", "行", "氏名", "内容", "解消方法"))
    for item in preview.issues:
        issues.append(
            (item.severity, item.source, item.row, item.person_name, item.message, item.resolution)
        )
        fill = _RED_FILL if item.severity == "error" else _YELLOW_FILL
        for cell in issues[issues.max_row]:
            cell.fill = fill
    _style_header(issues)
    validation = DataValidation(
        type="list",
        formula1='"共通名簿へ追加,体験生として登録,回答を修正,対応不要"',
        allow_blank=True,
    )
    issues.add_data_validation(validation)
    validation.add(f"F2:F{max(issues.max_row, 2)}")
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _read_first_table(path: Path) -> SourceTable:
    inspection = inspect_source(path, csv_encoding=CsvEncoding.AUTO)
    sheet_name = inspection.sheets[0].name if inspection.sheets else None
    return read_source_table(path, sheet_name=sheet_name, csv_encoding=CsvEncoding.AUTO)


def _date_headers(table: SourceTable, marker: str) -> dict[str, date]:
    result: dict[str, date] = {}
    for header in table.headers:
        if marker not in header:
            continue
        match = _DATE_PATTERN.search(header)
        if match:
            result[header] = date(*(int(value) for value in match.groups()))
    return result


def _validate_dates(
    headers: dict[str, date],
    open_dates: set[date],
    source: str,
    issues: list[CourseSurveyIssue],
) -> None:
    represented = set(headers.values())
    for missing in sorted(open_dates - represented):
        issues.append(
            _issue(
                "error",
                source,
                1,
                "",
                f"開校日 {missing.isoformat()} の不可時間列がありません。",
                "現在のプロジェクトからGoogleフォームを再生成",
            )
        )
    for extra in sorted(represented - open_dates):
        issues.append(
            _issue(
                "warning",
                source,
                1,
                "",
                f"現在は開校日でない {extra.isoformat()} の回答列は無視します。",
                "対応不要",
            )
        )


def _unavailable(
    values: object,
    date_headers: dict[str, date],
    slots: tuple[TimeSlot, ...],
) -> frozenset[tuple[date, str]]:
    mapping = values
    result: set[tuple[date, str]] = set()
    if not hasattr(mapping, "get"):
        return frozenset()
    for header, day in date_headers.items():
        cell = _text(mapping.get(header))
        for slot in slots:
            if re.search(
                rf"(?:^|[,、;\s]){re.escape(slot.code)}(?:$|[,、;\s])",
                cell,
            ):
                result.add((day, slot.code))
    return frozenset(result)


def _find_header(table: SourceTable, marker: str) -> str:
    return next((header for header in table.headers if marker in header), "")


def _require_headers(headers: tuple[tuple[str, str], ...], source: str) -> None:
    missing = [label for header, label in headers if not header]
    if missing:
        raise CourseSurveyError(f"{source}に必要な列がありません: {'、'.join(missing)}")


def _find_note(values: object, markers: tuple[str, ...]) -> str:
    if not hasattr(values, "items"):
        return ""
    for header, value in values.items():
        if any(marker in str(header) for marker in markers):
            return _text(value)
    return ""


def _issue(
    severity: IssueSeverity,
    source: str,
    row: int,
    name: str,
    message: str,
    resolution: str,
) -> CourseSurveyIssue:
    return CourseSurveyIssue(severity, source, row, name, message, resolution)


def _file_hash(path: Path) -> str:
    try:
        return sha256(path.read_bytes()).hexdigest()
    except OSError as exc:
        raise CourseSurveyError(f"回答ファイルを読み込めません: {path.name}") from exc


def _next_trial_id(session: Session) -> str:
    existing = {str(value) for value in session.scalars(select(Student.external_id))}
    number = 1
    while f"TRIAL-{number:04d}" in existing:
        number += 1
    return f"TRIAL-{number:04d}"


def _joined_note(profile_note: str | None, response_note: str) -> str | None:
    parts = [value.strip() for value in (profile_note or "", response_note) if value.strip()]
    return " / ".join(parts) or None


def _full_name(surname: object, given_name: object) -> str:
    return " ".join(value for value in (_text(surname), _text(given_name)) if value)


def _name_key(value: str) -> str:
    return "".join(value.replace("　", " ").split()).casefold()


def _text_key(value: str) -> str:
    return "".join(value.split()).casefold()


def _canonical_questionnaire_subject(value: str, header: str) -> str:
    """短縮したフォーム選択肢をDB・統合xlsxの正式科目名へ戻す。"""
    normalized = value.replace("（中学受験以外）", "（中学受験以外なら可能）")
    if normalized.startswith(("小学校・", "中学校・", "高校・")):
        return normalized
    if "他学年" in header:
        match = re.fullmatch(r"(.+)[(（]([小中高])[)）]", normalized)
        if match:
            prefix = {"小": "小学校・", "中": "中学校・", "高": "高校・"}[match.group(2)]
            return prefix + match.group(1)
    for marker, prefix in (
        ("小学校", "小学校・"),
        ("中学校", "中学校・"),
        ("高校", "高校・"),
    ):
        if marker in header:
            return prefix + normalized
    return normalized


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _style_header(sheet: object) -> None:
    for cell in sheet[1]:  # type: ignore[index]
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


__all__ = [
    "CourseSurveyApplyResult",
    "CourseSurveyError",
    "CourseSurveyIssue",
    "CourseSurveyPreview",
    "CourseSurveyService",
]
