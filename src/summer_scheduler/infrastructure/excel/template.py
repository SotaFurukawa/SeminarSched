"""master_data.xlsxテンプレートの生成。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import os
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Final, cast

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from summer_scheduler.domain.grades import EXCEL_GRADE_OPTIONS, grade_to_excel
from summer_scheduler.infrastructure.excel.schema import (
    LESSON_REQUEST_SHEET,
    MASTER_DATA_SHEETS,
    QUALIFICATION_SHEET,
    SHEET_NAMES,
    STUDENT_SHEET,
    SUBJECT_SHEET,
    TEACHER_SHEET,
    ColumnSpec,
    SheetSpec,
    ValueKind,
)

_HEADER_FILL: Final = PatternFill(fill_type="solid", fgColor="1F4E78")
_HELPER_HEADER_FILL: Final = PatternFill(fill_type="solid", fgColor="5B9BD5")
_EXAMPLE_FILL: Final = PatternFill(fill_type="solid", fgColor="FFF2CC")
_INACTIVE_FILL: Final = PatternFill(fill_type="solid", fgColor="E7E6E6")
_HEADER_FONT: Final = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT: Final = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MAX_INPUT_ROW: Final = 10_000
_FORMULA_INPUT_ROW: Final = 1_000
_formula_rule = cast(Callable[..., Rule], FormulaRule)


@dataclass(frozen=True, slots=True)
class _ReferenceHelper:
    """ID・code列へ名前選択と確認表示を追加する定義。"""

    key: str
    select_header: str
    confirm_header: str
    source_sheet: str
    source_id_header: str
    source_name_header: str


@dataclass(frozen=True, slots=True)
class _LayoutColumn:
    """Excel上のcanonical列または任意補助列。"""

    header: str
    width: float
    comment: str
    canonical: ColumnSpec | None = None
    helper: _ReferenceHelper | None = None
    helper_kind: str | None = None


_REFERENCE_HELPERS: Final = {
    QUALIFICATION_SHEET.name: (
        _ReferenceHelper(
            "teacher_external_id",
            "講師名から選択",
            "講師名（確認）",
            TEACHER_SHEET.name,
            "講師ID",
            "氏名",
        ),
        _ReferenceHelper(
            "subject_code",
            "科目名から選択",
            "科目名（確認）",
            SUBJECT_SHEET.name,
            "科目コード",
            "表示名",
        ),
    ),
    LESSON_REQUEST_SHEET.name: (
        _ReferenceHelper(
            "student_external_id",
            "生徒名から選択",
            "生徒名（確認）",
            STUDENT_SHEET.name,
            "生徒ID",
            "氏名",
        ),
        _ReferenceHelper(
            "subject_code",
            "科目名から選択",
            "科目名（確認）",
            SUBJECT_SHEET.name,
            "科目コード",
            "表示名",
        ),
        _ReferenceHelper(
            "regular_teacher_external_id",
            "通常担当講師名から選択",
            "通常担当講師名（確認）",
            TEACHER_SHEET.name,
            "講師ID",
            "氏名",
        ),
        _ReferenceHelper(
            "preferred_teacher_1_external_id",
            "第1希望講師名から選択",
            "第1希望講師名（確認）",
            TEACHER_SHEET.name,
            "講師ID",
            "氏名",
        ),
        _ReferenceHelper(
            "preferred_teacher_2_external_id",
            "第2希望講師名から選択",
            "第2希望講師名（確認）",
            TEACHER_SHEET.name,
            "講師ID",
            "氏名",
        ),
        _ReferenceHelper(
            "preferred_teacher_3_external_id",
            "第3希望講師名から選択",
            "第3希望講師名（確認）",
            TEACHER_SHEET.name,
            "講師ID",
            "氏名",
        ),
    ),
}


def write_master_data_workbook(
    path: Path,
    rows_by_sheet: Mapping[str, Sequence[Mapping[str, object]]],
) -> None:
    """5シートのマスターデータブックを原子的に保存する。"""
    destination = path.expanduser()
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.worksheets[0])
    workbook.properties.title = "夏期講習時間割 マスターデータ"
    workbook.properties.subject = "生徒・講師・科目・講師対応科目・受講希望"
    workbook.properties.creator = "夏期講習時間割作成アプリ"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    for sheet_spec in MASTER_DATA_SHEETS:
        worksheet = workbook.create_sheet(sheet_spec.name)
        _write_sheet(
            worksheet,
            sheet_spec,
            rows_by_sheet.get(sheet_spec.name, ()),
        )

    _add_cross_sheet_validations(workbook)
    _add_reference_helpers(workbook)

    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            prefix=f".{destination.stem}_",
            suffix=".xlsx.tmp",
            dir=destination.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
        temporary_path = None
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _write_sheet(
    worksheet: Any,
    sheet_spec: SheetSpec,
    rows: Sequence[Mapping[str, object]],
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 34

    layout = _sheet_layout(sheet_spec)
    worksheet.append([column.header for column in layout])
    worksheet.append(_row_values(layout, sheet_spec.example))
    for row in rows:
        actual_row = dict(row)
        actual_row["is_example"] = False
        worksheet.append(_row_values(layout, actual_row))

    for column_number, layout_column in enumerate(layout, start=1):
        cell = worksheet.cell(row=1, column=column_number)
        cell.fill = _HEADER_FILL if layout_column.canonical is not None else _HELPER_HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.comment = Comment(layout_column.comment, "夏期講習時間割作成アプリ")
        worksheet.column_dimensions[get_column_letter(column_number)].width = layout_column.width
        if layout_column.canonical is not None:
            _add_column_validation(worksheet, column_number, layout_column.canonical)

    last_column = get_column_letter(len(layout))
    last_row = max(worksheet.max_row, 2)
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    worksheet.auto_filter.add_filter_column(0, ["いいえ"])

    for cell in worksheet[2]:
        cell.fill = _EXAMPLE_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in worksheet.iter_rows(min_row=3, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.conditional_formatting.add(
        f"A2:{last_column}{_MAX_INPUT_ROW}",
        _formula_rule(formula=['$A2="はい"'], fill=_EXAMPLE_FILL),
    )
    active_header = next(
        (
            index
            for index, column in enumerate(layout, start=1)
            if column.canonical is not None and column.canonical.key == "active"
        ),
        None,
    )
    if active_header is not None:
        active_letter = get_column_letter(active_header)
        worksheet.conditional_formatting.add(
            f"A2:{last_column}{_MAX_INPUT_ROW}",
            _formula_rule(
                formula=[f'${active_letter}2="いいえ"'],
                fill=_INACTIVE_FILL,
            ),
        )


def _sheet_layout(sheet_spec: SheetSpec) -> tuple[_LayoutColumn, ...]:
    helpers = {helper.key: helper for helper in _REFERENCE_HELPERS.get(sheet_spec.name, ())}
    layout: list[_LayoutColumn] = []
    for column in sheet_spec.columns:
        layout.append(
            _LayoutColumn(column.template_header, column.width, column.comment, canonical=column)
        )
        helper = helpers.get(column.key)
        if helper is not None:
            layout.extend(
                (
                    _LayoutColumn(
                        helper.select_header,
                        22,
                        "任意。名前から選ぶと左のID・科目コードを自動算出します。"
                        "同姓同名の場合はIDを直接入力してください。",
                        helper=helper,
                        helper_kind="select",
                    ),
                    _LayoutColumn(
                        helper.confirm_header,
                        22,
                        "左のID・科目コードに対応する表示名です。確認用のため直接編集しません。",
                        helper=helper,
                        helper_kind="confirm",
                    ),
                )
            )
    return tuple(layout)


def _row_values(layout: Sequence[_LayoutColumn], row: Mapping[str, object]) -> list[object]:
    return [
        _excel_column_value(column.canonical, row.get(column.canonical.key))
        if column.canonical is not None
        else None
        for column in layout
    ]


def _excel_column_value(column: ColumnSpec, value: object) -> object:
    if column.key == "grade" and isinstance(value, str):
        return grade_to_excel(value)
    return _excel_value(value)


def _excel_value(value: object) -> object:
    if isinstance(value, bool):
        return "はい" if value else "いいえ"
    return value


def _add_column_validation(
    worksheet: Any,
    column_number: int,
    column: ColumnSpec,
) -> None:
    column_letter = get_column_letter(column_number)
    target = f"{column_letter}2:{column_letter}{_MAX_INPUT_ROW}"
    validation: DataValidation | None = None

    if column.kind is ValueKind.BOOLEAN:
        validation = DataValidation(
            type="list",
            formula1='"はい,いいえ"',
            allow_blank=not column.required or column.uses_default_when_blank,
        )
        validation.error = "「はい」または「いいえ」を選択してください。"
    elif column.kind is ValueKind.INTEGER:
        if column.minimum is not None and column.maximum is not None:
            validation = DataValidation(
                type="whole",
                operator="between",
                formula1=str(column.minimum),
                formula2=str(column.maximum),
                allow_blank=not column.required or column.uses_default_when_blank,
            )
        elif column.minimum is not None:
            validation = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1=str(column.minimum),
                allow_blank=not column.required or column.uses_default_when_blank,
            )
        if validation is not None:
            validation.error = "指定された範囲内の整数を入力してください。"
    elif column.key == "school_level":
        validation = DataValidation(
            type="list",
            formula1='"小学校,中学校,高校"',
            allow_blank=False,
        )
        validation.error = "小学校・中学校・高校から選択してください。"
    elif column.key == "grade":
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(EXCEL_GRADE_OPTIONS)}"',
            allow_blank=False,
        )
        validation.error = "S1～S6、J1～J3、H1～H3から選択してください。"

    if validation is not None:
        validation.errorTitle = "入力値を確認してください"
        validation.showErrorMessage = True
        validation.promptTitle = column.header
        validation.prompt = column.comment
        validation.showInputMessage = True
        worksheet.add_data_validation(validation)
        validation.add(target)


def _add_cross_sheet_validations(workbook: Any) -> None:
    student_ids = "=INDIRECT(\"'生徒'!$B$2:$B$10000\")"
    teacher_ids = "=INDIRECT(\"'講師'!$B$2:$B$10000\")"
    subject_codes = "=INDIRECT(\"'科目'!$B$2:$B$10000\")"

    qualification_sheet = workbook[QUALIFICATION_SHEET.name]
    _add_list_validation(
        qualification_sheet,
        _column_letter(qualification_sheet, "講師ID"),
        teacher_ids,
        False,
    )
    _add_list_validation(
        qualification_sheet,
        _column_letter(qualification_sheet, "科目コード"),
        subject_codes,
        False,
    )

    request_sheet = workbook[LESSON_REQUEST_SHEET.name]
    _add_list_validation(
        request_sheet,
        _column_letter(request_sheet, "生徒ID"),
        student_ids,
        False,
    )
    _add_list_validation(
        request_sheet,
        _column_letter(request_sheet, "科目コード"),
        subject_codes,
        False,
    )
    for header in (
        "通常担当講師ID",
        "第1希望講師ID",
        "第2希望講師ID",
        "第3希望講師ID",
    ):
        _add_list_validation(
            request_sheet,
            _column_letter(request_sheet, header),
            teacher_ids,
            True,
        )

    # 定義変更時にシート名の取りこぼしを即座に検出する。
    if tuple(workbook.sheetnames) != SHEET_NAMES:
        raise AssertionError("master_data.xlsxのシート順が定義と一致しません。")
    if workbook[STUDENT_SHEET.name] is None:
        raise AssertionError
    if workbook[TEACHER_SHEET.name] is None:
        raise AssertionError
    if workbook[SUBJECT_SHEET.name] is None:
        raise AssertionError


def _add_list_validation(
    worksheet: Any,
    column_letter: str,
    formula: str,
    allow_blank: bool,
) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
    )
    validation.errorTitle = "参照IDを確認してください"
    validation.error = "対応するマスターシートに存在するIDを選択してください。"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{_MAX_INPUT_ROW}")


def _add_reference_helpers(workbook: Any) -> None:
    """名前選択からIDを算出し、IDに対応する名前を表示する。"""
    for sheet_name, helpers in _REFERENCE_HELPERS.items():
        worksheet = workbook[sheet_name]
        first_input_row = max(worksheet.max_row + 1, 3)
        for helper in helpers:
            id_letter = _column_letter(
                worksheet,
                next(
                    column.header
                    for column in next(
                        spec for spec in MASTER_DATA_SHEETS if spec.name == sheet_name
                    ).columns
                    if column.key == helper.key
                ),
            )
            select_letter = _column_letter(worksheet, helper.select_header)
            confirm_letter = _column_letter(worksheet, helper.confirm_header)
            source_id_letter = _column_letter(
                workbook[helper.source_sheet], helper.source_id_header
            )
            source_name_letter = _column_letter(
                workbook[helper.source_sheet], helper.source_name_header
            )
            source_ids = (
                f"'{helper.source_sheet}'!${source_id_letter}$3:"
                f"${source_id_letter}${_MAX_INPUT_ROW}"
            )
            source_names = (
                f"'{helper.source_sheet}'!${source_name_letter}$3:"
                f"${source_name_letter}${_MAX_INPUT_ROW}"
            )
            _add_list_validation(
                worksheet,
                select_letter,
                f"=INDIRECT(\"'{helper.source_sheet}'!${source_name_letter}$3:"
                f'${source_name_letter}${_MAX_INPUT_ROW}")',
                True,
            )
            worksheet[f"{confirm_letter}2"] = _example_reference_name(helper)
            for row_number in range(3, _FORMULA_INPUT_ROW + 1):
                id_cell = f"{id_letter}{row_number}"
                select_cell = f"{select_letter}{row_number}"
                confirm_cell = f"{confirm_letter}{row_number}"
                if row_number >= first_input_row:
                    worksheet[id_cell] = (
                        f'=IF({select_cell}="","",IF(COUNTIF({source_names},'
                        f"{select_cell})=1,INDEX({source_ids},MATCH({select_cell},"
                        f'{source_names},0)),""))'
                    )
                worksheet[confirm_cell] = (
                    f'=IF({id_cell}="","",IFERROR(INDEX({source_names},'
                    f'MATCH({id_cell},{source_ids},0)),"ID不明"))'
                )
                worksheet[select_cell].alignment = Alignment(vertical="top", wrap_text=True)
                worksheet[confirm_cell].alignment = Alignment(vertical="top", wrap_text=True)


def _column_letter(worksheet: Any, header: str) -> str:
    sheet_spec = next(
        (spec for spec in MASTER_DATA_SHEETS if spec.name == worksheet.title),
        None,
    )
    accepted = {header}
    if sheet_spec is not None:
        accepted.update(
            column.template_header for column in sheet_spec.columns if column.header == header
        )
    for column_number in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=column_number).value in accepted:
            return get_column_letter(column_number)
    raise AssertionError(f"{worksheet.title}に列「{header}」がありません。")


def _example_reference_name(helper: _ReferenceHelper) -> str:
    source_spec = next(spec for spec in MASTER_DATA_SHEETS if spec.name == helper.source_sheet)
    id_key = next(
        column.key for column in source_spec.columns if column.header == helper.source_id_header
    )
    name_key = next(
        column.key for column in source_spec.columns if column.header == helper.source_name_header
    )
    if source_spec.example.get(id_key) is None:
        raise AssertionError("例示行の参照IDがありません。")
    value = source_spec.example.get(name_key)
    if not isinstance(value, str):
        raise AssertionError("例示行の参照名がありません。")
    return value
