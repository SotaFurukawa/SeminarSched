"""現実的な匿名大人数データでPhase 6帳票の分割とExcel往復を確認する。"""

from __future__ import annotations

from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

from summer_scheduler.infrastructure.exporting import ExcelRenderer
from summer_scheduler.reporting.data import (
    AssignmentRecord,
    DateRecord,
    GroupLessonRecord,
    LessonRequestRecord,
    OutputSnapshot,
    ProjectRecord,
    SlotRecord,
    StudentRecord,
    SubjectRecord,
    TeacherRecord,
    UnassignedRecord,
    WarningRecord,
)
from summer_scheduler.reporting.issue_builder import build_issues_document
from summer_scheduler.reporting.settings import OutputSettings
from summer_scheduler.reporting.student_builder import build_student_document
from summer_scheduler.reporting.teacher_builder import build_teacher_document
from summer_scheduler.reporting.timetable_builder import build_timetable_document


def test_150_students_40_teachers_40_days_round_trip_all_excel_reports(
    tmp_path: Path,
) -> None:
    snapshot = _large_anonymous_snapshot()
    settings = OutputSettings(
        project_id=1,
        days_per_page=7,
        teacher_columns_per_page=8,
        font_size=7.0,
    )
    documents = (
        build_timetable_document(snapshot, settings),
        build_student_document(snapshot, settings),
        build_teacher_document(snapshot, settings),
        build_issues_document(snapshot, settings),
    )

    assert documents[0].page_count == 32
    assert documents[1].page_count == 150
    assert documents[2].page_count == 242
    assert documents[3].page_count == 14

    for document in documents:
        destination = tmp_path / f"匿名大人数_{document.report_code}.xlsx"
        ExcelRenderer(settings.style_rules).render(document, destination)
        workbook = load_workbook(destination, read_only=False)
        try:
            assert workbook.sheetnames
            assert workbook.active is not None
            assert workbook.active["A1"].value == document.title
            assert all(worksheet.print_title_rows is not None for worksheet in workbook.worksheets)
            assert sum(
                len(worksheet.row_breaks.brk) for worksheet in workbook.worksheets
            ) == document.page_count - len(workbook.worksheets)
        finally:
            workbook.close()


def _large_anonymous_snapshot() -> OutputSnapshot:
    first_day = date(2026, 8, 1)
    dates = tuple(DateRecord(first_day + timedelta(days=index), True, "") for index in range(40))
    slots = tuple(
        SlotRecord(
            id=index,
            code=code,
            display_name=f"{code}コマ",
            start_time=time(9 + (index - 1) * 2),
            end_time=time(10 + (index - 1) * 2, 20),
            sort_order=index,
            enabled=True,
        )
        for index, code in enumerate(("Y", "Z", "A", "B", "C"), start=1)
    )
    students = tuple(
        StudentRecord(
            id=index,
            external_id=f"S-{index:03d}",
            name=f"匿名生徒{index:03d}",
            grade=f"架空学年{(index - 1) % 6 + 1}",
            note="",
            active=True,
        )
        for index in range(1, 151)
    )
    teachers = tuple(
        TeacherRecord(
            id=index,
            external_id=f"T-{index:03d}",
            name=f"匿名講師{index:02d}",
            note="",
            active=True,
        )
        for index in range(1, 41)
    )
    subject = SubjectRecord(1, "ANON", "架空科目", "架空校種")
    requests = tuple(
        LessonRequestRecord(
            id=index,
            student_id=index,
            subject_id=subject.id,
            required_sessions=2,
            regular_teacher_id_optional=(index - 1) % 40 + 1,
            regular_teacher_priority=3,
            one_to_one_required=False,
            note="",
        )
        for index in range(1, 151)
    )
    assignments = tuple(
        AssignmentRecord(
            id=index,
            lesson_request_id=index,
            session_index=1,
            day=dates[(index - 1) % 40].day,
            time_slot_id=(index - 1) // 40 + 1,
            teacher_id=(index - 1) % 40 + 1,
            is_locked=False,
            is_manual=False,
            note="",
        )
        for index in range(1, 151)
    )
    groups = tuple(
        GroupLessonRecord(
            id=index,
            group_code=f"G-{index:03d}",
            course_name=f"匿名コマ外講座{index:02d}",
            grade="架空学年",
            subject_id=subject.id,
            day=dates[index - 1].day,
            start_time=time(7),
            end_time=time(8),
            teacher_id_optional=index,
            student_ids=(index,),
            room="架空教室",
            note="",
        )
        for index in range(1, 31)
    )
    unassigned = tuple(
        UnassignedRecord(
            lesson_request_id=index,
            student_id=index,
            subject_id=subject.id,
            required_sessions=2,
            placed_sessions=1,
            missing_sessions=1,
            main_reason="匿名の候補不足",
            reason_codes=("no_candidate",),
            resolution_candidates=("設定を確認",),
            candidate_count=1,
            priority=3,
            regular_teacher_id_optional=(index - 1) % 40 + 1,
            one_to_one_required=False,
            note="",
        )
        for index in range(1, 151)
    )
    warnings = tuple(
        WarningRecord(
            severity="info",
            issue_type="anonymous_review",
            day_optional=assignment.day,
            slot_code=slots[assignment.time_slot_id - 1].code,
            student_name=students[index - 1].name,
            teacher_name=teachers[assignment.teacher_id - 1].name,
            content="匿名確認項目",
            status="未対応",
            student_ids=(index,),
            teacher_id_optional=assignment.teacher_id,
        )
        for index, assignment in enumerate(assignments, start=1)
    )
    return OutputSnapshot(
        project=ProjectRecord(
            id=1,
            title="匿名大人数 夏期講習",
            campus_name="架空校舎",
            start_date=dates[0].day,
            end_date=dates[-1].day,
            status="confirmed",
            generated_at=datetime(2026, 7, 29, 12, tzinfo=UTC),
        ),
        dates=dates,
        slots=slots,
        students=students,
        teachers=teachers,
        subjects=(subject,),
        lesson_requests=requests,
        assignments=assignments,
        group_lessons=groups,
        unassigned=unassigned,
        warnings=warnings,
    )
