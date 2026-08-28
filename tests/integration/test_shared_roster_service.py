"""講習から独立した共通名簿の統合テスト。"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from pathlib import Path
from typing import Any, cast
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from summer_scheduler.application.project_service import ProjectService
from summer_scheduler.application.shared_roster_service import SharedRosterService
from summer_scheduler.infrastructure.db import create_database, upgrade_database
from summer_scheduler.infrastructure.db.models import (
    RegularLessonProfile,
    Student,
    Subject,
    Teacher,
    TeacherQualification,
)
from summer_scheduler.infrastructure.excel.shared_roster import (
    SharedQualification,
    SharedRegularLesson,
    SharedRosterData,
    SharedStudent,
    SharedSubject,
    SharedTeacher,
    read_shared_roster,
    write_shared_roster,
)


@pytest.fixture
def roster_service(tmp_path: Path) -> Iterator[SharedRosterService]:
    registry = create_database(tmp_path / "registry.db")
    upgrade_database(registry.engine)
    projects = ProjectService(
        registry,
        tmp_path / "backups",
        workspace_directory=tmp_path / "workspace",
    )
    projects.create_project(
        tmp_path / "summer.jukuschedule",
        title="夏期講習",
        campus_name="テスト校",
        start_date=date(2026, 8, 1),
        end_date=date(2026, 8, 1),
    )
    service = SharedRosterService(projects)
    yield service
    projects.close_project()
    registry.dispose()


def test_shared_roster_syncs_people_qualifications_and_regular_lessons(
    roster_service: SharedRosterService,
) -> None:
    write_shared_roster(
        roster_service.path,
        SharedRosterData(
            students=(
                SharedStudent("S001", "山田", "花子", "中2"),
                SharedStudent("S002", "佐藤", "次郎", "中3", active=False),
            ),
            teachers=(SharedTeacher("T001", "田中", "太郎"),),
            subjects=(SharedSubject("JH_MATH", "中学校・数学", "junior_high", 1),),
            qualifications=(SharedQualification("T001", "JH_MATH"),),
            regular_lessons=(SharedRegularLesson("S001", "JH_MATH", "T001", 4, False),),
        ),
    )

    result = roster_service.sync_to_current_project()

    assert result.students == 2
    database = roster_service._projects.require_database()  # noqa: SLF001
    with database.session_factory() as session:
        students = list(session.scalars(select(Student).order_by(Student.external_id)))
        teacher = session.scalar(select(Teacher).where(Teacher.external_id == "T001"))
        qualification = session.scalar(select(TeacherQualification))
        profile = session.scalar(select(RegularLessonProfile))
        subjects = list(session.scalars(select(Subject).order_by(Subject.sort_order)))
    assert [(row.name, row.active) for row in students] == [
        ("山田 花子", True),
        ("佐藤 次郎", False),
    ]
    assert teacher is not None
    assert qualification is not None and qualification.can_teach is True
    assert profile is not None
    assert profile.regular_teacher_id_optional == teacher.id
    assert profile.regular_teacher_priority == 4
    assert len(subjects) == 26
    assert {row.display_name for row in subjects} >= {
        "小学校・算数（中学受験）",
        "小学校・算数（中学受験以外なら可能）",
        "小学校・国語（中学受験）",
        "小学校・国語（中学受験以外なら可能）",
        "高校・数学IA",
        "高校・数学IIBC",
        "高校・数学III",
    }

    workbook = load_workbook(roster_service.path, data_only=False)
    try:
        assert workbook["生徒"]["A2"].value is True
        assert workbook["生徒"]["A3"].value is False
        assert workbook["生徒"]["F2"].value == "J2"
        assert workbook["生徒"]["F3"].value == "J3"
        assert len(workbook["生徒"].conditional_formatting) > 0
        assert workbook["通常授業"]["G2"].value == 4
        assert workbook["科目"].max_row == 27
    finally:
        workbook.close()


def test_linked_names_follow_master_ids_and_can_be_reselected(tmp_path: Path) -> None:
    path = tmp_path / "linked_names.xlsx"
    write_shared_roster(
        path,
        SharedRosterData(
            students=(
                SharedStudent("S-0001", "山田", "花子", "中2"),
                SharedStudent("S-0002", "佐藤", "次郎", "中3"),
            ),
            teachers=(
                SharedTeacher("T-0001", "田中", "太郎"),
                SharedTeacher("T-0002", "鈴木", "一郎"),
            ),
            subjects=(SharedSubject("JH_MATH", "中学校・数学", "junior_high", 1),),
            qualifications=(SharedQualification("T-0001", "JH_MATH"),),
            regular_lessons=(SharedRegularLesson("S-0001", "JH_MATH", "T-0001", 3, False),),
        ),
    )

    workbook = load_workbook(path, data_only=False)
    try:
        qualification = workbook["講師対応科目"]
        regular = workbook["通常授業"]
        assert qualification["A2"].value.startswith('=IF(B2="","",IFERROR(INDEX(')
        assert qualification["B2"].value == "T-0001"
        assert regular["A2"].value.startswith('=IF(B2="","",IFERROR(INDEX(')
        assert regular["B2"].value == "S-0001"
        assert regular["C2"].value.startswith('=IF(D2="","",IFERROR(INDEX(')
        assert regular["D2"].value == "JH_MATH"
        assert regular["E2"].value.startswith('=IF(F2="","",IFERROR(INDEX(')
        assert regular["F2"].value == "T-0001"

        # マスター側の改名後も、通常授業は安定IDを参照するため同じ人物へ追従する。
        workbook["生徒"]["C2"] = "山本"
        workbook["講師"]["C2"] = "高橋"
        workbook.save(path)
    finally:
        workbook.close()

    renamed = read_shared_roster(path)
    assert renamed.students[0].name == "山本 花子"
    assert renamed.teachers[0].name == "高橋 太郎"
    assert renamed.regular_lessons[0].student_external_id == "S-0001"
    assert renamed.regular_lessons[0].regular_teacher_external_id == "T-0001"

    # 既存行の数式をプルダウン選択で上書きした場合は、新しい名前側を優先して関連を変更する。
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["講師対応科目"]["A2"] = "鈴木 一郎"
        workbook["通常授業"]["A2"] = "佐藤 次郎"
        workbook["通常授業"]["E2"] = "鈴木 一郎"
        workbook.save(path)
    finally:
        workbook.close()

    reselected = read_shared_roster(path)
    assert reselected.qualifications[0].teacher_external_id == "T-0002"
    assert reselected.regular_lessons[0].student_external_id == "S-0002"
    assert reselected.regular_lessons[0].regular_teacher_external_id == "T-0002"


def test_blank_ids_are_generated_and_persisted(roster_service: SharedRosterService) -> None:
    write_shared_roster(
        roster_service.path,
        SharedRosterData(
            students=(SharedStudent("", "山田", "花子", "小2"),),
            teachers=(SharedTeacher("", "田中", "太郎"),),
            subjects=(SharedSubject("ES_MATH", "小学校・算数", "elementary", 1),),
        ),
    )

    roster_service.sync_to_current_project()

    workbook = load_workbook(roster_service.path, data_only=False)
    try:
        assert workbook["生徒"]["B2"].value == "S-0001"
        assert workbook["生徒"]["F2"].value == "S2"
        assert workbook["講師"]["B2"].value == "T-0001"
    finally:
        workbook.close()


def test_shared_roster_template_export_and_import_keep_canonical_backup(
    roster_service: SharedRosterService,
    tmp_path: Path,
) -> None:
    write_shared_roster(
        roster_service.path,
        SharedRosterData(
            (SharedStudent("S-0001", "既存", "生徒", "中1"),),
            (),
            (),
        ),
    )
    template = roster_service.export_new_template(tmp_path / "新しい基本情報.xlsx")
    exported = read_shared_roster(template)
    assert exported.students == ()
    assert len(exported.subjects) >= 26

    incoming = tmp_path / "完成した基本情報.xlsx"
    write_shared_roster(
        incoming,
        SharedRosterData(
            (SharedStudent("S-0002", "取込", "生徒", "高1"),),
            (SharedTeacher("T-0001", "取込", "講師"),),
            (),
        ),
    )
    result = roster_service.import_workbook(incoming)

    assert result is not None and result.students == 1 and result.teachers == 1
    canonical = read_shared_roster(roster_service.path)
    assert canonical.students[0].external_id == "S-0002"
    backups = list((roster_service.path.parent / "基本情報バックアップ").glob("*.xlsx"))
    assert len(backups) == 1
    assert read_shared_roster(backups[0]).students[0].external_id == "S-0001"


def test_blank_ids_do_not_collide_with_ids_on_later_rows(
    roster_service: SharedRosterService,
) -> None:
    write_shared_roster(
        roster_service.path,
        SharedRosterData(
            students=(
                SharedStudent("", "青木", "花子", "小2"),
                SharedStudent("S-0001", "山田", "太郎", "中1"),
            ),
            teachers=(
                SharedTeacher("", "伊藤", "一郎"),
                SharedTeacher("T0001", "田中", "二郎"),
            ),
            subjects=(SharedSubject("ES_MATH", "小学校・算数", "elementary", 1),),
        ),
    )

    roster_service.sync_to_current_project()

    workbook = load_workbook(roster_service.path, data_only=False)
    try:
        student_ids = {workbook["生徒"][f"B{row}"].value for row in (2, 3)}
        teacher_ids = {workbook["講師"][f"B{row}"].value for row in (2, 3)}
        assert student_ids == {"S-0001", "S-0002"}
        assert teacher_ids == {"T0001", "T-0002"}
    finally:
        workbook.close()


def test_blank_ids_do_not_reuse_an_id_only_present_in_the_project(
    roster_service: SharedRosterService,
) -> None:
    database = roster_service._projects.require_database()  # noqa: SLF001
    with database.session_factory.begin() as session:
        session.add(Student(external_id="S-0001", name="過去 生徒", grade="中3", active=False))
        session.add(Teacher(external_id="T-0001", name="過去 講師", active=False))
    write_shared_roster(
        roster_service.path,
        SharedRosterData(
            students=(SharedStudent("", "新規", "生徒", "小2"),),
            teachers=(SharedTeacher("", "新規", "講師"),),
            subjects=(SharedSubject("ES_MATH", "小学校・算数", "elementary", 1),),
        ),
    )

    roster_service.sync_to_current_project()

    workbook = load_workbook(roster_service.path, data_only=False)
    try:
        assert workbook["生徒"]["B2"].value == "S-0002"
        assert workbook["講師"]["B2"].value == "T-0002"
    finally:
        workbook.close()


def test_shared_roster_prepares_status_first_defaults_and_required_cells(
    tmp_path: Path,
) -> None:
    path = tmp_path / "roster.xlsx"
    write_shared_roster(path, SharedRosterData((), (), ()))

    workbook = load_workbook(path, data_only=False)
    try:
        students = workbook["生徒"]
        teachers = workbook["講師"]
        assert students["A1"].value == "在籍"
        assert students["B1"].value == "生徒ID（自動・入力不要）"
        assert "デフォルトは2" in str(students["G1"].value)
        assert "デフォルトはなし" in str(students["H1"].value)
        assert students["A2"].value == '=C2<>""'
        assert students["B2"].value == (
            '=IF(C2="","",INDEX(\'_入力補助\'!$A$1:$A$999,COUNTIF($C$2:C2,"?*")))'
        )
        assert students["B3"].value == (
            '=IF(C3="","",INDEX(\'_入力補助\'!$A$1:$A$999,COUNTIF($C$2:C3,"?*")))'
        )
        assert students["G2"].value == '=IF(C2="","",2)'
        assert students["H2"].value == '=IF(C2="","","なし")'
        assert students["C2"].fill.fgColor.rgb == "FFFFF2CC"
        assert students["F2"].fill.fgColor.rgb == "FFFFF2CC"
        student_rules = [
            rule
            for rules in cast(Any, students.conditional_formatting)._cf_rules.values()  # noqa: SLF001
            for rule in rules
        ]
        assert any(rule.formula == ['AND($C2<>"",$A2=FALSE)'] for rule in student_rules)
        assert teachers["A2"].value == '=C2<>""'
        assert teachers["B2"].value == (
            '=IF(C2="","",INDEX(\'_入力補助\'!$B$1:$B$999,COUNTIF($C$2:C2,"?*")))'
        )
        assert teachers["F2"].value == '=IF(C2="","","なし")'
        assert teachers["C2"].fill.fgColor.rgb == "FFFFF2CC"
        qualifications = workbook["講師対応科目"]
        assert qualifications["A1"].value == "講師名から選択"
        assert qualifications["B1"].value == "講師ID（自動・入力不要）"
        assert qualifications["C1"].value == "科目名から選択"
        assert qualifications["D1"].value == "科目コード（自動・入力不要）"
        assert qualifications["B2"].value.startswith('=IF(A2="","",IFERROR(INDEX(')
        assert qualifications["D2"].value.startswith('=IF(C2="","",IFERROR(INDEX(')
        assert qualifications["E2"].value == '=IF(OR(A2="",C2=""),"","はい")'
        regular = workbook["通常授業"]
        assert regular["A1"].value == "生徒名から選択"
        assert regular["B1"].value == "生徒ID（自動・入力不要）"
        assert regular["G2"].value == '=IF(OR(A2="",C2=""),"",3)'
        assert regular["H2"].value == '=IF(OR(A2="",C2=""),"","いいえ")'
        assert "（確認）" not in "".join(
            str(regular.cell(1, column).value or "") for column in range(1, regular.max_column + 1)
        )
        assert students.auto_filter.ref == "A1:I2"
        assert teachers.auto_filter.ref == "A1:G2"
        helper = workbook["_入力補助"]
        assert helper.sheet_state == "veryHidden"
        assert helper["A1"].value == "S-0001"
        assert helper["A2"].value == "S-0002"
        assert helper["B1"].value == "T-0001"
        assert helper["B2"].value == "T-0002"
    finally:
        workbook.close()

    with ZipFile(path) as archive:
        assert "xl/featurePropertyBag/featurePropertyBag.xml" in archive.namelist()


def test_id_helper_lists_unused_ids_in_assignment_order(tmp_path: Path) -> None:
    path = tmp_path / "existing_ids.xlsx"
    write_shared_roster(
        path,
        SharedRosterData(
            (
                SharedStudent("S-0001", "一番", "", "小1"),
                SharedStudent("S-0003", "三番", "", "小3"),
            ),
            (SharedTeacher("T-0002", "二番", ""),),
            (),
        ),
    )

    workbook = load_workbook(path, data_only=False)
    try:
        helper = workbook["_入力補助"]
        assert [helper[f"A{row}"].value for row in range(1, 4)] == [
            "S-0002",
            "S-0004",
            "S-0005",
        ]
        assert [helper[f"B{row}"].value for row in range(1, 4)] == [
            "T-0001",
            "T-0003",
            "T-0004",
        ]
        assert workbook["生徒"]["B2"].value == "S-0001"
        assert workbook["生徒"]["B3"].value == "S-0003"
        assert workbook["講師"]["B2"].value == "T-0002"
        assert workbook["生徒"]["B4"].value == (
            '=IF(C4="","",INDEX(\'_入力補助\'!$A$1:$A$999,COUNTIF($C$4:C4,"?*")))'
        )
        assert workbook["講師"]["B3"].value == (
            '=IF(C3="","",INDEX(\'_入力補助\'!$B$1:$B$999,COUNTIF($C$3:C3,"?*")))'
        )
    finally:
        workbook.close()


def test_surname_only_input_gets_ids_and_defaults_when_imported(tmp_path: Path) -> None:
    path = tmp_path / "surname_only.xlsx"
    write_shared_roster(path, SharedRosterData((), (), ()))
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["生徒"]["C2"] = "新規"
        workbook["生徒"]["F2"] = "H2"
        workbook["講師"]["C2"] = "担当"
        workbook.save(path)
    finally:
        workbook.close()

    data = read_shared_roster(path)

    assert data.students == (SharedStudent("S-0001", "新規", "", "高2", 2, False, True, ""),)
    assert data.teachers == (SharedTeacher("T-0001", "担当", "", False, True, ""),)


def test_native_checkbox_false_is_imported_as_inactive(tmp_path: Path) -> None:
    path = tmp_path / "inactive.xlsx"
    write_shared_roster(path, SharedRosterData((), (), ()))
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["生徒"]["A2"] = False
        workbook["生徒"]["C2"] = "退籍"
        workbook["生徒"]["F2"] = "J3"
        workbook["講師"]["A2"] = False
        workbook["講師"]["C2"] = "退職"
        workbook.save(path)
    finally:
        workbook.close()

    data = read_shared_roster(path)

    assert data.students == (SharedStudent("S-0001", "退籍", "", "中3", 2, False, False, ""),)
    assert data.teachers == (SharedTeacher("T-0001", "退職", "", False, False, ""),)


def test_multiple_formula_id_rows_receive_distinct_ids_when_imported(tmp_path: Path) -> None:
    path = tmp_path / "multiple_new_people.xlsx"
    write_shared_roster(path, SharedRosterData((), (), ()))
    workbook = load_workbook(path, data_only=False)
    try:
        students = workbook["生徒"]
        students["C2"], students["F2"] = "一番", "S1"
        students["C3"], students["F3"] = "二番", "S2"
        teachers = workbook["講師"]
        teachers["C2"], teachers["C3"] = "一番", "二番"
        workbook.save(path)
    finally:
        workbook.close()

    data = read_shared_roster(path)

    assert [student.external_id for student in data.students] == ["S-0001", "S-0002"]
    assert [teacher.external_id for teacher in data.teachers] == ["T-0001", "T-0002"]


def test_shared_roster_reads_legacy_person_column_order(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xlsx"
    write_shared_roster(path, SharedRosterData((), (), ()))
    workbook = load_workbook(path)
    try:
        students = workbook["生徒"]
        students.delete_rows(1, students.max_row)
        students.append(
            (
                "生徒ID（自動・入力不要）",
                "姓（必須）",
                "名",
                "氏名（確認）",
                "学年（必須）",
                "標準最大連続コマ数",
                "空きコマ許可",
                "在籍",
                "備考",
            )
        )
        students.append(("S-0007", "旧式", "生徒", "旧式 生徒", "中2", 3, "はい", "☐ 退籍", "旧"))

        teachers = workbook["講師"]
        teachers.delete_rows(1, teachers.max_row)
        teachers.append(
            (
                "講師ID（自動・入力不要）",
                "姓（必須）",
                "名",
                "氏名（確認）",
                "空きコマ許可",
                "在籍",
                "備考",
            )
        )
        teachers.append(("T-0008", "旧式", "講師", "旧式 講師", "いいえ", "☑ 在籍", "旧"))
        workbook.save(path)
    finally:
        workbook.close()

    data = read_shared_roster(path)

    assert data.students == (SharedStudent("S-0007", "旧式", "生徒", "中2", 3, True, False, "旧"),)
    assert data.teachers == (SharedTeacher("T-0008", "旧式", "講師", False, True, "旧"),)


def test_shared_roster_reads_legacy_qualification_and_regular_lesson_columns(
    tmp_path: Path,
) -> None:
    path = tmp_path / "legacy_relations.xlsx"
    write_shared_roster(
        path,
        SharedRosterData(
            (SharedStudent("S-0001", "旧式", "生徒", "中2"),),
            (SharedTeacher("T-0001", "旧式", "講師"),),
            (SharedSubject("JH_MATH", "中学校・数学", "junior_high", 1),),
        ),
    )
    workbook = load_workbook(path)
    try:
        qualifications = workbook["講師対応科目"]
        qualifications.delete_rows(1, qualifications.max_row)
        qualifications.append(
            (
                "講師ID（自動・入力不要）",
                "講師名から選択",
                "講師名（確認）",
                "科目コード",
                "科目名から選択",
                "科目名（確認）",
                "指導可能",
                "備考",
            )
        )
        qualifications.append(
            ("T-0001", "", "旧式 講師", "JH_MATH", "", "中学校・数学", "はい", "旧資格")
        )
        regular = workbook["通常授業"]
        regular.delete_rows(1, regular.max_row)
        regular.append(
            (
                "生徒ID（自動・入力不要）",
                "生徒名から選択",
                "生徒名（確認）",
                "科目コード",
                "科目名から選択",
                "科目名（確認）",
                "通常担当講師ID（自動・入力不要）",
                "通常担当講師名から選択",
                "通常担当講師名（確認）",
                "担当講師優先度",
                "1対1必須",
                "備考",
            )
        )
        regular.append(
            (
                "S-0001",
                "",
                "旧式 生徒",
                "JH_MATH",
                "",
                "中学校・数学",
                "T-0001",
                "",
                "旧式 講師",
                4,
                "いいえ",
                "旧通常",
            )
        )
        workbook.save(path)
    finally:
        workbook.close()

    data = read_shared_roster(path)

    assert data.qualifications == (SharedQualification("T-0001", "JH_MATH", True, "旧資格"),)
    assert data.regular_lessons == (
        SharedRegularLesson("S-0001", "JH_MATH", "T-0001", 4, False, "旧通常"),
    )
