"""共通LayoutDocumentを編集可能な印刷用Excelへ変換する。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import re
from collections.abc import Sequence
from pathlib import Path
from typing import Final, cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from summer_scheduler.domain.grades import excelize_grades_in_text
from summer_scheduler.infrastructure.exporting.atomic_output import atomic_output_path
from summer_scheduler.infrastructure.exporting.errors import OutputRenderError
from summer_scheduler.reporting.layout import (
    LayoutCell,
    LayoutDocument,
    LayoutPage,
    LayoutRow,
    LayoutSection,
    LayoutTable,
)
from summer_scheduler.reporting.settings import (
    DEFAULT_STYLE_RULES,
    STYLE_RULE_PRIORITY,
    StyleRule,
)

_INVALID_SHEET_CHARACTER: Final = re.compile(r"[\\/*?:\[\]]")
_THIN_SIDE: Final = Side(style="thin", color="596579")
_CELL_BORDER: Final = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)
_ROLE_FILL: Final = {
    "title": "1F4E78",
    "subtitle": "D9EAF7",
    "metadata": "EAF0F6",
    "header": "1F4E78",
    "data": "FFFFFF",
    "legend": "F0F2F5",
    "closed": "333333",
}
_ROLE_TEXT: Final = {
    "title": "FFFFFF",
    "header": "FFFFFF",
    "closed": "FFFFFF",
}


class ExcelRenderer:
    """DBやQMLへ依存せず、共通ページモデルだけをxlsxへ描画する。"""

    def __init__(self, style_rules: Sequence[StyleRule] = DEFAULT_STYLE_RULES) -> None:
        self._style_rules = {rule.code: rule for rule in style_rules}
        if len(self._style_rules) != len(style_rules):
            raise ValueError("Excel表示ルールのコードが重複しています")
        for rule in style_rules:
            rule.validate()

    def render(
        self,
        document: LayoutDocument,
        destination: Path,
        *,
        overwrite: bool = False,
    ) -> Path:
        """LayoutDocumentを一時ファイルへ生成し、成功時だけ保存先へ置換する。"""

        target = _require_suffix(destination, ".xlsx")
        if document.page_count < 1:
            raise OutputRenderError("Excelへ出力するページがありません")

        workbook = Workbook()
        initial_sheet = workbook.active
        if initial_sheet is None:
            raise OutputRenderError("Excelブックの初期シートを作成できませんでした")
        workbook.remove(initial_sheet)
        workbook.properties.title = document.title
        workbook.properties.subject = document.report_code
        workbook.properties.creator = "季節講習時間割作成アプリ"
        used_names: set[str] = set()
        try:
            for section in document.sections:
                if not section.pages:
                    continue
                worksheet = workbook.create_sheet(_unique_sheet_name(section.name, used_names))
                self._render_section(worksheet, document, section)
            if not workbook.worksheets:
                raise OutputRenderError("Excelへ出力するセクションがありません")

            with atomic_output_path(target, overwrite=overwrite) as temporary:
                try:
                    workbook.save(temporary)
                except PermissionError:
                    raise
                except Exception as exc:
                    raise OutputRenderError(
                        "Excelファイルの生成に失敗しました。レイアウト設定を確認してください。"
                    ) from exc
        finally:
            workbook.close()
        return target.expanduser().resolve()

    def _render_section(
        self,
        worksheet: Worksheet,
        document: LayoutDocument,
        section: LayoutSection,
    ) -> None:
        max_columns = max(
            (len(table.column_widths) for page in section.pages for table in page.tables),
            default=1,
        )
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A4"
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(
            fitToPage=True,
            autoPageBreaks=False,
        )
        worksheet.page_setup.paperSize = (
            worksheet.PAPERSIZE_A3 if document.page_size == "A3" else worksheet.PAPERSIZE_A4
        )
        worksheet.page_setup.orientation = (
            worksheet.ORIENTATION_LANDSCAPE
            if document.orientation == "landscape"
            else worksheet.ORIENTATION_PORTRAIT
        )
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        margin_inches = document.margin_mm / 25.4
        worksheet.page_margins.left = margin_inches
        worksheet.page_margins.right = margin_inches
        worksheet.page_margins.top = margin_inches
        worksheet.page_margins.bottom = margin_inches
        worksheet.print_options.horizontalCentered = True

        self._write_merged_heading(
            worksheet,
            row_number=1,
            text=document.title,
            column_count=max_columns,
            role="title",
            font_size=max(14.0, document.font_size + 5.0),
        )
        metadata = f"{document.campus_name}／{document.course_name}"
        self._write_merged_heading(
            worksheet,
            row_number=2,
            text=metadata,
            column_count=max_columns,
            role="subtitle",
            font_size=max(10.0, document.font_size + 1.0),
        )
        self._write_merged_heading(
            worksheet,
            row_number=3,
            text=document.updated_text,
            column_count=max_columns,
            role="metadata",
            font_size=document.font_size,
        )
        repeat_header_rows = _first_table_repeat_header_rows(section)
        worksheet.print_title_rows = f"6:{5 + repeat_header_rows}" if repeat_header_rows else "1:3"

        row_cursor = 4
        page_end_rows: list[int] = []
        for page in section.pages:
            row_cursor = self._write_page(
                worksheet,
                document,
                page,
                start_row=row_cursor,
                max_columns=max_columns,
            )
            page_end_rows.append(row_cursor - 1)
            row_cursor += 1

        for page_end in page_end_rows[:-1]:
            worksheet.row_breaks.append(Break(id=page_end))

        max_row = max(1, row_cursor - 2)
        worksheet.print_area = f"A1:{get_column_letter(max_columns)}{max_row}"
        worksheet.auto_filter.ref = None
        odd_header = worksheet.oddHeader
        odd_footer = worksheet.oddFooter
        if odd_header is None or odd_footer is None:
            raise OutputRenderError("Excelのヘッダー・フッターを設定できませんでした")
        odd_header.left.text = document.campus_name
        odd_header.center.text = document.title
        odd_header.right.text = document.updated_text
        odd_footer.left.text = section.name
        odd_footer.center.text = "ページ &[Page] / &[Pages]"
        odd_footer.right.text = "個人情報を含みます"

    def _write_page(
        self,
        worksheet: Worksheet,
        document: LayoutDocument,
        page: LayoutPage,
        *,
        start_row: int,
        max_columns: int,
    ) -> int:
        row_cursor = start_row
        self._write_merged_heading(
            worksheet,
            row_number=row_cursor,
            text=page.heading,
            column_count=max_columns,
            role="subtitle",
            font_size=max(document.font_size + 2.0, 11.0),
        )
        row_cursor += 1
        self._write_merged_heading(
            worksheet,
            row_number=row_cursor,
            text=page.subheading,
            column_count=max_columns,
            role="metadata",
            font_size=document.font_size,
        )
        row_cursor += 1

        for table_index, table in enumerate(page.tables):
            if table_index:
                row_cursor += 1
            row_cursor = self._write_table(
                worksheet,
                table,
                start_row=row_cursor,
                font_size=document.font_size,
            )

        if page.footer_note:
            self._write_merged_heading(
                worksheet,
                row_number=row_cursor,
                text=page.footer_note,
                column_count=max_columns,
                role="legend",
                font_size=max(7.0, document.font_size - 1.0),
            )
            row_cursor += 1
        return row_cursor

    def _write_table(
        self,
        worksheet: Worksheet,
        table: LayoutTable,
        *,
        start_row: int,
        font_size: float,
    ) -> int:
        for column_index, width in enumerate(table.column_widths, start=1):
            letter = get_column_letter(column_index)
            existing = worksheet.column_dimensions[letter].width or 0
            worksheet.column_dimensions[letter].width = max(existing, width)

        occupied_until: dict[int, int] = {}
        for row_offset, layout_row in enumerate(table.rows):
            row_number = start_row + row_offset
            if layout_row.height_points_optional is not None:
                worksheet.row_dimensions[row_number].height = layout_row.height_points_optional
            self._write_layout_row(
                worksheet,
                layout_row,
                row_number=row_number,
                column_count=len(table.column_widths),
                occupied_until=occupied_until,
                font_size=font_size,
            )
        return start_row + len(table.rows)

    def _write_layout_row(
        self,
        worksheet: Worksheet,
        layout_row: LayoutRow,
        *,
        row_number: int,
        column_count: int,
        occupied_until: dict[int, int],
        font_size: float,
    ) -> None:
        column = 1
        for layout_cell in layout_row.cells:
            column = _next_available_column(
                column,
                row_number=row_number,
                occupied_until=occupied_until,
            )
            while not _span_is_available(
                column,
                layout_cell.column_span,
                row_number=row_number,
                occupied_until=occupied_until,
            ):
                column = _next_available_column(
                    column + 1,
                    row_number=row_number,
                    occupied_until=occupied_until,
                )
            last_column = column + layout_cell.column_span - 1
            if last_column > column_count:
                raise OutputRenderError("Excelレイアウトのセル数が定義された列幅を超えています")
            last_row = row_number + layout_cell.row_span - 1
            self._write_cell_range(
                worksheet,
                layout_cell,
                first_row=row_number,
                last_row=last_row,
                first_column=column,
                last_column=last_column,
                font_size=font_size,
            )
            if layout_cell.row_span > 1:
                for occupied_column in range(column, last_column + 1):
                    occupied_until[occupied_column] = last_row
            column = last_column + 1

    def _write_cell_range(
        self,
        worksheet: Worksheet,
        layout_cell: LayoutCell,
        *,
        first_row: int,
        last_row: int,
        first_column: int,
        last_column: int,
        font_size: float,
    ) -> None:
        fill_color, text_color = self._colors(layout_cell)
        for row_number in range(first_row, last_row + 1):
            for column_number in range(first_column, last_column + 1):
                cell = worksheet.cell(row=row_number, column=column_number)
                cell.border = _CELL_BORDER
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                cell.font = Font(
                    name="Yu Gothic UI",
                    size=font_size,
                    bold=layout_cell.role in {"title", "subtitle", "header"},
                    color=text_color,
                )
                cell.alignment = Alignment(
                    horizontal=layout_cell.alignment,
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=True,
                )
        value_cell = cast(Cell, worksheet.cell(row=first_row, column=first_column))
        _set_explicit_text(value_cell, excelize_grades_in_text(layout_cell.text))
        if first_row != last_row or first_column != last_column:
            worksheet.merge_cells(
                start_row=first_row,
                start_column=first_column,
                end_row=last_row,
                end_column=last_column,
            )

    def _write_merged_heading(
        self,
        worksheet: Worksheet,
        *,
        row_number: int,
        text: str,
        column_count: int,
        role: str,
        font_size: float,
    ) -> None:
        layout_cell = LayoutCell(
            text=text,
            role=role,  # type: ignore[arg-type]
            column_span=column_count,
            alignment="center",
        )
        self._write_cell_range(
            worksheet,
            layout_cell,
            first_row=row_number,
            last_row=row_number,
            first_column=1,
            last_column=column_count,
            font_size=font_size,
        )
        worksheet.row_dimensions[row_number].height = max(18.0, font_size * 1.8)

    def _colors(self, cell: LayoutCell) -> tuple[str, str]:
        for code in STYLE_RULE_PRIORITY:
            if code in cell.style_codes and code in self._style_rules:
                style = self._style_rules[code]
                return style.fill_color.removeprefix("#"), style.text_color.removeprefix("#")
        return _ROLE_FILL[cell.role], _ROLE_TEXT.get(cell.role, "18212F")


def _next_available_column(
    start: int,
    *,
    row_number: int,
    occupied_until: dict[int, int],
) -> int:
    column = start
    while occupied_until.get(column, 0) >= row_number:
        column += 1
    return column


def _span_is_available(
    start: int,
    span: int,
    *,
    row_number: int,
    occupied_until: dict[int, int],
) -> bool:
    return all(occupied_until.get(column, 0) < row_number for column in range(start, start + span))


def _set_explicit_text(cell: Cell, value: str) -> None:
    """先頭 ``=`` も数式ではなく利用者データとして保存する。"""

    cell.value = value
    cell.data_type = "s"


def _require_suffix(path: Path, suffix: str) -> Path:
    if path.suffix.casefold() != suffix:
        raise OutputRenderError(f"出力ファイルの拡張子は{suffix}を指定してください")
    return path


def _unique_sheet_name(value: str, used_names: set[str]) -> str:
    base = _INVALID_SHEET_CHARACTER.sub("_", value).strip("'").strip() or "出力"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.casefold() in used_names:
        suffix = f"_{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate.casefold())
    return candidate


def _first_table_repeat_header_rows(section: LayoutSection) -> int:
    """Excelの自動改ページ時に、先頭表の指定見出し行を反復する。"""

    for page in section.pages:
        if page.tables:
            return page.tables[0].repeat_header_rows
    return 0


__all__ = ["ExcelRenderer"]
