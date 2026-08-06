"""アンケート・集団授業xlsxテンプレートの原子的生成。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import os
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from summer_scheduler.infrastructure.importing.readers import (
    GROUP_LESSON_SHEET,
    GROUP_PARTICIPANT_SHEET,
)
from summer_scheduler.infrastructure.importing.schemas import (
    DEFAULT_SLOT_CODES,
    slot_key,
)

STUDENT_AVAILABILITY_SHEET = "生徒希望"
TEACHER_AVAILABILITY_SHEET = "講師希望"
INSTRUCTIONS_SHEET = "説明"
STUDENT_REFERENCE_SHEET = "生徒マスター参照"
TEACHER_REFERENCE_SHEET = "講師マスター参照"
SUBJECT_REFERENCE_SHEET = "科目マスター参照"

_HEADER_FILL: Final = PatternFill(fill_type="solid", fgColor="1F4E78")
_EXAMPLE_FILL: Final = PatternFill(fill_type="solid", fgColor="FFF2CC")
_HEADER_FONT: Final = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT: Final = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MAX_INPUT_ROW: Final = 10_000


@dataclass(frozen=True, slots=True)
class _TemplateColumn:
    key: str
    header: str
    width: float = 16
    comment: str = ""
    availability: bool = False
    date_value: bool = False
    time_value: bool = False
    required: bool = False

    @property
    def display_header(self) -> str:
        if self.required:
            return f"{self.header}（必須）"
        return self.header


def write_student_availability_template(
    path: Path,
    slot_codes: Sequence[str] = DEFAULT_SLOT_CODES,
    rows: Sequence[Mapping[str, object]] = (),
    reference_students: Sequence[Mapping[str, object]] = (),
    reference_teachers: Sequence[Mapping[str, object]] = (),
    reference_subjects: Sequence[Mapping[str, object]] = (),
) -> None:
    """生徒アンケートテンプレートを既存／架空行付きで出力する。"""
    slots = _validated_slots(slot_codes)
    columns = (
        _example_column(),
        _TemplateColumn("student_id", "生徒ID", comment="生徒マスターのID", required=True),
        _TemplateColumn("student_name", "生徒名", width=20, required=True),
        _TemplateColumn(
            "subject_code",
            "科目コード",
            comment="科目マスターのコード",
            required=True,
        ),
        _TemplateColumn("subject_name_confirm", "科目名（確認）", width=22),
        _TemplateColumn("date", "日付", date_value=True, required=True),
        *(_slot_column(code) for code in slots),
        _TemplateColumn("preferred_teacher_1", "第1希望講師ID", width=18),
        _TemplateColumn("preferred_teacher_2", "第2希望講師ID", width=18),
        _TemplateColumn("preferred_teacher_3", "第3希望講師ID", width=18),
        _TemplateColumn("note", "備考", width=28),
    )
    example = {
        "example": True,
        "student_id": "SAMPLE-S001",
        "student_name": "架空 花子",
        "subject_code": "JH_MATH",
        "date": "2026-08-01",
        **{slot_key(code): index % 3 for index, code in enumerate(slots)},
        "preferred_teacher_1": "SAMPLE-T001",
        "preferred_teacher_2": "",
        "preferred_teacher_3": "",
        "note": "この例示行は取込み対象外です",
    }
    workbook = _new_workbook("生徒アンケート入力テンプレート")
    _write_data_sheet(
        workbook,
        STUDENT_AVAILABILITY_SHEET,
        columns,
        example,
        rows,
    )
    _write_master_reference_sheets(
        workbook,
        reference_students,
        reference_teachers,
        reference_subjects,
    )
    _add_student_reference_helpers(workbook)
    _write_instructions_sheet(
        workbook,
        (
            ("用途", "生徒の受講可能日時・希望日時を入力します。"),
            ("コマ値", "0 = 不可、1 = 可能、2 = 希望"),
            ("例示行", "「例示行」が「はい」の行は取込み対象外です。"),
            ("ID", "人物は名前ではなく生徒ID・講師IDで照合します。"),
            (
                "変更禁止",
                "通常担当講師、優先度5、1対1契約はこの回答から変更できません。",
            ),
        ),
    )
    _atomic_save(workbook, path)


def write_teacher_availability_template(
    path: Path,
    slot_codes: Sequence[str] = DEFAULT_SLOT_CODES,
    rows: Sequence[Mapping[str, object]] = (),
    reference_students: Sequence[Mapping[str, object]] = (),
    reference_teachers: Sequence[Mapping[str, object]] = (),
    reference_subjects: Sequence[Mapping[str, object]] = (),
) -> None:
    """講師アンケートテンプレートを既存／架空行付きで出力する。"""
    slots = _validated_slots(slot_codes)
    columns = (
        _example_column(),
        _TemplateColumn("teacher_id", "講師ID", comment="講師マスターのID", required=True),
        _TemplateColumn("name", "講師名", width=20, required=True),
        _TemplateColumn("date", "日付", date_value=True, required=True),
        *(_slot_column(code) for code in slots),
        _TemplateColumn("note", "備考", width=28),
    )
    example = {
        "example": True,
        "teacher_id": "SAMPLE-T001",
        "name": "架空 一郎",
        "date": "2026-08-01",
        **{slot_key(code): (index + 1) % 3 for index, code in enumerate(slots)},
        "note": "この例示行は取込み対象外です",
    }
    workbook = _new_workbook("講師アンケート入力テンプレート")
    _write_data_sheet(
        workbook,
        TEACHER_AVAILABILITY_SHEET,
        columns,
        example,
        rows,
    )
    _write_master_reference_sheets(
        workbook,
        reference_students,
        reference_teachers,
        reference_subjects,
    )
    _add_teacher_reference_helpers(workbook)
    _write_instructions_sheet(
        workbook,
        (
            ("用途", "講師の出勤可能日時・希望日時を入力します。"),
            ("コマ値", "0 = 不可、1 = 可能、2 = 希望"),
            ("例示行", "「例示行」が「はい」の行は取込み対象外です。"),
            ("ID", "人物は名前ではなく講師IDで照合します。"),
            ("資格", "指導可能科目は講師マスターで管理します。"),
        ),
    )
    _atomic_save(workbook, path)


def write_group_lessons_template(
    path: Path,
    lessons: Sequence[Mapping[str, object]] = (),
    participants: Sequence[Mapping[str, object]] = (),
) -> None:
    """集団授業と受講者の2シートを持つテンプレートを出力する。"""
    lesson_columns = (
        _example_column(),
        _TemplateColumn("group_lesson_id", "集団授業ID", width=18, required=True),
        _TemplateColumn("grade", "学年", required=True),
        _TemplateColumn("subject_code", "科目コード", required=True),
        _TemplateColumn("course_name", "コース名", width=22),
        _TemplateColumn("date", "日付", date_value=True, required=True),
        _TemplateColumn("start_time", "開始時刻", time_value=True, required=True),
        _TemplateColumn("end_time", "終了時刻", time_value=True, required=True),
        _TemplateColumn("teacher_id", "担当講師ID", width=18),
        _TemplateColumn("room", "教室"),
        _TemplateColumn("note", "備考", width=28),
    )
    participant_columns = (
        _example_column(),
        _TemplateColumn("group_lesson_id", "集団授業ID", width=18, required=True),
        _TemplateColumn("student_id", "生徒ID", width=18, required=True),
    )
    workbook = _new_workbook("集団授業入力テンプレート")
    _write_data_sheet(
        workbook,
        GROUP_LESSON_SHEET,
        lesson_columns,
        {
            "example": True,
            "group_lesson_id": "SAMPLE-G001",
            "grade": "中学3年",
            "subject_code": "JH_MATH",
            "course_name": "架空 夏期数学",
            "date": "2026-08-01",
            "start_time": "17:20",
            "end_time": "18:20",
            "teacher_id": "SAMPLE-T001",
            "room": "架空教室",
            "note": "任意時刻の例です",
        },
        lessons,
    )
    _write_data_sheet(
        workbook,
        GROUP_PARTICIPANT_SHEET,
        participant_columns,
        {
            "example": True,
            "group_lesson_id": "SAMPLE-G001",
            "student_id": "SAMPLE-S001",
        },
        participants,
    )
    _write_instructions_sheet(
        workbook,
        (
            ("用途", "集団授業と受講者を固定予定として入力します。"),
            ("シート", "「集団授業」と「受講者」の両方を入力してください。"),
            ("時刻", "コマと完全一致しない開始・終了時刻も入力できます。"),
            ("例示行", "「例示行」が「はい」の行は取込み対象外です。"),
            ("ID", "受講者の集団授業IDは集団授業シートのIDを参照します。"),
        ),
    )
    _atomic_save(workbook, path)


def _new_workbook(title: str) -> Any:
    workbook = Workbook()
    workbook.remove(workbook.worksheets[0])
    workbook.properties.title = title
    workbook.properties.creator = "夏期講習時間割作成アプリ"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def _write_data_sheet(
    workbook: Any,
    sheet_name: str,
    columns: Sequence[_TemplateColumn],
    example: Mapping[str, object],
    rows: Sequence[Mapping[str, object]],
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.append([column.display_header for column in columns])
    worksheet.append([_excel_value(example.get(column.key)) for column in columns])
    for source_row in rows:
        row = dict(source_row)
        row.setdefault("example", False)
        worksheet.append(
            [
                _excel_value(
                    row.get(column.key, row.get(column.header)),
                )
                for column in columns
            ]
        )

    for column_number, column in enumerate(columns, start=1):
        letter = get_column_letter(column_number)
        header_cell = worksheet.cell(row=1, column=column_number)
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        if column.comment:
            header_cell.comment = Comment(column.comment, "夏期講習時間割作成アプリ")
        worksheet.column_dimensions[letter].width = column.width
        if column.availability:
            _add_availability_validation(worksheet, letter)
        if column.date_value:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = "yyyy-mm-dd"
        if column.time_value:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = "hh:mm"

    for cell in worksheet[2]:
        cell.fill = _EXAMPLE_FILL
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    last_column = get_column_letter(len(columns))
    worksheet.auto_filter.ref = f"A1:{last_column}{max(worksheet.max_row, 2)}"


def _write_instructions_sheet(
    workbook: Any,
    entries: Sequence[tuple[str, str]],
) -> None:
    worksheet = workbook.create_sheet(INSTRUCTIONS_SHEET)
    worksheet.sheet_view.showGridLines = False
    worksheet.append(["項目", "説明"])
    for label, description in entries:
        worksheet.append([label, description])
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 72
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_availability_validation(worksheet: Any, column_letter: str) -> None:
    validation = DataValidation(
        type="list",
        formula1='"0,1,2"',
        allow_blank=False,
    )
    validation.errorTitle = "入力値を確認してください"
    validation.error = "0（不可）、1（可能）、2（希望）から選択してください。"
    validation.promptTitle = "コマ希望"
    validation.prompt = "0 = 不可、1 = 可能、2 = 希望"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{_MAX_INPUT_ROW}")


def _atomic_save(workbook: Any, path: Path) -> None:
    destination = path.expanduser()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            prefix=f".{destination.stem}_",
            suffix=".xlsx.tmp",
            dir=destination.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
        temporary_path = None
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _example_column() -> _TemplateColumn:
    return _TemplateColumn(
        "example",
        "例示行",
        width=12,
        comment="「はい」の行は取込み対象外です。",
    )


def _slot_column(code: str) -> _TemplateColumn:
    return _TemplateColumn(
        slot_key(code),
        code,
        width=10,
        comment="0 = 不可、1 = 可能、2 = 希望",
        availability=True,
        required=True,
    )


def _write_master_reference_sheets(
    workbook: Any,
    students: Sequence[Mapping[str, object]],
    teachers: Sequence[Mapping[str, object]],
    subjects: Sequence[Mapping[str, object]],
) -> None:
    """アンケートと同じブックへ、編集時に照合できるマスターを同梱する。"""
    definitions = (
        (
            STUDENT_REFERENCE_SHEET,
            ("生徒ID", "氏名", "学年"),
            ("external_id", "name", "grade"),
            students,
        ),
        (
            TEACHER_REFERENCE_SHEET,
            ("講師ID", "氏名"),
            ("external_id", "name"),
            teachers,
        ),
        (
            SUBJECT_REFERENCE_SHEET,
            ("科目コード", "表示名"),
            ("code", "display_name"),
            subjects,
        ),
    )
    for sheet_name, headers, keys, rows in definitions:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append([row.get(key) for key in keys])
        for cell in worksheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGNMENT
        for column_number in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(column_number)].width = 20
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{max(worksheet.max_row, 2)}"


def _add_student_reference_helpers(workbook: Any) -> None:
    worksheet = workbook[STUDENT_AVAILABILITY_SHEET]
    first_input_row = max(worksheet.max_row + 1, 3)
    student_id = _template_column_letter(worksheet, "生徒ID")
    student_name = _template_column_letter(worksheet, "生徒名")
    subject_code = _template_column_letter(worksheet, "科目コード")
    subject_name = _template_column_letter(worksheet, "科目名（確認）")
    _add_reference_validation(
        worksheet,
        student_id,
        f"=INDIRECT(\"'{STUDENT_REFERENCE_SHEET}'!$A$2:$A$10000\")",
    )
    _add_reference_validation(
        worksheet,
        subject_code,
        f"=INDIRECT(\"'{SUBJECT_REFERENCE_SHEET}'!$A$2:$A$10000\")",
    )
    for header in ("第1希望講師ID", "第2希望講師ID", "第3希望講師ID"):
        _add_reference_validation(
            worksheet,
            _template_column_letter(worksheet, header),
            f"=INDIRECT(\"'{TEACHER_REFERENCE_SHEET}'!$A$2:$A$10000\")",
        )
    for row_number in range(3, 1001):
        if row_number >= first_input_row:
            worksheet[f"{student_name}{row_number}"] = (
                f'=IF({student_id}{row_number}="","",IFERROR(INDEX('
                f"'{STUDENT_REFERENCE_SHEET}'!$B$2:$B$10000,"
                f"MATCH({student_id}{row_number},'{STUDENT_REFERENCE_SHEET}'!"
                '$A$2:$A$10000,0)),"ID不明"))'
            )
        worksheet[f"{subject_name}{row_number}"] = (
            f'=IF({subject_code}{row_number}="","",IFERROR(INDEX('
            f"'{SUBJECT_REFERENCE_SHEET}'!$B$2:$B$10000,"
            f"MATCH({subject_code}{row_number},'{SUBJECT_REFERENCE_SHEET}'!"
            '$A$2:$A$10000,0)),"コード不明"))'
        )


def _add_teacher_reference_helpers(workbook: Any) -> None:
    worksheet = workbook[TEACHER_AVAILABILITY_SHEET]
    first_input_row = max(worksheet.max_row + 1, 3)
    teacher_id = _template_column_letter(worksheet, "講師ID")
    teacher_name = _template_column_letter(worksheet, "講師名")
    _add_reference_validation(
        worksheet,
        teacher_id,
        f"=INDIRECT(\"'{TEACHER_REFERENCE_SHEET}'!$A$2:$A$10000\")",
    )
    for row_number in range(3, 1001):
        if row_number >= first_input_row:
            worksheet[f"{teacher_name}{row_number}"] = (
                f'=IF({teacher_id}{row_number}="","",IFERROR(INDEX('
                f"'{TEACHER_REFERENCE_SHEET}'!$B$2:$B$10000,"
                f"MATCH({teacher_id}{row_number},'{TEACHER_REFERENCE_SHEET}'!"
                '$A$2:$A$10000,0)),"ID不明"))'
            )


def _add_reference_validation(worksheet: Any, column_letter: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.errorTitle = "マスターを確認してください"
    validation.error = "同じブックのマスター参照シートにあるIDを選択してください。"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{_MAX_INPUT_ROW}")


def _template_column_letter(worksheet: Any, header: str) -> str:
    for column_number in range(1, worksheet.max_column + 1):
        actual = str(worksheet.cell(row=1, column=column_number).value or "")
        if actual == header or actual == f"{header}（必須）":
            return get_column_letter(column_number)
    raise AssertionError(f"{worksheet.title}に列「{header}」がありません。")


def _validated_slots(slot_codes: Sequence[str]) -> tuple[str, ...]:
    slots = tuple(code.strip() for code in slot_codes)
    if not slots or any(not code for code in slots):
        raise ValueError("コマコードを1件以上指定してください。")
    if len(slots) != len(set(slots)):
        raise ValueError("コマコードが重複しています。")
    return slots


def _excel_value(value: object) -> object:
    if isinstance(value, bool):
        return "はい" if value else "いいえ"
    return value
