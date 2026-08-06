"""Phase 1～6の主要Application Serviceを一つの匿名業務フローで確認する。"""

from __future__ import annotations

from dataclasses import replace
from datetime import date, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PySide6.QtGui import QGuiApplication

from summer_scheduler.application.availability_import_service import (
    AvailabilityImportService,
)
from summer_scheduler.application.group_lesson_service import GroupLessonService
from summer_scheduler.application.optimization_run_service import (
    OptimizationRunService,
)
from summer_scheduler.application.output_service import OutputService
from summer_scheduler.application.phase5_dto import ScheduleBoardDto
from summer_scheduler.application.project_service import ProjectService
from summer_scheduler.application.project_validation_service import (
    ProjectValidationService,
)
from summer_scheduler.application.sample_project_service import SampleProjectService
from summer_scheduler.application.schedule_edit_service import ScheduleEditService
from summer_scheduler.infrastructure.db import create_database, upgrade_database
from summer_scheduler.infrastructure.importing import (
    GROUP_LESSON_SHEET,
    GROUP_PARTICIPANT_SHEET,
    STUDENT_AVAILABILITY_SHEET,
    TEACHER_AVAILABILITY_SHEET,
)
from summer_scheduler.infrastructure.importing.templates import (
    STUDENT_REFERENCE_SHEET,
    SUBJECT_REFERENCE_SHEET,
    TEACHER_REFERENCE_SHEET,
)
from summer_scheduler.optimization.solver import solve_optimization
from summer_scheduler.shared.settings import load_settings


def test_anonymous_project_survives_import_optimize_edit_output_and_restart(
    tmp_path: Path,
    qt_gui_app: QGuiApplication,
) -> None:
    """必須フローをサービス境界で通し、再起動後の永続化まで確認する。"""
    registry_path = tmp_path / "アプリ管理" / "registry.db"
    backup_directory = tmp_path / "安全なバックアップ"
    project_path = tmp_path / "日本語の長い保存先" / "Phase 7 品質確認.jukuschedule"
    output_directory = tmp_path / "日本語 出力"
    settings = load_settings()
    optimization_settings = replace(
        settings.optimization,
        fast_time_limit_seconds=5.0,
        standard_time_limit_seconds=5.0,
        high_quality_time_limit_seconds=5.0,
    )

    registry = create_database(registry_path)
    upgrade_database(registry.engine)
    projects = ProjectService(registry, backup_directory)
    try:
        # 新規プロジェクトと、コマ・開校日・マスター・LessonRequestの匿名初期値。
        summary = SampleProjectService(projects).create_anonymous_sample(project_path)
        assert summary.path == project_path.resolve()

        # 生徒・講師アンケートは、テンプレート出力→編集→preview→applyを通す。
        availability = AvailabilityImportService(projects)
        student_book = tmp_path / "架空 生徒アンケート.xlsx"
        teacher_book = tmp_path / "架空 講師アンケート.xlsx"
        availability.export_student_template(student_book)
        availability.export_teacher_template(teacher_book)
        student_template = load_workbook(student_book, data_only=False)
        try:
            assert student_template[STUDENT_REFERENCE_SHEET]["A2"].value == "S-001"
            assert student_template[TEACHER_REFERENCE_SHEET]["A2"].value == "T-001"
            assert "JH_ENG" in {
                cell.value for cell in student_template[SUBJECT_REFERENCE_SHEET]["A"]
            }
            assert student_template.calculation.forceFullCalc is True
        finally:
            student_template.close()
        _append_student_availability(student_book)
        _append_teacher_availability(teacher_book)
        student_preview = availability.prepare_import(
            "student",
            student_book,
            sheet_name=STUDENT_AVAILABILITY_SHEET,
        )
        teacher_preview = availability.prepare_import(
            "teacher",
            teacher_book,
            sheet_name=TEACHER_AVAILABILITY_SHEET,
        )
        assert not student_preview.has_errors
        assert not teacher_preview.has_errors
        availability.apply_import(student_preview)
        availability.apply_import(teacher_preview)

        # 集団授業も2シートのテンプレートをpreviewしてから反映する。
        groups = GroupLessonService(projects)
        group_book = tmp_path / "架空 集団授業.xlsx"
        groups.export_template(group_book)
        _append_group_lesson(group_book)
        group_preview = groups.prepare_group_import(group_book)
        assert not group_preview.has_errors
        groups.apply_group_import(group_preview)

        issues = ProjectValidationService(projects).run_validation()
        assert not [issue for issue in issues if issue.severity == "error"]

        # 自動最適化と未配置一覧の構築。
        run_service = OptimizationRunService(projects, optimization_settings)
        prepared = run_service.prepare("fast", log_directory=tmp_path / "最適化ログ")
        first_result = solve_optimization(prepared.input)
        finalized = run_service.finalize(prepared, first_result)
        assert finalized.assignment_count > 0

        # 手動編集、ロック、checkpoint後の再最適化でロック配置を保持する。
        editor = ScheduleEditService(projects, optimization_settings)
        board = editor.load_board()
        assert board.cards
        edited = board.cards[0]
        editor.edit_assignment(
            lesson_request_id=edited.lesson_request_id,
            session_index=edited.session_index,
            day=edited.day,
            time_slot_id=edited.time_slot_id,
            teacher_id=edited.teacher_id,
            is_locked=False,
            note="Phase 7 匿名回帰確認",
            reason="匿名回帰確認の手動編集",
        )
        editor.set_lock(
            lesson_request_id=edited.lesson_request_id,
            session_index=edited.session_index,
            is_locked=True,
            reason="再最適化で保持する匿名ロック",
        )
        locked_before = _card_placement(
            editor.load_board(),
            edited.lesson_request_id,
            edited.session_index,
        )
        checkpoint = editor.create_checkpoint_backup()
        assert checkpoint.path.is_file()

        prepared_again = run_service.prepare(
            "fast",
            log_directory=tmp_path / "最適化ログ",
        )
        second_result = solve_optimization(prepared_again.input)
        run_service.finalize(prepared_again, second_result)
        locked_after = _card_placement(
            ScheduleEditService(projects, optimization_settings).load_board(),
            edited.lesson_request_id,
            edited.session_index,
        )
        assert locked_after == locked_before
        assert locked_after[-1] is True

        # Excel/PDFを実体生成し、明示保存点を作る。
        assert qt_gui_app is QGuiApplication.instance()
        output = OutputService(
            projects,
            optimization_settings,
            output_defaults=settings.output,
        )
        workspace = output.load_workspace()
        assert workspace.assignment_count == finalized.assignment_count
        excel_result = output.export_excel(
            "overall",
            output_directory / "Phase 7 時間割.xlsx",
        )
        pdf_result = output.export_pdf(
            "overall",
            output_directory / "Phase 7 時間割.pdf",
        )
        assert excel_result.path.stat().st_size > 0
        assert pdf_result.path.read_bytes().startswith(b"%PDF-")
        saved = projects.backup(tmp_path / "手動保存点.jukuschedule")
        assert saved.is_file()
    finally:
        projects.close_project()
        registry.dispose()

    # 終了・再起動相当としてregistry/ProjectServiceを作り直して再読込みする。
    restarted_registry = create_database(registry_path)
    upgrade_database(restarted_registry.engine)
    restarted_projects = ProjectService(restarted_registry, backup_directory)
    try:
        reopened = restarted_projects.open_project(project_path)
        assert reopened.title == "匿名サンプル 2026夏期講習"
        restarted_board = ScheduleEditService(
            restarted_projects,
            optimization_settings,
        ).load_board()
        assert restarted_board.cards
        assert (
            _card_placement(
                restarted_board,
                edited.lesson_request_id,
                edited.session_index,
            )
            == locked_after
        )
    finally:
        restarted_projects.close_project()
        restarted_registry.dispose()


def _append_student_availability(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        _write_template_row(
            workbook[STUDENT_AVAILABILITY_SHEET],
            {
                "例示行": False,
                "生徒ID": "S-001",
                "生徒名": "架空 青空",
                "科目コード": "JH_ENG",
                "日付": date(2026, 8, 3),
                "Y": 1,
                "Z": 1,
                "A": 1,
                "B": 1,
                "C": 1,
                "第1希望講師ID": "T-001",
                "備考": "Phase 7 匿名入力",
            },
        )
        workbook.save(path)
    finally:
        workbook.close()


def _append_teacher_availability(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        _write_template_row(
            workbook[TEACHER_AVAILABILITY_SHEET],
            {
                "例示行": False,
                "講師ID": "T-001",
                "講師名": "架空 講師あおい",
                "日付": date(2026, 8, 3),
                "Y": 1,
                "Z": 1,
                "A": 1,
                "B": 1,
                "C": 1,
                "備考": "Phase 7 匿名入力",
            },
        )
        workbook.save(path)
    finally:
        workbook.close()


def _write_template_row(
    worksheet: Worksheet,
    values: dict[str, bool | float | str | date | None],
) -> None:
    target = worksheet
    row_number = next(
        number
        for number in range(3, target.max_row + 2)
        if target.cell(row=number, column=1).value is None
    )
    for column_number, cell in enumerate(target[1], start=1):
        header = str(cell.value).removesuffix("（必須）")
        if header in values:
            target.cell(row=row_number, column=column_number).value = values[header]


def _append_group_lesson(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        workbook[GROUP_LESSON_SHEET].append(
            (
                False,
                "GROUP-PHASE7",
                "中1",
                "JH_ENG",
                "架空 品質確認講座",
                date(2026, 8, 7),
                time(10, 0),
                time(11, 0),
                "T-001",
                "架空教室B",
                "Phase 7 匿名入力",
            )
        )
        workbook[GROUP_PARTICIPANT_SHEET].append((False, "GROUP-PHASE7", "S-001"))
        workbook.save(path)
    finally:
        workbook.close()


def _card_placement(
    board: ScheduleBoardDto,
    lesson_request_id: int,
    session_index: int,
) -> tuple[date, int, int, bool]:
    for card in board.cards:
        if card.lesson_request_id == lesson_request_id and card.session_index == session_index:
            return (
                card.day,
                card.time_slot_id,
                card.teacher_id,
                card.is_locked,
            )
    raise AssertionError("対象の配置が見つかりません")
