"""年度をまたいで利用する生徒・講師基本情報Excelの入出力。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import os
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Final

import xlsxwriter
from openpyxl import load_workbook

from summer_scheduler.domain.defaults import DEFAULT_SUBJECTS
from summer_scheduler.domain.grades import (
    EXCEL_GRADE_OPTIONS,
    INTERNAL_GRADE_OPTIONS,
    grade_from_excel,
    grade_to_excel,
)
from summer_scheduler.domain.identifiers import PersonIdPrefix, next_person_external_id

SHARED_ROSTER_FILENAME: Final = "生徒・講師_基本情報.xlsx"

_STUDENT_HEADERS: Final = (
    "在籍",
    "生徒ID（自動・入力不要）",
    "姓（必須）",
    "名",
    "氏名（確認）",
    "学年（必須）",
    "標準最大連続コマ数（デフォルトは2）",
    "空きコマ許可（デフォルトはなし）",
    "備考",
)
_TEACHER_HEADERS: Final = (
    "在籍",
    "講師ID（自動・入力不要）",
    "姓（必須）",
    "名",
    "氏名（確認）",
    "空きコマ許可（デフォルトはなし）",
    "備考",
)
_SUBJECT_HEADERS: Final = (
    "科目コード（必須）",
    "表示名（必須）",
    "学校段階（必須）",
    "並び順（必須）",
    "有効",
)
_QUALIFICATION_HEADERS: Final = (
    "講師名から選択",
    "講師ID（自動・入力不要）",
    "科目名から選択",
    "科目コード（自動・入力不要）",
    "指導可能（デフォルトははい）",
    "備考",
)
_REGULAR_HEADERS: Final = (
    "生徒名から選択",
    "生徒ID（自動・入力不要）",
    "科目名から選択",
    "科目コード（自動・入力不要）",
    "通常担当講師名から選択",
    "通常担当講師ID（自動・入力不要）",
    "担当講師優先度（デフォルトは3）",
    "1対1必須（デフォルトはいいえ）",
    "備考",
)
_LEGACY_QUALIFICATION_WIDTH: Final = 8
_LEGACY_REGULAR_WIDTH: Final = 12
_SHEETS: Final = ("生徒", "講師", "科目", "講師対応科目", "通常授業")
_ID_HELPER_SHEET: Final = "_入力補助"
_MAX_ROW: Final = 10_000
_FORMULA_TEMPLATE_MAX_ROW: Final = 1_000


class SharedRosterError(ValueError):
    """共通名簿の形式または参照関係に問題がある。"""


@dataclass(frozen=True, slots=True)
class SharedStudent:
    external_id: str
    surname: str
    given_name: str
    grade: str
    max_consecutive_slots: int = 2
    allow_gap: bool = False
    active: bool = True
    note: str = ""

    @property
    def name(self) -> str:
        return _full_name(self.surname, self.given_name)


@dataclass(frozen=True, slots=True)
class SharedTeacher:
    external_id: str
    surname: str
    given_name: str
    allow_gap: bool = False
    active: bool = True
    note: str = ""

    @property
    def name(self) -> str:
        return _full_name(self.surname, self.given_name)


@dataclass(frozen=True, slots=True)
class SharedSubject:
    code: str
    display_name: str
    school_level: str
    sort_order: int
    active: bool = True


@dataclass(frozen=True, slots=True)
class SharedQualification:
    teacher_external_id: str
    subject_code: str
    can_teach: bool = True
    note: str = ""


@dataclass(frozen=True, slots=True)
class SharedRegularLesson:
    student_external_id: str
    subject_code: str
    regular_teacher_external_id: str = ""
    regular_teacher_priority: int = 3
    one_to_one_required: bool = False
    note: str = ""


@dataclass(frozen=True, slots=True)
class SharedRosterData:
    students: tuple[SharedStudent, ...]
    teachers: tuple[SharedTeacher, ...]
    subjects: tuple[SharedSubject, ...]
    qualifications: tuple[SharedQualification, ...] = ()
    regular_lessons: tuple[SharedRegularLesson, ...] = ()


def empty_shared_roster() -> SharedRosterData:
    """既定科目だけを含む最初の共通名簿を返す。"""
    return SharedRosterData(
        students=(),
        teachers=(),
        subjects=tuple(
            SharedSubject(
                item.code,
                item.display_name,
                item.school_level,
                item.sort_order,
                True,
            )
            for item in DEFAULT_SUBJECTS
        ),
    )


def write_shared_roster(
    path: Path,
    data: SharedRosterData,
    *,
    reserved_student_ids: Iterable[str] = (),
    reserved_teacher_ids: Iterable[str] = (),
) -> None:
    """共通名簿を入力規則・参照表示・退籍行の灰色表示付きで保存する。"""
    destination = path.expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    workbook: Any | None = None
    closed = False
    try:
        with NamedTemporaryFile(
            prefix=f".{destination.stem}_",
            suffix=".xlsx.tmp",
            dir=destination.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        workbook = xlsxwriter.Workbook(str(temporary))
        workbook.set_properties(
            {
                "title": "生徒・講師 基本情報",
                "subject": "講習に依存しない在籍者・通常授業情報",
                "author": "夏期講習時間割作成アプリ",
            }
        )
        workbook.set_calc_mode("auto")
        formats = _xlsx_formats(workbook)

        student_rows = sorted(data.students, key=lambda item: (not item.active, item.external_id))
        student_sheet = workbook.add_worksheet("生徒")
        _setup_sheet(
            student_sheet,
            _STUDENT_HEADERS,
            (20, 18, 15, 15, 24, 12, 25, 24, 32),
            formats,
        )
        _mark_auto_id_header(student_sheet, "B1", "生徒", "S-0001")
        for excel_row, student in enumerate(student_rows, start=2):
            row = excel_row - 1
            student_sheet.insert_checkbox(row, 0, student.active, formats["checkbox"])
            student_sheet.write(row, 1, student.external_id, formats["normal"])
            student_sheet.write(row, 2, student.surname, formats["required"])
            student_sheet.write(row, 3, student.given_name, formats["normal"])
            student_sheet.write_formula(
                row,
                4,
                f'=IF(COUNTA(C{excel_row}:D{excel_row})=0,"",TRIM(C{excel_row}&" "&D{excel_row}))',
                formats["normal"],
                student.name,
            )
            student_sheet.write(row, 5, grade_to_excel(student.grade), formats["required"])
            student_sheet.write(row, 6, student.max_consecutive_slots, formats["normal"])
            student_sheet.write(row, 7, _gap_label(student.allow_gap), formats["normal"])
            student_sheet.write(row, 8, student.note, formats["normal"])
        _prepare_student_input_rows(student_sheet, len(student_rows) + 2, formats)
        _add_list(student_sheet, "F2:F10000", list(EXCEL_GRADE_OPTIONS))
        _add_whole(student_sheet, "G2:G10000", 1, 3, allow_blank=True)
        _add_list(student_sheet, "H2:H10000", ["あり", "なし"], allow_blank=True)
        _grey_inactive(student_sheet, len(_STUDENT_HEADERS), formats["inactive"])
        _finish_rows(student_sheet, len(student_rows), len(_STUDENT_HEADERS))

        teacher_rows = sorted(data.teachers, key=lambda item: (not item.active, item.external_id))
        teacher_sheet = workbook.add_worksheet("講師")
        _setup_sheet(
            teacher_sheet,
            _TEACHER_HEADERS,
            (20, 18, 15, 15, 24, 24, 32),
            formats,
        )
        _mark_auto_id_header(teacher_sheet, "B1", "講師", "T-0001")
        for excel_row, teacher in enumerate(teacher_rows, start=2):
            row = excel_row - 1
            teacher_sheet.insert_checkbox(row, 0, teacher.active, formats["checkbox"])
            teacher_sheet.write(row, 1, teacher.external_id, formats["normal"])
            teacher_sheet.write(row, 2, teacher.surname, formats["required"])
            teacher_sheet.write(row, 3, teacher.given_name, formats["normal"])
            teacher_sheet.write_formula(
                row,
                4,
                f'=IF(COUNTA(C{excel_row}:D{excel_row})=0,"",TRIM(C{excel_row}&" "&D{excel_row}))',
                formats["normal"],
                teacher.name,
            )
            teacher_sheet.write(row, 5, _gap_label(teacher.allow_gap), formats["normal"])
            teacher_sheet.write(row, 6, teacher.note, formats["normal"])
        _prepare_teacher_input_rows(teacher_sheet, len(teacher_rows) + 2, formats)
        _add_list(teacher_sheet, "F2:F10000", ["あり", "なし"], allow_blank=True)
        _grey_inactive(teacher_sheet, len(_TEACHER_HEADERS), formats["inactive"])
        _finish_rows(teacher_sheet, len(teacher_rows), len(_TEACHER_HEADERS))

        subject_rows = sorted(data.subjects, key=lambda item: (item.sort_order, item.code))
        subject_sheet = workbook.add_worksheet("科目")
        _setup_sheet(subject_sheet, _SUBJECT_HEADERS, (20, 28, 15, 13, 12), formats)
        for row, subject in enumerate(subject_rows, start=1):
            _write_plain_row(
                subject_sheet,
                row,
                (
                    subject.code,
                    subject.display_name,
                    _school_level_label(subject.school_level),
                    subject.sort_order,
                    _yes_no(subject.active),
                ),
                formats["normal"],
            )
        _add_list(subject_sheet, "C2:C10000", ["小学校", "中学校", "高校"])
        _add_whole(subject_sheet, "D2:D10000", 1, 9999)
        _add_list(subject_sheet, "E2:E10000", ["はい", "いいえ"], allow_blank=True)
        _finish_rows(subject_sheet, len(subject_rows), len(_SUBJECT_HEADERS))

        qualification_sheet = workbook.add_worksheet("講師対応科目")
        _setup_sheet(
            qualification_sheet,
            _QUALIFICATION_HEADERS,
            (28, 22, 32, 24, 24, 32),
            formats,
            helper_columns={2, 4},
        )
        teacher_name_by_id = {row.external_id: row.name for row in data.teachers}
        subject_name_by_code = {row.code: row.display_name for row in data.subjects}
        for excel_row, qualification in enumerate(data.qualifications, start=2):
            row = excel_row - 1
            qualification_sheet.write_formula(
                row,
                0,
                _linked_display_formula(excel_row, "B", "講師", "B", "E"),
                formats["normal"],
                teacher_name_by_id.get(qualification.teacher_external_id, ""),
            )
            qualification_sheet.write(row, 1, qualification.teacher_external_id, formats["normal"])
            qualification_sheet.write_formula(
                row,
                2,
                _linked_display_formula(excel_row, "D", "科目", "A", "B"),
                formats["normal"],
                subject_name_by_code.get(qualification.subject_code, ""),
            )
            qualification_sheet.write(row, 3, qualification.subject_code, formats["normal"])
            qualification_sheet.write(row, 4, _yes_no(qualification.can_teach), formats["normal"])
            qualification_sheet.write(row, 5, qualification.note, formats["normal"])
        _prepare_qualification_input_rows(
            qualification_sheet, len(data.qualifications) + 2, formats
        )
        _add_list(
            qualification_sheet,
            "A2:A10000",
            "=INDIRECT(\"'講師'!$E$2:$E$10000\")",
            allow_blank=True,
        )
        _add_list(
            qualification_sheet,
            "C2:C10000",
            "=INDIRECT(\"'科目'!$B$2:$B$10000\")",
            allow_blank=True,
        )
        _add_list(qualification_sheet, "E2:E10000", ["はい", "いいえ"], allow_blank=True)
        _finish_rows(qualification_sheet, len(data.qualifications), len(_QUALIFICATION_HEADERS))

        regular_sheet = workbook.add_worksheet("通常授業")
        _setup_sheet(
            regular_sheet,
            _REGULAR_HEADERS,
            (28, 22, 32, 24, 28, 24, 24, 26, 32),
            formats,
            helper_columns={2, 4, 6},
        )
        student_name_by_id = {row.external_id: row.name for row in data.students}
        for excel_row, lesson in enumerate(data.regular_lessons, start=2):
            row = excel_row - 1
            regular_sheet.write_formula(
                row,
                0,
                _linked_display_formula(excel_row, "B", "生徒", "B", "E"),
                formats["normal"],
                student_name_by_id.get(lesson.student_external_id, ""),
            )
            regular_sheet.write(row, 1, lesson.student_external_id, formats["normal"])
            regular_sheet.write_formula(
                row,
                2,
                _linked_display_formula(excel_row, "D", "科目", "A", "B"),
                formats["normal"],
                subject_name_by_code.get(lesson.subject_code, ""),
            )
            regular_sheet.write(row, 3, lesson.subject_code, formats["normal"])
            regular_sheet.write_formula(
                row,
                4,
                _linked_display_formula(excel_row, "F", "講師", "B", "E"),
                formats["normal"],
                teacher_name_by_id.get(lesson.regular_teacher_external_id, ""),
            )
            regular_sheet.write(
                row,
                5,
                lesson.regular_teacher_external_id,
                formats["normal"],
            )
            regular_sheet.write(row, 6, lesson.regular_teacher_priority, formats["normal"])
            regular_sheet.write(row, 7, _yes_no(lesson.one_to_one_required), formats["normal"])
            regular_sheet.write(row, 8, lesson.note, formats["normal"])
        _prepare_regular_lesson_input_rows(regular_sheet, len(data.regular_lessons) + 2, formats)
        _add_list(
            regular_sheet,
            "A2:A10000",
            "=INDIRECT(\"'生徒'!$E$2:$E$10000\")",
            allow_blank=True,
        )
        _add_list(
            regular_sheet,
            "C2:C10000",
            "=INDIRECT(\"'科目'!$B$2:$B$10000\")",
            allow_blank=True,
        )
        _add_list(
            regular_sheet,
            "E2:E10000",
            "=INDIRECT(\"'講師'!$E$2:$E$10000\")",
            allow_blank=True,
        )
        _add_whole(regular_sheet, "G2:G10000", 1, 5, allow_blank=True)
        _add_list(regular_sheet, "H2:H10000", ["はい", "いいえ"], allow_blank=True)
        _finish_rows(regular_sheet, len(data.regular_lessons), len(_REGULAR_HEADERS))

        _create_id_helper_sheet(
            workbook,
            data,
            reserved_student_ids=reserved_student_ids,
            reserved_teacher_ids=reserved_teacher_ids,
        )
        workbook.close()
        closed = True
        os.replace(temporary, destination)
        temporary = None
    finally:
        if workbook is not None and not closed:
            workbook.close()
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def read_shared_roster(
    path: Path,
    *,
    reserved_student_ids: Iterable[str] = (),
    reserved_teacher_ids: Iterable[str] = (),
) -> SharedRosterData:
    """共通名簿を検証し、名前選択列も解決して正規化したデータを返す。"""
    source = path.expanduser().resolve()
    try:
        workbook = load_workbook(source, data_only=False, read_only=False)
    except (OSError, ValueError) as exc:
        raise SharedRosterError("共通名簿を読み込めませんでした。") from exc
    try:
        missing = [name for name in _SHEETS if name not in workbook.sheetnames]
        if missing:
            raise SharedRosterError(f"必要なシートがありません: {', '.join(missing)}")
        errors: list[str] = []
        students = _read_students(workbook["生徒"], errors, reserved_student_ids)
        teachers = _read_teachers(workbook["講師"], errors, reserved_teacher_ids)
        subjects = _read_subjects(workbook["科目"], errors)
        qualifications = _read_qualifications(
            workbook["講師対応科目"], students, teachers, subjects, errors
        )
        regular_lessons = _read_regular_lessons(
            workbook["通常授業"], students, teachers, subjects, errors
        )
        if errors:
            detail = "\n".join(f"・{message}" for message in errors[:20])
            suffix = "\n（ほかにもエラーがあります）" if len(errors) > 20 else ""
            raise SharedRosterError(f"共通名簿に入力エラーがあります。\n{detail}{suffix}")
        return SharedRosterData(
            tuple(students),
            tuple(teachers),
            tuple(subjects),
            tuple(qualifications),
            tuple(regular_lessons),
        )
    finally:
        workbook.close()


def _read_students(
    sheet: Any,
    errors: list[str],
    reserved_external_ids: Iterable[str] = (),
) -> list[SharedStudent]:
    result: list[SharedStudent] = []
    used: set[str] = set()
    rows = _rows(sheet, len(_STUDENT_HEADERS))
    status_first = _status_is_first_column(sheet)
    active_index, id_index, surname_index, given_name_index = (
        (0, 1, 2, 3) if status_first else (7, 0, 1, 2)
    )
    grade_index, maximum_index, gap_index, note_index = (
        (5, 6, 7, 8) if status_first else (4, 5, 6, 8)
    )
    reserved = {value.strip() for value in reserved_external_ids if value.strip()}
    reserved.update(
        _text(values[id_index]) for _row_number, values in rows if _text(values[id_index])
    )
    for row_number, values in rows:
        if status_first and not any(
            _text(values[index])
            for index in (id_index, surname_index, given_name_index, grade_index, note_index)
        ):
            continue
        if _empty(values):
            continue
        external_id = _text(values[id_index])
        if not external_id:
            external_id = next_person_external_id(reserved, prefix="S")
            reserved.add(external_id)
        surname = _text(values[surname_index])
        given_name = _text(values[given_name_index])
        grade = grade_from_excel(_text(values[grade_index]))
        if external_id in used:
            errors.append(f"生徒 {row_number}行: 生徒ID「{external_id}」が重複しています")
            continue
        used.add(external_id)
        if not surname:
            errors.append(f"生徒 {row_number}行: 姓は必須です")
        if grade not in INTERNAL_GRADE_OPTIONS:
            errors.append(f"生徒 {row_number}行: 学年はS1～S6、J1～J3、H1～H3から選択してください")
        maximum = _integer(values[maximum_index], 2)
        if not 1 <= maximum <= 3:
            errors.append(f"生徒 {row_number}行: 最大連続コマ数は1～3です")
        result.append(
            SharedStudent(
                external_id,
                surname,
                given_name,
                grade,
                maximum,
                _boolean(values[gap_index], False),
                _active(values[active_index]),
                _text(values[note_index]),
            )
        )
    return result


def _read_teachers(
    sheet: Any,
    errors: list[str],
    reserved_external_ids: Iterable[str] = (),
) -> list[SharedTeacher]:
    result: list[SharedTeacher] = []
    used: set[str] = set()
    rows = _rows(sheet, len(_TEACHER_HEADERS))
    status_first = _status_is_first_column(sheet)
    active_index, id_index, surname_index, given_name_index, gap_index, note_index = (
        (0, 1, 2, 3, 5, 6) if status_first else (5, 0, 1, 2, 4, 6)
    )
    reserved = {value.strip() for value in reserved_external_ids if value.strip()}
    reserved.update(
        _text(values[id_index]) for _row_number, values in rows if _text(values[id_index])
    )
    for row_number, values in rows:
        if status_first and not any(
            _text(values[index])
            for index in (id_index, surname_index, given_name_index, note_index)
        ):
            continue
        if _empty(values):
            continue
        external_id = _text(values[id_index])
        if not external_id:
            external_id = next_person_external_id(reserved, prefix="T")
            reserved.add(external_id)
        surname = _text(values[surname_index])
        given_name = _text(values[given_name_index])
        if external_id in used:
            errors.append(f"講師 {row_number}行: 講師ID「{external_id}」が重複しています")
            continue
        used.add(external_id)
        if not surname:
            errors.append(f"講師 {row_number}行: 姓は必須です")
        result.append(
            SharedTeacher(
                external_id,
                surname,
                given_name,
                _boolean(values[gap_index], False),
                _active(values[active_index]),
                _text(values[note_index]),
            )
        )
    return result


def _read_subjects(sheet: Any, errors: list[str]) -> list[SharedSubject]:
    result: list[SharedSubject] = []
    used: set[str] = set()
    for row_number, values in _rows(sheet, len(_SUBJECT_HEADERS)):
        if _empty(values):
            continue
        code, name, level = _text(values[0]), _text(values[1]), _text(values[2])
        level_code = _school_level_code(level)
        if not code or not name or not level_code:
            errors.append(f"科目 {row_number}行: コード・表示名・学校段階は必須です")
            continue
        if code in used:
            errors.append(f"科目 {row_number}行: 科目コード「{code}」が重複しています")
            continue
        used.add(code)
        order = _integer(values[3], len(result) + 1)
        result.append(SharedSubject(code, name, level_code, order, _boolean(values[4], True)))
    return result


def _read_qualifications(
    sheet: Any,
    students: list[SharedStudent],
    teachers: list[SharedTeacher],
    subjects: list[SharedSubject],
    errors: list[str],
) -> list[SharedQualification]:
    del students
    teacher_by_name = _unique_name_map(teachers)
    teacher_ids = {row.external_id for row in teachers}
    subject_by_name = _unique_subject_name_map(subjects)
    subject_codes = {row.code for row in subjects}
    result: list[SharedQualification] = []
    used: set[tuple[str, str]] = set()
    new_layout = _text(sheet.cell(1, 1).value).startswith("講師名から選択")
    width = len(_QUALIFICATION_HEADERS) if new_layout else _LEGACY_QUALIFICATION_WIDTH
    for row_number, values in _rows(sheet, width):
        if _empty(values):
            continue
        if new_layout:
            selected_teacher_name = _text(values[0])
            selected_subject_name = _text(values[2])
            teacher_id = (
                teacher_by_name.get(selected_teacher_name, "")
                if selected_teacher_name
                else _text(values[1])
            )
            subject_code = (
                subject_by_name.get(selected_subject_name, "")
                if selected_subject_name
                else _text(values[3])
            )
            can_teach_value, note_value = values[4], values[5]
        else:
            teacher_id = _text(values[0]) or teacher_by_name.get(_text(values[1]), "")
            subject_code = _text(values[3]) or subject_by_name.get(_text(values[4]), "")
            can_teach_value, note_value = values[6], values[7]
        if teacher_id not in teacher_ids:
            errors.append(f"講師対応科目 {row_number}行: 講師が基本情報にありません")
        if subject_code not in subject_codes:
            errors.append(f"講師対応科目 {row_number}行: 科目が科目シートにありません")
        key = (teacher_id, subject_code)
        if key in used:
            errors.append(f"講師対応科目 {row_number}行: 講師と科目の組が重複しています")
            continue
        used.add(key)
        result.append(
            SharedQualification(
                teacher_id,
                subject_code,
                _boolean(can_teach_value, True),
                _text(note_value),
            )
        )
    return result


def _read_regular_lessons(
    sheet: Any,
    students: list[SharedStudent],
    teachers: list[SharedTeacher],
    subjects: list[SharedSubject],
    errors: list[str],
) -> list[SharedRegularLesson]:
    student_by_name = _unique_name_map(students)
    teacher_by_name = _unique_name_map(teachers)
    subject_by_name = _unique_subject_name_map(subjects)
    student_ids = {row.external_id for row in students}
    teacher_ids = {row.external_id for row in teachers}
    subject_codes = {row.code for row in subjects}
    result: list[SharedRegularLesson] = []
    used: set[tuple[str, str]] = set()
    new_layout = _text(sheet.cell(1, 1).value).startswith("生徒名から選択")
    width = len(_REGULAR_HEADERS) if new_layout else _LEGACY_REGULAR_WIDTH
    for row_number, values in _rows(sheet, width):
        if _empty(values):
            continue
        if new_layout:
            selected_student_name = _text(values[0])
            selected_subject_name = _text(values[2])
            selected_teacher_name = _text(values[4])
            student_id = (
                student_by_name.get(selected_student_name, "")
                if selected_student_name
                else _text(values[1])
            )
            subject_code = (
                subject_by_name.get(selected_subject_name, "")
                if selected_subject_name
                else _text(values[3])
            )
            teacher_id = (
                teacher_by_name.get(selected_teacher_name, "")
                if selected_teacher_name
                else _text(values[5])
            )
            priority_value, one_to_one_value, note_value = values[6], values[7], values[8]
        else:
            student_id = _text(values[0]) or student_by_name.get(_text(values[1]), "")
            subject_code = _text(values[3]) or subject_by_name.get(_text(values[4]), "")
            teacher_id = _text(values[6]) or teacher_by_name.get(_text(values[7]), "")
            priority_value, one_to_one_value, note_value = values[9], values[10], values[11]
        if student_id not in student_ids:
            errors.append(f"通常授業 {row_number}行: 生徒が基本情報にありません")
        if subject_code not in subject_codes:
            errors.append(f"通常授業 {row_number}行: 科目が科目シートにありません")
        if teacher_id and teacher_id not in teacher_ids:
            errors.append(f"通常授業 {row_number}行: 通常担当講師が基本情報にありません")
        key = (student_id, subject_code)
        if key in used:
            errors.append(f"通常授業 {row_number}行: 生徒と科目の組が重複しています")
            continue
        used.add(key)
        priority = _integer(priority_value, 3)
        if not 1 <= priority <= 5:
            errors.append(f"通常授業 {row_number}行: 優先度は1～5です")
        result.append(
            SharedRegularLesson(
                student_id,
                subject_code,
                teacher_id,
                priority,
                _boolean(one_to_one_value, False),
                _text(note_value),
            )
        )
    return result


def _xlsx_formats(workbook: Any) -> dict[str, Any]:
    common = {
        "valign": "top",
        "text_wrap": True,
        "bottom": 1,
        "bottom_color": "#D9DEE7",
    }
    return {
        "header": workbook.add_format(
            {
                "bg_color": "#1F4E78",
                "font_color": "#FFFFFF",
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
            }
        ),
        "helper_header": workbook.add_format(
            {
                "bg_color": "#5B9BD5",
                "font_color": "#FFFFFF",
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
            }
        ),
        "normal": workbook.add_format(common),
        "required": workbook.add_format({**common, "bg_color": "#FFF2CC"}),
        "checkbox": workbook.add_format(
            {**common, "checkbox": True, "align": "center", "valign": "vcenter"}
        ),
        "inactive": workbook.add_format({"bg_color": "#E7E6E6"}),
    }


def _prepare_student_input_rows(sheet: Any, first_blank_row: int, formats: dict[str, Any]) -> None:
    """姓を入力するとID・既定値が表示される、手動変更可能な在籍行を用意する。"""
    for excel_row in range(first_blank_row, _FORMULA_TEMPLATE_MAX_ROW + 1):
        row = excel_row - 1
        # 数式セルはExcelで直接オン・オフできないため、空行も初期値オンの
        # ネイティブセルチェックボックスにする。姓が空欄の行は読込み時に無視される。
        sheet.insert_checkbox(row, 0, True, formats["checkbox"])
        sheet.write_formula(
            row,
            1,
            _next_id_formula(excel_row, "S", first_blank_row),
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 2, None, formats["required"])
        sheet.write_blank(row, 3, None, formats["normal"])
        sheet.write_formula(
            row,
            4,
            f'=IF(COUNTA(C{excel_row}:D{excel_row})=0,"",TRIM(C{excel_row}&" "&D{excel_row}))',
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 5, None, formats["required"])
        sheet.write_formula(
            row,
            6,
            f'=IF(C{excel_row}="","",2)',
            formats["normal"],
            "",
        )
        sheet.write_formula(
            row,
            7,
            f'=IF(C{excel_row}="","","なし")',
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 8, None, formats["normal"])


def _prepare_teacher_input_rows(sheet: Any, first_blank_row: int, formats: dict[str, Any]) -> None:
    """姓を入力するとID・既定値が表示される、手動変更可能な在籍行を用意する。"""
    for excel_row in range(first_blank_row, _FORMULA_TEMPLATE_MAX_ROW + 1):
        row = excel_row - 1
        sheet.insert_checkbox(row, 0, True, formats["checkbox"])
        sheet.write_formula(
            row,
            1,
            _next_id_formula(excel_row, "T", first_blank_row),
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 2, None, formats["required"])
        sheet.write_blank(row, 3, None, formats["normal"])
        sheet.write_formula(
            row,
            4,
            f'=IF(COUNTA(C{excel_row}:D{excel_row})=0,"",TRIM(C{excel_row}&" "&D{excel_row}))',
            formats["normal"],
            "",
        )
        sheet.write_formula(
            row,
            5,
            f'=IF(C{excel_row}="","","なし")',
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 6, None, formats["normal"])


def _prepare_qualification_input_rows(
    sheet: Any,
    first_blank_row: int,
    formats: dict[str, Any],
) -> None:
    """名前を選ぶとID・科目コード・既定値が埋まる入力行を用意する。"""
    for excel_row in range(first_blank_row, _FORMULA_TEMPLATE_MAX_ROW + 1):
        row = excel_row - 1
        sheet.write_blank(row, 0, None, formats["required"])
        sheet.write_formula(
            row,
            1,
            _selected_lookup_formula(excel_row, "A", "講師", "E", "B"),
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 2, None, formats["required"])
        sheet.write_formula(
            row,
            3,
            _selected_lookup_formula(excel_row, "C", "科目", "B", "A"),
            formats["normal"],
            "",
        )
        sheet.write_formula(
            row,
            4,
            f'=IF(OR(A{excel_row}="",C{excel_row}=""),"","はい")',
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 5, None, formats["normal"])


def _prepare_regular_lesson_input_rows(
    sheet: Any,
    first_blank_row: int,
    formats: dict[str, Any],
) -> None:
    """生徒・科目・講師名を選ぶと参照IDと既定値が埋まる入力行を用意する。"""
    for excel_row in range(first_blank_row, _FORMULA_TEMPLATE_MAX_ROW + 1):
        row = excel_row - 1
        sheet.write_blank(row, 0, None, formats["required"])
        sheet.write_formula(
            row,
            1,
            _selected_lookup_formula(excel_row, "A", "生徒", "E", "B"),
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 2, None, formats["required"])
        sheet.write_formula(
            row,
            3,
            _selected_lookup_formula(excel_row, "C", "科目", "B", "A"),
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 4, None, formats["normal"])
        sheet.write_formula(
            row,
            5,
            _selected_lookup_formula(excel_row, "E", "講師", "E", "B"),
            formats["normal"],
            "",
        )
        sheet.write_formula(
            row,
            6,
            f'=IF(OR(A{excel_row}="",C{excel_row}=""),"",3)',
            formats["normal"],
            "",
        )
        sheet.write_formula(
            row,
            7,
            f'=IF(OR(A{excel_row}="",C{excel_row}=""),"","いいえ")',
            formats["normal"],
            "",
        )
        sheet.write_blank(row, 8, None, formats["normal"])


def _next_id_formula(row: int, prefix: PersonIdPrefix, first_blank_row: int) -> str:
    """新規姓の順番に対応する未使用ID候補を補助シートから表示する。"""
    helper_column = "A" if prefix == "S" else "B"
    surname_count = f'COUNTIF($C${first_blank_row}:C{row},"?*")'
    return (
        f'=IF(C{row}="","",INDEX(\'{_ID_HELPER_SHEET}\'!'
        f"${helper_column}$1:${helper_column}$999,{surname_count}))"
    )


def _create_id_helper_sheet(
    workbook: Any,
    data: SharedRosterData,
    *,
    reserved_student_ids: Iterable[str] = (),
    reserved_teacher_ids: Iterable[str] = (),
) -> None:
    """数式同士の配列計算に依存せず、一意なID候補を参照できるようにする。"""
    sheet = workbook.add_worksheet(_ID_HELPER_SHEET)
    student_ids = _available_person_ids(
        (
            *(student.external_id for student in data.students),
            *reserved_student_ids,
        ),
        prefix="S",
    )
    teacher_ids = _available_person_ids(
        (
            *(teacher.external_id for teacher in data.teachers),
            *reserved_teacher_ids,
        ),
        prefix="T",
    )
    for row, (student_id, teacher_id) in enumerate(zip(student_ids, teacher_ids, strict=True)):
        sheet.write_string(row, 0, student_id)
        sheet.write_string(row, 1, teacher_id)
    sheet.very_hidden()


def _available_person_ids(
    existing_ids: Iterable[str], *, prefix: PersonIdPrefix
) -> tuple[str, ...]:
    reserved = {value.strip() for value in existing_ids if value.strip()}
    result: list[str] = []
    for _ in range(_FORMULA_TEMPLATE_MAX_ROW - 1):
        external_id = next_person_external_id(reserved, prefix=prefix)
        result.append(external_id)
        reserved.add(external_id)
    return tuple(result)


def _status_is_first_column(sheet: Any) -> bool:
    return _text(sheet.cell(1, 1).value).startswith("在籍")


def _setup_sheet(
    sheet: Any,
    headers: tuple[str, ...],
    widths: tuple[int, ...],
    formats: dict[str, Any],
    *,
    helper_columns: set[int] | None = None,
) -> None:
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    sheet.set_row(0, 34)
    helpers = helper_columns or set()
    for column, (header, width) in enumerate(zip(headers, widths, strict=True)):
        header_format = formats["helper_header"] if column + 1 in helpers else formats["header"]
        sheet.write(0, column, header, header_format)
        sheet.write_comment(
            0,
            column,
            "（必須）は入力必須です。空欄時の既定値は列の入力規則に従います。",
            {"author": "夏期講習時間割作成アプリ"},
        )
        sheet.set_column(column, column, width)


def _write_plain_row(sheet: Any, row: int, values: tuple[object, ...], cell_format: Any) -> None:
    for column, value in enumerate(values):
        sheet.write(row, column, value, cell_format)


def _finish_rows(sheet: Any, data_count: int, column_count: int) -> None:
    # 入力開始用の空行は含めるが、数式だけの将来行をソート対象にしない。
    last_row = max(data_count + 1, 1)
    sheet.autofilter(0, 0, last_row, column_count - 1)


def _mark_auto_id_header(sheet: Any, coordinate: str, person_label: str, example: str) -> None:
    sheet.write_comment(
        coordinate,
        f"入力不要です。{person_label}の姓を入力すると{example}形式の候補を表示します。"
        "アプリで反映すると、既存IDと衝突しない正式IDとしてこの列へ書き戻します。",
        {"author": "夏期講習時間割作成アプリ"},
    )


def _grey_inactive(sheet: Any, column_count: int, inactive_format: Any) -> None:
    sheet.conditional_format(
        1,
        0,
        _MAX_ROW - 1,
        column_count - 1,
        {
            "type": "formula",
            "criteria": '=AND($C2<>"",$A2=FALSE)',
            "format": inactive_format,
        },
    )


def _add_list(
    sheet: Any,
    cells: str,
    source: list[str] | str,
    allow_blank: bool = False,
) -> None:
    sheet.data_validation(
        cells,
        {
            "validate": "list",
            "source": source,
            "ignore_blank": allow_blank,
            "error_title": "入力値を確認してください",
            "error_message": "一覧から選択してください。",
            "show_error": True,
        },
    )


def _add_whole(
    sheet: Any,
    cells: str,
    minimum: int,
    maximum: int,
    allow_blank: bool = False,
) -> None:
    sheet.data_validation(
        cells,
        {
            "validate": "integer",
            "criteria": "between",
            "minimum": minimum,
            "maximum": maximum,
            "ignore_blank": allow_blank,
            "error_title": "入力値を確認してください",
            "error_message": f"{minimum}～{maximum}の整数を入力してください。",
            "show_error": True,
        },
    )


def _linked_display_formula(
    row: int,
    id_column: str,
    source_sheet: str,
    source_id_column: str,
    source_name_column: str,
) -> str:
    """安定IDを参照し、マスター側の最新表示名へ追従する。"""
    return (
        f'=IF({id_column}{row}="","",IFERROR(INDEX(\'{source_sheet}\'!${source_name_column}:${source_name_column},'
        f"MATCH({id_column}{row},'{source_sheet}'!${source_id_column}:${source_id_column},0)),\"⚠ 未登録\"))"
    )


def _selected_lookup_formula(
    row: int,
    selected_column: str,
    source_sheet: str,
    source_lookup_column: str,
    source_result_column: str,
) -> str:
    """プルダウンで選んだ表示名からIDまたはコードを引く。"""
    return (
        f'=IF({selected_column}{row}="","",IFERROR(INDEX(\'{source_sheet}\'!'
        f"${source_result_column}:${source_result_column},MATCH({selected_column}{row},"
        f"'{source_sheet}'!${source_lookup_column}:${source_lookup_column},0)),\"⚠ 未登録\"))"
    )


def _rows(sheet: Any, width: int) -> list[tuple[int, tuple[object, ...]]]:
    return [
        (row_number, tuple(sheet.cell(row_number, column).value for column in range(1, width + 1)))
        for row_number in range(2, sheet.max_row + 1)
    ]


def _empty(values: tuple[object, ...]) -> bool:
    return all(not _text(value) for value in values if not _is_formula(value))


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _text(value: object) -> str:
    if value is None or _is_formula(value):
        return ""
    return str(value).strip()


def _integer(value: object, default: int) -> int:
    text = _text(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return -1


def _boolean(value: object, default: bool) -> bool:
    text = _text(value).casefold()
    if not text:
        return default
    if text in {"はい", "あり", "yes", "true", "1", "☑", "可", "在籍"}:
        return True
    if text in {"いいえ", "なし", "no", "false", "0", "☐", "不可", "退籍"}:
        return False
    return default


def _active(value: object) -> bool:
    text = _text(value).casefold()
    if not text:
        return True
    return text not in {"☐ 退籍", "退籍", "いいえ", "false", "0"}


def _yes_no(value: bool) -> str:
    return "はい" if value else "いいえ"


def _gap_label(value: bool) -> str:
    return "あり" if value else "なし"


def _school_level_label(value: str) -> str:
    return {
        "elementary": "小学校",
        "junior_high": "中学校",
        "high_school": "高校",
    }.get(value, value)


def _school_level_code(value: str) -> str:
    return {
        "小学校": "elementary",
        "中学校": "junior_high",
        "高校": "high_school",
        "elementary": "elementary",
        "junior_high": "junior_high",
        "high_school": "high_school",
    }.get(value, "")


def _full_name(surname: str, given_name: str) -> str:
    return " ".join(part for part in (surname.strip(), given_name.strip()) if part)


def _unique_name_map(rows: list[SharedStudent] | list[SharedTeacher]) -> dict[str, str]:
    grouped: dict[str, list[str]] = {}
    for row in rows:
        grouped.setdefault(row.name, []).append(row.external_id)
    return {name: ids[0] for name, ids in grouped.items() if name and len(ids) == 1}


def _unique_subject_name_map(rows: list[SharedSubject]) -> dict[str, str]:
    grouped: dict[str, list[str]] = {}
    for row in rows:
        grouped.setdefault(row.display_name, []).append(row.code)
    return {name: codes[0] for name, codes in grouped.items() if name and len(codes) == 1}


__all__ = [
    "SHARED_ROSTER_FILENAME",
    "SharedQualification",
    "SharedRegularLesson",
    "SharedRosterData",
    "SharedRosterError",
    "SharedStudent",
    "SharedSubject",
    "SharedTeacher",
    "empty_shared_roster",
    "read_shared_roster",
    "write_shared_roster",
]
