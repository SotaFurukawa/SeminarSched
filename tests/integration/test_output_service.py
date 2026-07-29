"""Phase 6出力ServiceのDB再診断・設定・保存結合テスト。"""

from __future__ import annotations

from dataclasses import replace
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from summer_scheduler.application.output_service import (
    OutputDataIntegrityError,
    OutputService,
)
from summer_scheduler.application.project_service import ProjectService
from summer_scheduler.infrastructure.db import Database, create_database, upgrade_database
from summer_scheduler.infrastructure.db.models import (
    Assignment,
    LessonRequest,
    Student,
    StudentAvailability,
    Subject,
    Teacher,
    TeacherAvailability,
    TeacherQualification,
    TimeSlot,
)
from summer_scheduler.reporting.data import OutputSelection
from summer_scheduler.shared.settings import (
    OptimizationAppSettings,
    load_settings,
)


def test_output_service_rebuilds_unassigned_and_exports_excel_csv(
    tmp_path: Path,
) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        graph = _seed_valid_schedule(projects)
        service = OutputService(projects, _optimization_settings())

        workspace = service.load_workspace()

        assert workspace.assignment_count == 2
        assert workspace.unassigned_count == 1
        assert workspace.warning_count >= 0
        assert workspace.settings.paper_size == "A3"
        assert service.suggested_filename("overall", "xlsx") == "夏期講習時間割.xlsx"
        assert service.suggested_filename("raw", ".csv") == "割当て生データ.csv"

        issues = service.build_document(
            "issues", OutputSelection(student_ids=(graph.student_1_id,))
        )
        issue_text = "\n".join(
            cell.text
            for section in issues.sections
            for page in section.pages
            for table in page.tables
            for row in table.rows
            for cell in row.cells
        )
        assert "架空 生徒一" in issue_text
        assert "単独配置可" in issue_text

        output_dir = tmp_path / "日本語の出力先"
        excel_result = service.export_excel(
            "overall",
            output_dir / "夏期講習時間割.xlsx",
        )
        csv_result = service.export_csv(output_dir / "割当て生データ.csv")
        filtered_csv_result = service.export_csv(
            output_dir / "対象生徒のみ.csv",
            OutputSelection(student_ids=(graph.student_1_id,)),
        )

        assert excel_result.path.is_file()
        workbook = load_workbook(excel_result.path, read_only=False, data_only=False)
        try:
            assert workbook.sheetnames[0].startswith("全体時間割")
            assert workbook[workbook.sheetnames[0]]["A1"].value == "夏期講習時間割"
        finally:
            workbook.close()
        assert csv_result.path.read_bytes().startswith(b"\xef\xbb\xbf")
        csv_text = csv_result.path.read_text(encoding="utf-8-sig")
        assert "架空 生徒一" in csv_text
        assert "group_lesson" in csv_text.splitlines()[0]
        assert filtered_csv_result.record_count == 1
        assert len(filtered_csv_result.path.read_text(encoding="utf-8-sig").splitlines()) == 2
    finally:
        projects.close_project()
        registry.dispose()


def test_output_settings_round_trip_and_filename_sanitization(tmp_path: Path) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        _seed_valid_schedule(projects)
        service = OutputService(projects, _optimization_settings())
        workspace = service.load_workspace()
        changed = replace(
            workspace.settings,
            paper_size="A4",
            orientation="portrait",
            days_per_page=1,
            teacher_columns_per_page=3,
            font_size=9.0,
            margin_mm=12.0,
            file_name_pattern="{project}_{report}_{date}",
            csv_with_bom=False,
        )

        saved = service.save_settings(changed)
        service.invalidate()
        reloaded = service.load_workspace(refresh=False).settings

        assert saved == reloaded
        assert reloaded.paper_size == "A4"
        today = datetime.now().astimezone().strftime("%Y%m%d")
        assert service.suggested_filename("overall", "pdf") == (
            f"架空校 夏期講習_夏期講習時間割_{today}.pdf"
        )
        assert (
            service.suggested_filename(
                "overall",
                "xlsx",
                settings_override=replace(reloaded, file_name_pattern="CON.foo"),
            )
            == "_CON.foo.xlsx"
        )
    finally:
        projects.close_project()
        registry.dispose()


def test_unassigned_resolution_omits_candidates_that_break_current_hard_constraints(
    tmp_path: Path,
) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        graph = _seed_valid_schedule(projects)
        database = projects.require_database()
        with database.session_factory.begin() as session:
            request = session.scalar(
                select(LessonRequest).where(
                    LessonRequest.project_id == graph.project_id,
                    LessonRequest.student_id == graph.student_1_id,
                )
            )
            slot_z = session.scalar(
                select(TimeSlot).where(
                    TimeSlot.project_id == graph.project_id,
                    TimeSlot.code == "Z",
                )
            )
            other_assignment = session.scalar(
                select(Assignment)
                .join(
                    LessonRequest,
                    LessonRequest.id == Assignment.lesson_request_id,
                )
                .where(
                    Assignment.project_id == graph.project_id,
                    LessonRequest.student_id != graph.student_1_id,
                )
            )
            assert request is not None
            assert slot_z is not None
            assert other_assignment is not None
            request.one_to_one_required = True
            other_assignment.time_slot_id = slot_z.id

        issues = OutputService(
            projects,
            _optimization_settings(),
        ).build_document("issues", OutputSelection(student_ids=(graph.student_1_id,)))
        text = "\n".join(
            cell.text
            for section in issues.sections
            for page in section.pages
            for table in page.tables
            for row in table.rows
            for cell in row.cells
        )

        assert "単独配置可" not in text
        assert "1対1専用の講師枠を確保" in text
    finally:
        projects.close_project()
        registry.dispose()


def test_yaml_defaults_apply_only_until_project_settings_are_saved(
    tmp_path: Path,
) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        _seed_valid_schedule(projects)
        yaml_defaults = replace(
            load_settings().output,
            paper_size="A4",
            orientation="portrait",
            days_per_page=4,
            teacher_columns_per_page=5,
            font_size=10.0,
            margin_mm=7.5,
            csv_with_bom=False,
            visible_fields=("subject", "grade", "warning"),
            file_name_pattern="{project}_{report}",
            default_output_directory_optional=r"C:\架空既定出力",
        )
        service = OutputService(
            projects,
            _optimization_settings(),
            output_defaults=yaml_defaults,
        )

        workspace = service.load_workspace()

        assert workspace.settings == yaml_defaults.for_project(workspace.project_id)
        assert service.suggested_filename("overall", "pdf") == (
            "架空校 夏期講習_夏期講習時間割.pdf"
        )

        project_settings = replace(
            workspace.settings,
            paper_size="A3",
            days_per_page=1,
            csv_with_bom=True,
        )
        service.save_settings(project_settings)
        different_defaults = replace(
            yaml_defaults,
            paper_size="A4",
            days_per_page=7,
            csv_with_bom=False,
        )
        reloaded = OutputService(
            projects,
            _optimization_settings(),
            output_defaults=different_defaults,
        ).load_workspace()

        assert reloaded.settings == project_settings
    finally:
        projects.close_project()
        registry.dispose()


def test_file_export_refreshes_database_after_workspace_was_cached(
    tmp_path: Path,
) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        graph = _seed_valid_schedule(projects)
        service = OutputService(projects, _optimization_settings())
        assert service.load_workspace().assignment_count == 2

        database = projects.require_database()
        with database.session_factory.begin() as session:
            request = session.scalar(
                select(LessonRequest).where(
                    LessonRequest.project_id == graph.project_id,
                    LessonRequest.student_id == graph.student_1_id,
                )
            )
            slot_z = session.scalar(
                select(TimeSlot).where(
                    TimeSlot.project_id == graph.project_id,
                    TimeSlot.code == "Z",
                )
            )
            assert request is not None
            assert slot_z is not None
            session.add(
                Assignment(
                    project_id=graph.project_id,
                    lesson_request_id=request.id,
                    session_index=2,
                    date=graph.day,
                    time_slot_id=slot_z.id,
                    teacher_id=graph.teacher_id,
                    is_locked=False,
                    is_manual=True,
                    created_by="test",
                    note="画面読込み後の変更",
                )
            )

        destination = tmp_path / "最新化確認" / "割当て生データ.csv"
        result = service.export_csv(destination)

        assert result.record_count == 3
        rows = destination.read_text(encoding="utf-8-sig").splitlines()
        assert len(rows) == 4
        assert any("画面読込み後の変更" in row and ",Z," in row for row in rows)
    finally:
        projects.close_project()
        registry.dispose()


def test_hard_constraint_violation_blocks_all_output(tmp_path: Path) -> None:
    projects, registry = _project_service(tmp_path)
    try:
        graph = _seed_valid_schedule(projects)
        service = OutputService(projects, _optimization_settings())
        service.load_workspace()
        database = projects.require_database()
        with database.session_factory.begin() as session:
            subject = session.get(Subject, graph.subject_id)
            assert subject is not None
            student = Student(
                external_id="S-P6-003",
                name="架空 生徒三",
                grade="中学3年",
                default_max_consecutive_slots=2,
                allow_gap=False,
                active=True,
            )
            session.add(student)
            session.flush()
            request = LessonRequest(
                project_id=graph.project_id,
                student_id=student.id,
                subject_id=subject.id,
                required_sessions=1,
                regular_teacher_id_optional=graph.teacher_id,
                regular_teacher_priority=3,
                one_to_one_required=False,
            )
            session.add(request)
            session.flush()
            session.add(
                StudentAvailability(
                    project_id=graph.project_id,
                    student_id=student.id,
                    date=graph.day,
                    time_slot_id=graph.slot_y_id,
                    availability_level=2,
                )
            )
            session.add(
                Assignment(
                    project_id=graph.project_id,
                    lesson_request_id=request.id,
                    session_index=1,
                    date=graph.day,
                    time_slot_id=graph.slot_y_id,
                    teacher_id=graph.teacher_id,
                    is_locked=False,
                    is_manual=True,
                    created_by="test",
                )
            )
        service.invalidate()

        with pytest.raises(OutputDataIntegrityError, match="ハード制約違反"):
            service.load_workspace(refresh=False)

        assert not (tmp_path / "違反時間割.xlsx").exists()
    finally:
        projects.close_project()
        registry.dispose()


class _Graph:
    def __init__(
        self,
        *,
        project_id: int,
        day: date,
        student_1_id: int,
        teacher_id: int,
        subject_id: int,
        slot_y_id: int,
    ) -> None:
        self.project_id = project_id
        self.day = day
        self.student_1_id = student_1_id
        self.teacher_id = teacher_id
        self.subject_id = subject_id
        self.slot_y_id = slot_y_id


def _project_service(tmp_path: Path) -> tuple[ProjectService, Database]:
    registry = create_database(tmp_path / "registry.db")
    upgrade_database(registry.engine)
    projects = ProjectService(registry, tmp_path / "バックアップ")
    projects.create_project(
        tmp_path / "日本語プロジェクト" / "Phase6出力.jukuschedule",
        title="架空校 夏期講習",
        campus_name="架空みらい校",
        start_date=date(2026, 8, 1),
        end_date=date(2026, 8, 1),
    )
    return projects, registry


def _seed_valid_schedule(projects: ProjectService) -> _Graph:
    project_id = projects.require_project().project_id
    database = projects.require_database()
    day = date(2026, 8, 1)
    with database.session_factory.begin() as session:
        subject = session.scalar(select(Subject).where(Subject.code == "JH_MATH"))
        slot_y = session.scalar(select(TimeSlot).where(TimeSlot.code == "Y"))
        slot_z = session.scalar(select(TimeSlot).where(TimeSlot.code == "Z"))
        assert subject is not None
        assert slot_y is not None
        assert slot_z is not None
        students = (
            Student(
                external_id="S-P6-001",
                name="架空 生徒一",
                grade="中学1年",
                default_max_consecutive_slots=2,
                allow_gap=False,
                active=True,
            ),
            Student(
                external_id="S-P6-002",
                name="架空 生徒二",
                grade="中学2年",
                default_max_consecutive_slots=2,
                allow_gap=False,
                active=True,
            ),
        )
        teacher = Teacher(
            external_id="T-P6-001",
            name="架空 講師一",
            allow_gap=False,
            active=True,
        )
        session.add_all([*students, teacher])
        session.flush()
        session.add(
            TeacherQualification(
                teacher_id=teacher.id,
                subject_id=subject.id,
                can_teach=True,
            )
        )
        requests = (
            LessonRequest(
                project_id=project_id,
                student_id=students[0].id,
                subject_id=subject.id,
                required_sessions=2,
                regular_teacher_id_optional=teacher.id,
                regular_teacher_priority=3,
                one_to_one_required=False,
            ),
            LessonRequest(
                project_id=project_id,
                student_id=students[1].id,
                subject_id=subject.id,
                required_sessions=1,
                regular_teacher_id_optional=teacher.id,
                regular_teacher_priority=3,
                one_to_one_required=False,
            ),
        )
        session.add_all(requests)
        session.flush()
        for student in students:
            for slot in (slot_y, slot_z):
                session.add(
                    StudentAvailability(
                        project_id=project_id,
                        student_id=student.id,
                        date=day,
                        time_slot_id=slot.id,
                        availability_level=2,
                    )
                )
        for slot in (slot_y, slot_z):
            session.add(
                TeacherAvailability(
                    project_id=project_id,
                    teacher_id=teacher.id,
                    date=day,
                    time_slot_id=slot.id,
                    availability_level=2,
                )
            )
        session.add_all(
            (
                Assignment(
                    project_id=project_id,
                    lesson_request_id=requests[0].id,
                    session_index=1,
                    date=day,
                    time_slot_id=slot_y.id,
                    teacher_id=teacher.id,
                    is_locked=False,
                    is_manual=False,
                    created_by="solver",
                ),
                Assignment(
                    project_id=project_id,
                    lesson_request_id=requests[1].id,
                    session_index=1,
                    date=day,
                    time_slot_id=slot_y.id,
                    teacher_id=teacher.id,
                    is_locked=False,
                    is_manual=False,
                    created_by="solver",
                ),
            )
        )
        return _Graph(
            project_id=project_id,
            day=day,
            student_1_id=students[0].id,
            teacher_id=teacher.id,
            subject_id=subject.id,
            slot_y_id=slot_y.id,
        )


def _optimization_settings() -> OptimizationAppSettings:
    return OptimizationAppSettings(
        default_preset="fast",
        fast_time_limit_seconds=1.0,
        standard_time_limit_seconds=2.0,
        high_quality_time_limit_seconds=3.0,
        random_seed=20260729,
        num_search_workers=1,
        regular_teacher_priority_weights=(1, 3, 6, 10),
        preferred_teacher_rank_weights=(9, 6, 3),
        student_preferred_time_weight=1,
        teacher_preferred_time_weight=1,
        preserve_existing_assignment_weight=1,
        optional_balance_weight=0,
    )
