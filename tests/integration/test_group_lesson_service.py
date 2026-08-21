"""Phase 3 集団授業取込みサービスの統合テスト。"""

from __future__ import annotations

import json
from collections.abc import Iterator
from datetime import date, time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from summer_scheduler.application.group_lesson_service import (
    GroupLessonImportError,
    GroupLessonService,
)
from summer_scheduler.application.project_service import ProjectService
from summer_scheduler.infrastructure.db import create_database, upgrade_database
from summer_scheduler.infrastructure.db.models import (
    AuditLog,
    GroupLesson,
    ImportBatch,
    Student,
    Subject,
    Teacher,
    TeacherQualification,
)
from summer_scheduler.infrastructure.importing import (
    GROUP_LESSON_SHEET,
    GROUP_PARTICIPANT_SHEET,
)
from summer_scheduler.infrastructure.repositories import MasterRepository


@pytest.fixture
def group_service(tmp_path: Path) -> Iterator[GroupLessonService]:
    registry = create_database(tmp_path / "registry.db")
    upgrade_database(registry.engine)
    projects = ProjectService(registry, tmp_path / "backups")
    projects.create_project(
        tmp_path / "集団授業.jukuschedule",
        title="夏期講習",
        campus_name="テスト校",
        start_date=date(2026, 8, 1),
        end_date=date(2026, 8, 2),
    )
    _seed_masters(projects)
    yield GroupLessonService(projects)
    projects.close_project()
    registry.dispose()


def test_import_diff_apply_and_explicit_deletion(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "日本語の集団授業.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-数学",
                "J2",
                "JH_MATH",
                "夏期数学",
                date(2026, 8, 1),
                time(17, 10),
                time(18, 10),
                "T-001",
                "教室A",
                "日本語名のテスト",
            )
        ],
        participants=[("G-数学", "S-001"), ("G-数学", "S-002")],
    )

    preview = group_service.prepare_group_import(workbook)
    assert not preview.has_errors
    assert [diff.operation for diff in preview.diffs] == ["add"]

    result = group_service.apply_group_import(preview)
    assert result.added == 1
    database = group_service._projects.require_database()
    with database.session_factory() as session:
        audit = session.scalar(select(AuditLog).order_by(AuditLog.id.desc()))
        assert audit is not None
        assert audit.before_json is None
        assert audit.after_json is not None
        audit_payload = json.loads(audit.after_json)
        assert set(audit_payload) == {
            "added",
            "changed",
            "deleted",
            "source_file_name",
            "unchanged",
        }
        assert "日本語名のテスト" not in audit.after_json
        assert "S-001" not in audit.after_json
        assert str(workbook.parent) not in audit.after_json
    listed = group_service.list_group_lessons()
    assert [(row.group_code, row.student_count, row.teacher_name) for row in listed] == [
        ("G-数学", 2, "講師 一郎"),
    ]
    assert listed[0].grade == "中2"

    _write_group_rows(workbook, lessons=[], participants=[])
    deletion_preview = group_service.prepare_group_import(workbook)
    assert [diff.operation for diff in deletion_preview.diffs] == ["delete_candidate"]
    group_service.apply_group_import(deletion_preview, include_deletes=False)
    assert len(group_service.list_group_lessons()) == 1
    group_service.apply_group_import(deletion_preview, include_deletes=True)
    assert group_service.list_group_lessons() == ()


def test_calendar_options_create_and_delete_group_lesson(
    group_service: GroupLessonService,
) -> None:
    options = group_service.calendar_options()
    assert {row["value"] for row in options["dates"]} == {
        "2026-08-01",
        "2026-08-02",
    }
    assert any(row["code"] == "JH_MATH" for row in options["subjects"])
    assert any(row["externalId"] == "T-001" for row in options["teachers"])

    group_id = group_service.create_calendar_lesson(
        grade="中2",
        subject_code="JH_MATH",
        day=date(2026, 8, 1),
        start_time=time(17, 10),
        end_time=time(18, 30),
        course_name="受験数学",
        teacher_external_id="T-001",
        room="教室A",
    )
    listed = group_service.list_group_lessons()
    assert len(listed) == 1
    assert listed[0].grade == "中2"
    assert listed[0].subject_name == "中学校・数学"
    assert listed[0].teacher_name == "講師 一郎"

    assert group_service.delete_calendar_lesson(group_id)
    assert group_service.list_group_lessons() == ()


def test_boundary_touch_is_allowed_but_overlapping_student_is_rejected(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "境界.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-1",
                "中2",
                "JH_MATH",
                "前半",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            ),
            (
                "G-2",
                "中2",
                "JH_MATH",
                "後半",
                date(2026, 8, 1),
                time(18, 0),
                time(19, 0),
                "T-001",
                "A",
                "",
            ),
        ],
        participants=[("G-1", "S-001"), ("G-2", "S-001")],
    )
    assert not group_service.prepare_group_import(workbook).has_errors

    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-1",
                "中2",
                "JH_MATH",
                "前半",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            ),
            (
                "G-2",
                "中2",
                "JH_MATH",
                "重複",
                date(2026, 8, 1),
                time(17, 30),
                time(18, 30),
                "T-002",
                "A",
                "",
            ),
        ],
        participants=[("G-1", "S-001"), ("G-2", "S-001")],
    )
    preview = group_service.prepare_group_import(workbook)
    assert preview.has_errors
    assert "student_time_conflict" in {issue.code for issue in preview.issues}
    with pytest.raises(GroupLessonImportError):
        group_service.apply_group_import(preview)


def test_overlapping_same_teacher_is_rejected_without_shared_student(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "same_teacher_overlap.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-TEACHER-1",
                "中2",
                "JH_MATH",
                "前半",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            ),
            (
                "G-TEACHER-2",
                "中2",
                "JH_MATH",
                "後半",
                date(2026, 8, 1),
                time(17, 30),
                time(18, 30),
                "T-001",
                "B",
                "",
            ),
        ],
        participants=[("G-TEACHER-1", "S-001"), ("G-TEACHER-2", "S-002")],
    )

    preview = group_service.prepare_group_import(workbook)

    codes = {issue.code for issue in preview.issues}
    assert "teacher_time_conflict" in codes
    assert "student_time_conflict" not in codes


def test_existing_group_conflicts_with_new_teacher_and_student_are_rejected(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "existing_group_conflicts.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-EXISTING",
                "中2",
                "JH_MATH",
                "既存",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-EXISTING", "S-001")],
    )
    group_service.apply_group_import(group_service.prepare_group_import(workbook))

    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-TEACHER-CONFLICT",
                "中2",
                "JH_MATH",
                "講師重複",
                date(2026, 8, 1),
                time(17, 30),
                time(18, 30),
                "T-001",
                "B",
                "",
            ),
            (
                "G-STUDENT-CONFLICT",
                "中2",
                "JH_MATH",
                "生徒重複",
                date(2026, 8, 1),
                time(17, 30),
                time(18, 30),
                "",
                "C",
                "",
            ),
        ],
        participants=[
            ("G-TEACHER-CONFLICT", "S-002"),
            ("G-STUDENT-CONFLICT", "S-001"),
        ],
    )

    preview = group_service.prepare_group_import(workbook)

    codes = {issue.code for issue in preview.issues}
    assert "teacher_time_conflict" in codes
    assert "student_time_conflict" in codes
    with pytest.raises(GroupLessonImportError):
        group_service.apply_group_import(preview)
    assert [row.group_code for row in group_service.list_group_lessons()] == ["G-EXISTING"]


def test_apply_rejects_changed_file_after_preview_without_changing_database(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "changed_after_preview.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-PREVIEW",
                "中2",
                "JH_MATH",
                "確認前",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-PREVIEW", "S-001")],
    )
    preview = group_service.prepare_group_import(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-MODIFIED",
                "中2",
                "JH_MATH",
                "確認後の変更",
                date(2026, 8, 1),
                time(17, 0),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-MODIFIED", "S-002")],
    )

    with pytest.raises(GroupLessonImportError, match="再度検証"):
        group_service.apply_group_import(preview)

    database = group_service._projects.require_database()
    with database.session_factory() as session:
        assert list(session.scalars(select(GroupLesson))) == []
        assert list(session.scalars(select(ImportBatch))) == []
        assert list(session.scalars(select(AuditLog))) == []


def test_unqualified_teacher_is_rejected(group_service: GroupLessonService, tmp_path: Path) -> None:
    workbook = tmp_path / "資格.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-資格",
                "中2",
                "JH_MATH",
                "数学",
                date(2026, 8, 1),
                time(17, 10),
                time(18, 0),
                "T-002",
                "A",
                "",
            )
        ],
        participants=[("G-資格", "S-001")],
    )
    preview = group_service.prepare_group_import(workbook)
    assert "teacher_unqualified" in {issue.code for issue in preview.issues}


def test_participant_cannot_reference_missing_group_lesson(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "存在しない集団授業参照.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-存在",
                "中2",
                "JH_MATH",
                "数学",
                date(2026, 8, 1),
                time(17, 10),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-不存在", "S-001")],
    )

    preview = group_service.prepare_group_import(workbook)
    assert preview.has_errors
    missing_reference = [
        issue for issue in preview.issues if issue.code == "unknown_group_lesson_reference"
    ]
    assert len(missing_reference) == 1
    assert missing_reference[0].sheet == GROUP_PARTICIPANT_SHEET
    assert missing_reference[0].row == 2
    assert missing_reference[0].column == "集団授業ID（必須）"


def test_unknown_participant_reports_actual_sheet_row_and_source_header(
    group_service: GroupLessonService,
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "存在しない参加生徒.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-存在",
                "中2",
                "JH_MATH",
                "数学",
                date(2026, 8, 1),
                time(17, 10),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-存在", "S-不存在")],
    )

    preview = group_service.prepare_group_import(workbook)

    unknown_students = [issue for issue in preview.issues if issue.code == "unknown_student"]
    assert len(unknown_students) == 1
    assert unknown_students[0].sheet == GROUP_PARTICIPANT_SHEET
    assert unknown_students[0].row == 2
    assert unknown_students[0].column == "生徒ID（必須）"


def test_failed_apply_rolls_back_all_group_changes(
    group_service: GroupLessonService,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = tmp_path / "rollback.xlsx"
    group_service.export_template(workbook)
    _write_group_rows(
        workbook,
        lessons=[
            (
                "G-RB",
                "中2",
                "JH_MATH",
                "ロールバック",
                date(2026, 8, 1),
                time(17, 10),
                time(18, 0),
                "T-001",
                "A",
                "",
            )
        ],
        participants=[("G-RB", "S-001")],
    )
    preview = group_service.prepare_group_import(workbook)

    def raise_after_write(self: MasterRepository, batch: ImportBatch) -> ImportBatch:
        raise RuntimeError("audit storage unavailable")

    monkeypatch.setattr(MasterRepository, "create_import_batch", raise_after_write)
    with pytest.raises(RuntimeError, match="audit storage unavailable"):
        group_service.apply_group_import(preview)
    assert group_service.list_group_lessons() == ()


def _seed_masters(projects: ProjectService) -> None:
    database = projects.require_database()
    with database.session_factory.begin() as session:
        subject = session.scalar(select(Subject).where(Subject.code == "JH_MATH"))
        assert subject is not None
        students = [
            Student(
                external_id="S-001",
                name="生徒 太郎",
                grade="中2",
                default_max_consecutive_slots=2,
                allow_gap=False,
                active=True,
            ),
            Student(
                external_id="S-002",
                name="生徒 花子",
                grade="中2",
                default_max_consecutive_slots=2,
                allow_gap=False,
                active=True,
            ),
        ]
        teachers = [
            Teacher(external_id="T-001", name="講師 一郎", allow_gap=False, active=True),
            Teacher(external_id="T-002", name="講師 次郎", allow_gap=False, active=True),
        ]
        session.add_all([*students, *teachers])
        session.flush()
        session.add(
            TeacherQualification(
                teacher_id=teachers[0].id,
                subject_id=subject.id,
                can_teach=True,
            )
        )


def _write_group_rows(
    path: Path,
    *,
    lessons: list[tuple[Any, ...]],
    participants: list[tuple[str, str]],
) -> None:
    workbook = load_workbook(path)
    try:
        lesson_sheet = workbook[GROUP_LESSON_SHEET]
        participant_sheet = workbook[GROUP_PARTICIPANT_SHEET]
        lesson_sheet.delete_rows(2, lesson_sheet.max_row)
        participant_sheet.delete_rows(2, participant_sheet.max_row)
        for row in lessons:
            lesson_sheet.append((False, *row))
        for group_code, student_id in participants:
            participant_sheet.append((False, group_code, student_id))
        workbook.save(path)
    finally:
        workbook.close()
