"""架空master_data.xlsx生成物の契約テスト。"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from tools.generate_master_data_example import build_example_rows, generate_example_workbook

from summer_scheduler.infrastructure.db.base import Base
from summer_scheduler.infrastructure.db.database import create_database
from summer_scheduler.infrastructure.db.models import (
    Campus,
    CourseProject,
    LessonRequest,
    Student,
    Subject,
    Teacher,
    TeacherQualification,
)
from summer_scheduler.infrastructure.excel import MasterDataExcelService
from summer_scheduler.infrastructure.excel.schema import (
    LESSON_REQUEST_SHEET,
    QUALIFICATION_SHEET,
    SHEET_NAMES,
    STUDENT_SHEET,
    SUBJECT_SHEET,
    TEACHER_SHEET,
)


def test_generated_example_imports_and_regular_teachers_are_qualified(
    tmp_path: Path,
) -> None:
    source = generate_example_workbook(tmp_path / "架空マスター50名20名.xlsx")
    rows = build_example_rows()

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        expected_counts = {
            STUDENT_SHEET.name: 50,
            TEACHER_SHEET.name: 20,
            SUBJECT_SHEET.name: 26,
            QUALIFICATION_SHEET.name: 520,
            LESSON_REQUEST_SHEET.name: len(rows[LESSON_REQUEST_SHEET.name]),
        }
        for sheet_name, expected_count in expected_counts.items():
            actual_count = sum(
                row[0] == "いいえ"
                for row in workbook[sheet_name].iter_rows(min_row=3, values_only=True)
            )
            assert actual_count == expected_count
    finally:
        workbook.close()

    database = create_database(tmp_path / "取込検証.jukuschedule")
    Base.metadata.create_all(database.engine)
    with database.session_factory() as session:
        campus = Campus(name="架空校", address_optional=None, logo_path_optional=None)
        session.add(campus)
        session.flush()
        project = CourseProject(
            campus_id=campus.id,
            title="架空講習",
            start_date=date(2026, 7, 20),
            end_date=date(2026, 8, 31),
            status="draft",
            file_version=1,
        )
        session.add(project)
        session.flush()

        service = MasterDataExcelService(session, project.id)
        preview = service.preview_import(source)
        assert not preview.has_errors
        assert not preview.issues
        service.apply_import(preview)
        session.commit()

        assert session.scalar(select(func.count()).select_from(Student)) == 50
        assert session.scalar(select(func.count()).select_from(Teacher)) == 20
        assert session.scalar(select(func.count()).select_from(Subject)) == 26
        assert session.scalar(select(func.count()).select_from(TeacherQualification)) == 20 * 26

        unqualified_regular_count = session.scalar(
            select(func.count())
            .select_from(LessonRequest)
            .outerjoin(
                TeacherQualification,
                (TeacherQualification.teacher_id == LessonRequest.regular_teacher_id_optional)
                & (TeacherQualification.subject_id == LessonRequest.subject_id),
            )
            .where(
                (LessonRequest.regular_teacher_id_optional.is_(None))
                | (TeacherQualification.can_teach.is_(False))
                | (TeacherQualification.teacher_id.is_(None))
            )
        )
        assert unqualified_regular_count == 0
