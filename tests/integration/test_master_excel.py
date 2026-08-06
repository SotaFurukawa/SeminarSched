"""master_data.xlsx入出力とトランザクションの統合テスト。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

from collections.abc import Iterator, Mapping
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import event, func, select
from sqlalchemy.orm import Session

from summer_scheduler.infrastructure.db.base import Base
from summer_scheduler.infrastructure.db.database import Database, create_database
from summer_scheduler.infrastructure.db.models import (
    Campus,
    CourseProject,
    LessonRequest,
    Student,
    Subject,
    Teacher,
    TeacherQualification,
)
from summer_scheduler.infrastructure.excel import (
    IssueSeverity,
    MasterDataExcelService,
    MasterDataImportError,
)
from summer_scheduler.infrastructure.excel.schema import MASTER_DATA_SHEETS, SHEET_NAMES


@pytest.fixture
def project_session(tmp_path: Path) -> Iterator[tuple[Database, Session, int]]:
    """日本語パス上の空プロジェクトSession。"""
    database = create_database(tmp_path / "日本語DB" / "講習.jukuschedule")
    Base.metadata.create_all(database.engine)
    session = database.session_factory()
    campus = Campus(
        name="架空校",
        address_optional=None,
        logo_path_optional=None,
    )
    session.add(campus)
    session.flush()
    project = CourseProject(
        campus_id=campus.id,
        title="架空の夏期講習",
        start_date=date(2026, 7, 20),
        end_date=date(2026, 8, 31),
        status="draft",
        file_version=1,
    )
    session.add(project)
    session.commit()

    try:
        yield database, session, project.id
    finally:
        session.close()
        database.dispose()


def test_template_has_all_sheets_guidance_and_skippable_examples(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    path = service.export_template(tmp_path / "日本語テンプレート.xlsx")

    workbook = load_workbook(path)
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        for sheet_name in SHEET_NAMES:
            worksheet = workbook[sheet_name]
            assert worksheet.freeze_panes == "A2"
            assert worksheet.auto_filter.ref is not None
            assert worksheet["A2"].value == "はい"
            assert all(cell.comment is not None for cell in worksheet[1])
            assert len(worksheet.data_validations.dataValidation) >= 1
            assert all(
                worksheet.column_dimensions[get_column_letter(column_number)].width is not None
                for column_number, _cell in enumerate(worksheet[1], start=1)
            )
    finally:
        workbook.close()

    preview = service.preview_import(path)
    assert not preview.has_errors
    assert preview.rows == ()
    assert dict(preview.new_counts) == dict.fromkeys(SHEET_NAMES, 0)


def test_template_supports_id_entry_and_name_selection_helpers(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    path = MasterDataExcelService(session, project_id).export_template(
        tmp_path / "名前選択付きテンプレート.xlsx"
    )

    workbook = load_workbook(path, data_only=False)
    try:
        assert workbook.calculation.calcMode == "auto"
        assert workbook.calculation.fullCalcOnLoad is True
        assert workbook.calculation.forceFullCalc is True
        qualification = workbook["講師対応科目"]
        assert [cell.value for cell in qualification[1]][1:7] == [
            "講師ID（必須）",
            "講師名から選択",
            "講師名（確認）",
            "科目コード（必須）",
            "科目名から選択",
            "科目名（確認）",
        ]
        assert str(qualification["B3"].value).startswith("=IF(C3=")
        assert "MATCH(B3" in str(qualification["D3"].value)

        requests = workbook["受講希望"]
        assert [cell.value for cell in requests[1]][1:7] == [
            "生徒ID（必須）",
            "生徒名から選択",
            "生徒名（確認）",
            "科目コード（必須）",
            "科目名から選択",
            "科目名（確認）",
        ]
        assert str(requests["B3"].value).startswith("=IF(C3=")
        assert "MATCH(B3" in str(requests["D3"].value)
        assert any(
            "講師'!$C$3" in str(validation.formula1)
            for validation in requests.data_validations.dataValidation
        )
        student = workbook["生徒"]
        assert "生徒ID（必須）" in [cell.value for cell in student[1]]
        assert any(
            validation.formula1 == '"小1,小2,小3,小4,小5,小6,中1,中2,中3,高1,高2,高3"'
            for validation in student.data_validations.dataValidation
        )
    finally:
        workbook.close()


def test_legacy_headers_without_required_marker_remain_importable(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "旧ヘッダー互換.xlsx")
    workbook = load_workbook(source)
    try:
        for sheet_spec in MASTER_DATA_SHEETS:
            worksheet = workbook[sheet_spec.name]
            for cell in worksheet[1]:
                if isinstance(cell.value, str):
                    cell.value = cell.value.removesuffix("（必須）")
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    assert not preview.has_errors


def test_blank_student_teacher_and_subject_defaults_are_applied(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "空欄既定値.xlsx")

    workbook = load_workbook(source)
    try:
        student = _student_row()
        student["標準最大連続コマ数"] = None
        student["空きコマ許可"] = None
        student["有効"] = None
        _append_row(workbook["生徒"], student)

        teacher = _teacher_row()
        teacher["空きコマ許可"] = None
        teacher["有効"] = None
        _append_row(workbook["講師"], teacher)

        subject = _subject_row()
        subject["有効"] = None
        _append_row(workbook["科目"], subject)
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    assert not preview.has_errors
    service.apply_import(preview)
    session.commit()

    imported_student = session.scalar(select(Student).where(Student.external_id == "S-001"))
    imported_teacher = session.scalar(select(Teacher).where(Teacher.external_id == "T-001"))
    imported_subject = session.scalar(select(Subject).where(Subject.code == "JH_CUSTOM_MATH"))
    assert imported_student is not None
    assert imported_student.default_max_consecutive_slots == 2
    assert imported_student.allow_gap is False
    assert imported_student.active is True
    assert imported_teacher is not None
    assert imported_teacher.allow_gap is False
    assert imported_teacher.active is True
    assert imported_subject is not None
    assert imported_subject.active is True


def test_blank_qualification_and_lesson_request_defaults_are_applied(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "受講希望既定値.xlsx")
    workbook = load_workbook(source)
    try:
        _append_row(workbook["生徒"], _student_row())
        _append_row(workbook["講師"], _teacher_row())
        _append_row(workbook["科目"], _subject_row())
        qualification = _qualification_row()
        qualification["指導可能"] = None
        _append_row(workbook["講師対応科目"], qualification)
        request = _lesson_request_row()
        request["担当講師優先度"] = None
        request["1対1必須"] = None
        _append_row(workbook["受講希望"], request)
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    assert not preview.has_errors
    service.apply_import(preview)
    session.commit()

    qualification_entity = session.scalar(select(TeacherQualification))
    lesson_request = session.scalar(select(LessonRequest))
    assert qualification_entity is not None
    assert qualification_entity.can_teach is True
    assert lesson_request is not None
    assert lesson_request.regular_teacher_priority == 3
    assert lesson_request.one_to_one_required is False


def test_japanese_workbook_import_and_export_round_trip(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "架空データ_取込み.xlsx")
    _append_complete_valid_master(source)

    preview = service.preview_import(source)
    assert not preview.has_errors
    assert dict(preview.new_counts) == dict.fromkeys(SHEET_NAMES, 1)
    assert dict(preview.update_counts) == dict.fromkeys(SHEET_NAMES, 0)

    result = service.apply_import(preview)
    session.commit()

    assert dict(result.new_counts) == dict.fromkeys(SHEET_NAMES, 1)
    student = session.scalar(select(Student).where(Student.external_id == "S-001"))
    teacher = session.scalar(select(Teacher).where(Teacher.external_id == "T-001"))
    subject = session.scalar(select(Subject).where(Subject.code == "JH_CUSTOM_MATH"))
    assert student is not None and student.name == "架空 花子"
    assert teacher is not None and teacher.name == "架空 太郎"
    assert subject is not None and subject.school_level == "junior_high"
    assert session.scalar(select(func.count()).select_from(TeacherQualification)) == 1
    request = session.scalar(select(LessonRequest))
    assert request is not None
    assert request.required_sessions == 4
    assert request.regular_teacher_id_optional == teacher.id

    exported = service.export_template(tmp_path / "再出力_日本語.xlsx")
    second_preview = service.preview_import(exported)
    assert not second_preview.has_errors
    assert dict(second_preview.new_counts) == dict.fromkeys(SHEET_NAMES, 0)
    assert dict(second_preview.update_counts) == dict.fromkeys(SHEET_NAMES, 1)


def test_invalid_excel_row_blocks_all_updates_and_reports_location(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    session.add(
        Student(
            external_id="S-001",
            name="架空 変更前",
            grade="中学1年",
            default_max_consecutive_slots=2,
            allow_gap=False,
            note=None,
            active=True,
        ),
    )
    session.commit()
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "不正行を含む取込み.xlsx")

    workbook = load_workbook(source)
    try:
        _replace_or_append(
            workbook["生徒"],
            "生徒ID",
            "S-001",
            _student_row(name="架空 変更後"),
        )
        _append_row(workbook["講師"], _teacher_row())
        _append_row(workbook["科目"], _subject_row())
        _append_row(workbook["講師対応科目"], _qualification_row())
        invalid_request = _lesson_request_row()
        invalid_request["必要授業回数"] = 0
        _append_row(workbook["受講希望"], invalid_request)
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    matching_issues = [
        issue
        for issue in preview.issues
        if issue.sheet_name == "受講希望"
        and issue.row_number == 3
        and issue.column_name == "必要授業回数"
    ]
    assert matching_issues
    assert matching_issues[0].severity is IssueSeverity.ERROR

    with pytest.raises(MasterDataImportError, match="取込みエラー"):
        service.apply_import(preview)

    unchanged = session.scalar(select(Student).where(Student.external_id == "S-001"))
    assert unchanged is not None and unchanged.name == "架空 変更前"
    assert session.scalar(select(func.count()).select_from(Teacher)) == 0
    assert session.scalar(select(func.count()).select_from(Subject)) == 0


def test_database_error_rolls_back_every_imported_sheet(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "ロールバック確認.xlsx")
    _append_complete_valid_master(source)
    preview = service.preview_import(source)
    assert not preview.has_errors

    def fail_during_flush(
        _session: Session,
        _flush_context: object,
        _instances: object,
    ) -> None:
        raise RuntimeError("テスト用のDB書込み失敗")

    event.listen(session, "before_flush", fail_during_flush, once=True)
    with pytest.raises(RuntimeError, match="DB書込み失敗"):
        service.apply_import(preview)

    assert session.scalar(select(func.count()).select_from(Student)) == 0
    assert session.scalar(select(func.count()).select_from(Teacher)) == 0
    assert session.scalar(select(func.count()).select_from(Subject)) == 0
    assert session.scalar(select(func.count()).select_from(TeacherQualification)) == 0
    assert session.scalar(select(func.count()).select_from(LessonRequest)) == 0


def test_preview_reports_priority_qualification_duplicate_and_inactive_teacher(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "業務検証.xlsx")

    workbook = load_workbook(source)
    try:
        _append_row(workbook["生徒"], _student_row())
        _append_row(workbook["講師"], _teacher_row())
        _append_row(
            workbook["講師"],
            _teacher_row(
                teacher_id="T-STOP",
                name="架空 休止",
                active="いいえ",
            ),
        )
        _append_row(workbook["科目"], _subject_row())
        qualification = _qualification_row()
        qualification["指導可能"] = "いいえ"
        _append_row(workbook["講師対応科目"], qualification)
        request = _lesson_request_row()
        request["担当講師優先度"] = 5
        request["第1希望講師ID"] = "T-001"
        request["第2希望講師ID"] = "T-001"
        request["第3希望講師ID"] = "T-STOP"
        _append_row(workbook["受講希望"], request)
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    issues_by_code = {issue.code: issue for issue in preview.issues}
    assert issues_by_code["regular_teacher_not_qualified"].severity is IssueSeverity.ERROR
    assert issues_by_code["duplicate_preferred_teacher"].severity is IssueSeverity.WARNING
    assert issues_by_code["inactive_teacher"].column_name == "第3希望講師ID"
    assert issues_by_code["inactive_teacher"].row_number == 3


def test_preview_validates_duplicates_types_ranges_and_reference_ids(
    project_session: tuple[Database, Session, int],
    tmp_path: Path,
) -> None:
    _, session, project_id = project_session
    service = MasterDataExcelService(session, project_id)
    source = service.export_template(tmp_path / "行検証.xlsx")

    workbook = load_workbook(source)
    try:
        _append_row(workbook["生徒"], _student_row())
        _append_row(workbook["生徒"], _student_row(name="架空 重複ID"))
        second_student = _student_row(name="架空 次郎")
        second_student["生徒ID"] = "S-002"
        _append_row(workbook["生徒"], second_student)
        _append_row(workbook["科目"], _subject_row())

        unknown_qualification = _qualification_row()
        unknown_qualification["講師ID"] = "T-UNKNOWN"
        unknown_qualification["科目コード"] = "SUBJECT-UNKNOWN"
        _append_row(workbook["講師対応科目"], unknown_qualification)

        priority_five_without_teacher = _lesson_request_row()
        priority_five_without_teacher["生徒ID"] = "S-002"
        priority_five_without_teacher["通常担当講師ID"] = None
        _append_row(workbook["受講希望"], priority_five_without_teacher)

        invalid_ranges = _lesson_request_row()
        invalid_ranges["必要授業回数"] = "回数不正"
        invalid_ranges["担当講師優先度"] = 6
        _append_row(workbook["受講希望"], invalid_ranges)
        workbook.save(source)
    finally:
        workbook.close()

    preview = service.preview_import(source)
    codes = {issue.code for issue in preview.issues}
    assert "duplicate_row" in codes
    assert "unknown_teacher" in codes
    assert "unknown_subject" in codes
    assert "priority_five_without_regular_teacher" in codes
    assert any(
        issue.code == "invalid_type"
        and issue.column_name == "必要授業回数"
        and issue.row_number == 4
        for issue in preview.issues
    )
    assert any(
        issue.code == "invalid_type"
        and issue.column_name == "担当講師優先度"
        and issue.row_number == 4
        for issue in preview.issues
    )


def _append_complete_valid_master(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        _append_row(workbook["生徒"], _student_row())
        _append_row(workbook["講師"], _teacher_row())
        _append_row(workbook["科目"], _subject_row())
        _append_row(workbook["講師対応科目"], _qualification_row())
        _append_row(workbook["受講希望"], _lesson_request_row())
        workbook.save(path)
    finally:
        workbook.close()


def _student_row(*, name: str = "架空 花子") -> dict[str, object]:
    return {
        "例示行": "いいえ",
        "生徒ID": "S-001",
        "氏名": name,
        "学年": "中学2年",
        "標準最大連続コマ数": 2,
        "空きコマ許可": "いいえ",
        "備考": "架空のテストデータ",
        "有効": "はい",
    }


def _teacher_row(
    *,
    teacher_id: str = "T-001",
    name: str = "架空 太郎",
    active: str = "はい",
) -> dict[str, object]:
    return {
        "例示行": "いいえ",
        "講師ID": teacher_id,
        "氏名": name,
        "空きコマ許可": "いいえ",
        "備考": "架空のテストデータ",
        "有効": active,
    }


def _subject_row() -> dict[str, object]:
    return {
        "例示行": "いいえ",
        "科目コード": "JH_CUSTOM_MATH",
        "表示名": "中学校・数学（架空）",
        "学校段階": "中学校",
        "並び順": 100,
        "有効": "はい",
    }


def _qualification_row() -> dict[str, object]:
    return {
        "例示行": "いいえ",
        "講師ID": "T-001",
        "科目コード": "JH_CUSTOM_MATH",
        "指導可能": "はい",
        "備考": "明示資格",
    }


def _lesson_request_row() -> dict[str, object]:
    return {
        "例示行": "いいえ",
        "生徒ID": "S-001",
        "科目コード": "JH_CUSTOM_MATH",
        "必要授業回数": 4,
        "通常担当講師ID": "T-001",
        "担当講師優先度": 5,
        "第1希望講師ID": None,
        "第2希望講師ID": None,
        "第3希望講師ID": None,
        "1対1必須": "いいえ",
        "最大連続コマ数上書き": None,
        "空きコマ許可上書き": None,
        "備考": "架空の受講希望",
    }


def _append_row(worksheet: Worksheet, values_by_header: Mapping[str, object]) -> None:
    headers = [cell.value for cell in worksheet[1]]
    target_row = next(
        row_number
        for row_number in range(3, worksheet.max_row + 2)
        if worksheet.cell(row=row_number, column=1).value is None
    )
    for column_number, header in enumerate(headers, start=1):
        canonical_header = _canonical_header(str(header))
        if canonical_header in values_by_header:
            worksheet.cell(row=target_row, column=column_number).value = _cell_value(
                values_by_header[canonical_header]
            )


def _replace_or_append(
    worksheet: Worksheet,
    key_header: str,
    key_value: str,
    values_by_header: Mapping[str, object],
) -> None:
    headers = [cell.value for cell in worksheet[1]]
    canonical_headers = [_canonical_header(str(header)) for header in headers]
    key_column = canonical_headers.index(key_header) + 1
    target_row = next(
        (
            row_number
            for row_number in range(2, worksheet.max_row + 1)
            if worksheet.cell(row=row_number, column=key_column).value == key_value
        ),
        None,
    )
    if target_row is None:
        _append_row(worksheet, values_by_header)
        return
    for column_number, header in enumerate(headers, start=1):
        worksheet.cell(row=target_row, column=column_number).value = _cell_value(
            values_by_header.get(_canonical_header(str(header))),
        )


def _canonical_header(header: str) -> str:
    return header.removesuffix("（必須）")


def _cell_value(value: object) -> str | int | float | bool | date | None:
    if value is None or isinstance(value, str | int | float | bool | date):
        return value
    raise TypeError(f"Excelセルへ設定できないテスト値です: {type(value).__name__}")
