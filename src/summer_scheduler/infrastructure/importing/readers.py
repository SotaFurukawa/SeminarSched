"""xlsx／CSVを元の列名・セル値を保ったまま読み取る。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

import csv
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from summer_scheduler.infrastructure.importing.contracts import (
    CsvEncoding,
    ImportSourceError,
    SheetSummary,
    SourceFormat,
    SourceInspection,
    SourceRow,
    SourceTable,
    immutable_mapping,
)

GROUP_LESSON_SHEET = "集団授業"
GROUP_PARTICIPANT_SHEET = "受講者"


def inspect_source(
    path: Path,
    *,
    csv_encoding: CsvEncoding = CsvEncoding.AUTO,
) -> SourceInspection:
    """シート候補、ヘッダー、CSV文字コードを列挙する。"""
    source = _existing_source(path)
    source_format = _source_format(source)
    if source_format is SourceFormat.CSV:
        encoding = detect_csv_encoding(source, requested=csv_encoding)
        headers, rows = _read_csv_rows(source, encoding, preview_limit=None)
        return SourceInspection(
            source_path=source,
            source_format=source_format,
            sheets=(
                SheetSummary(
                    name=source.stem,
                    headers=headers,
                    data_row_count=len(rows),
                ),
            ),
            detected_encoding=encoding,
        )

    workbook = _open_workbook(source, read_only=True)
    try:
        sheets: list[SheetSummary] = []
        for worksheet in workbook.worksheets:
            first_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            headers = _headers_from_values(first_row)
            sheets.append(
                SheetSummary(
                    name=str(worksheet.title),
                    headers=headers,
                    data_row_count=max(int(worksheet.max_row) - 1, 0),
                ),
            )
    finally:
        workbook.close()
    return SourceInspection(source, source_format, tuple(sheets))


def read_source_table(
    path: Path,
    *,
    sheet_name: str | None = None,
    csv_encoding: CsvEncoding = CsvEncoding.AUTO,
    preview_limit: int | None = None,
) -> SourceTable:
    """選択したシートまたはCSVを読み、先頭行previewにも利用できる表を返す。"""
    if preview_limit is not None and preview_limit < 0:
        raise ValueError("preview_limitは0以上で指定してください。")

    source = _existing_source(path)
    source_format = _source_format(source)
    if source_format is SourceFormat.CSV:
        if sheet_name not in (None, source.stem):
            raise ImportSourceError("CSVではシートを選択できません。")
        encoding = detect_csv_encoding(source, requested=csv_encoding)
        headers, rows = _read_csv_rows(source, encoding, preview_limit=preview_limit)
        return SourceTable(
            source_path=source,
            source_format=source_format,
            sheet_name=source.stem,
            headers=headers,
            rows=rows,
            detected_encoding=encoding,
        )

    workbook = _open_workbook(source, read_only=True)
    try:
        selected_name = _selected_sheet_name(workbook.sheetnames, sheet_name)
        worksheet = workbook[selected_name]
        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = _headers_from_values(first_row)
        rows = _xlsx_rows(worksheet, headers, preview_limit)
    finally:
        workbook.close()
    return SourceTable(
        source_path=source,
        source_format=source_format,
        sheet_name=selected_name,
        headers=headers,
        rows=rows,
    )


def preview_source_table(
    path: Path,
    *,
    sheet_name: str | None = None,
    csv_encoding: CsvEncoding = CsvEncoding.AUTO,
    row_limit: int = 20,
) -> SourceTable:
    """ウィザードの先頭行preview用に件数を制限して読む。"""
    return read_source_table(
        path,
        sheet_name=sheet_name,
        csv_encoding=csv_encoding,
        preview_limit=row_limit,
    )


def read_group_workbook(
    path: Path,
    *,
    preview_limit: int | None = None,
) -> tuple[SourceTable, SourceTable]:
    """集団授業ブックの必須2シートを明示的に読み取る。"""
    inspection = inspect_source(path)
    if inspection.source_format is not SourceFormat.XLSX:
        raise ImportSourceError("集団授業ブックにはxlsxファイルを指定してください。")
    sheet_names = {sheet.name for sheet in inspection.sheets}
    missing = {
        GROUP_LESSON_SHEET,
        GROUP_PARTICIPANT_SHEET,
    }.difference(sheet_names)
    if missing:
        joined = "、".join(sorted(missing))
        raise ImportSourceError(f"集団授業ブックに必須シートがありません: {joined}")
    return (
        read_source_table(
            path,
            sheet_name=GROUP_LESSON_SHEET,
            preview_limit=preview_limit,
        ),
        read_source_table(
            path,
            sheet_name=GROUP_PARTICIPANT_SHEET,
            preview_limit=preview_limit,
        ),
    )


def detect_csv_encoding(
    path: Path,
    *,
    requested: CsvEncoding = CsvEncoding.AUTO,
) -> CsvEncoding:
    """BOMとstrict decodeでUTF-8-sig／UTF-8／CP932を判定する。"""
    source = _existing_source(path)
    if requested is not CsvEncoding.AUTO:
        try:
            source.read_text(encoding=requested.value)
        except UnicodeDecodeError as exc:
            raise ImportSourceError(
                f"CSVを指定された文字コード（{requested.value}）で読み込めません。"
            ) from exc
        return requested

    content = source.read_bytes()
    if content.startswith(b"\xef\xbb\xbf"):
        return CsvEncoding.UTF_8_SIG
    try:
        content.decode(CsvEncoding.UTF_8.value, errors="strict")
    except UnicodeDecodeError:
        try:
            content.decode(CsvEncoding.CP932.value, errors="strict")
        except UnicodeDecodeError as exc:
            raise ImportSourceError(
                "CSVの文字コードを判定できません。UTF-8またはCP932を明示してください。"
            ) from exc
        return CsvEncoding.CP932
    return CsvEncoding.UTF_8


def _read_csv_rows(
    source: Path,
    encoding: CsvEncoding,
    preview_limit: int | None,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    try:
        with source.open("r", encoding=encoding.value, newline="") as stream:
            reader = csv.reader(stream)
            header_values = next(reader, ())
            headers = _headers_from_values(header_values)
            rows = _source_rows(
                ((row_number, values) for row_number, values in enumerate(reader, start=2)),
                headers,
                preview_limit,
            )
    except (UnicodeDecodeError, csv.Error) as exc:
        raise ImportSourceError(f"CSVを読み込めません: {exc}") from exc
    return headers, rows


def _xlsx_rows(
    worksheet: Any,
    headers: Sequence[str],
    preview_limit: int | None,
) -> tuple[SourceRow, ...]:
    raw_rows = (
        (row_number, values)
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        )
    )
    return _source_rows(raw_rows, headers, preview_limit)


def _source_rows(
    raw_rows: Iterable[tuple[int, Sequence[object]]],
    headers: Sequence[str],
    preview_limit: int | None,
) -> tuple[SourceRow, ...]:
    rows: list[SourceRow] = []
    for row_number, raw_values in raw_rows:
        values = tuple(raw_values)
        if _is_blank_row(values):
            continue
        padded = (*values, *(None for _ in range(max(len(headers) - len(values), 0))))
        row_mapping = {
            header: padded[index] if index < len(padded) else None
            for index, header in enumerate(headers)
        }
        rows.append(SourceRow(row_number, immutable_mapping(row_mapping)))
        if preview_limit is not None and len(rows) >= preview_limit:
            break
    return tuple(rows)


def _headers_from_values(values: Sequence[object]) -> tuple[str, ...]:
    headers: list[str] = []
    seen: set[str] = set()
    for column_number, value in enumerate(values, start=1):
        header = str(value).strip() if value is not None else ""
        if not header:
            header = f"__column_{column_number}"
        if header in seen:
            raise ImportSourceError(f"ヘッダー「{header}」が重複しています。")
        seen.add(header)
        headers.append(header)
    return tuple(headers)


def _is_blank_row(values: Sequence[object]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _selected_sheet_name(sheet_names: Sequence[str], requested: str | None) -> str:
    if requested is not None:
        if requested not in sheet_names:
            raise ImportSourceError(f"シート「{requested}」が見つかりません。")
        return requested
    if len(sheet_names) == 1:
        return sheet_names[0]
    if not sheet_names:
        raise ImportSourceError("xlsxにシートがありません。")
    raise ImportSourceError("複数シートがあるため、取込み対象シートを選択してください。")


def _open_workbook(source: Path, *, read_only: bool) -> Any:
    try:
        return load_workbook(source, read_only=read_only, data_only=True)
    except Exception as exc:
        raise ImportSourceError(f"xlsxファイルを読み込めません: {exc}") from exc


def _existing_source(path: Path) -> Path:
    source = path.expanduser()
    if not source.is_file():
        raise ImportSourceError(f"入力ファイルが見つかりません: {source}")
    return source


def _source_format(source: Path) -> SourceFormat:
    suffix = source.suffix.casefold()
    if suffix == ".xlsx":
        return SourceFormat.XLSX
    if suffix == ".csv":
        return SourceFormat.CSV
    raise ImportSourceError("対応する入力形式は.xlsxまたは.csvです。")
