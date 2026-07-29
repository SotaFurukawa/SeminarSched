"""Phase 6のExcel・CSVレンダラーと原子的保存の結合テスト。"""

from __future__ import annotations

import csv
import logging
import os
import subprocess
from dataclasses import replace
from datetime import UTC, date, datetime, time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from summer_scheduler.infrastructure.exporting import (
    ASSIGNMENT_CSV_COLUMNS,
    CsvRenderer,
    ExcelRenderer,
    OutputDataError,
    OutputDestinationExistsError,
    OutputPermissionError,
    OutputRenderError,
)
from summer_scheduler.infrastructure.exporting.atomic_output import atomic_output_path
from summer_scheduler.infrastructure.exporting.html_renderer import HtmlRenderer
from summer_scheduler.reporting.data import (
    AssignmentRecord,
    DateRecord,
    GroupLessonRecord,
    LessonRequestRecord,
    OutputSelection,
    OutputSnapshot,
    ProjectRecord,
    SlotRecord,
    StudentRecord,
    SubjectRecord,
    TeacherRecord,
)
from summer_scheduler.reporting.issue_builder import build_issues_document
from summer_scheduler.reporting.layout import (
    LayoutCell,
    LayoutDocument,
    LayoutPage,
    LayoutRow,
    LayoutSection,
    LayoutTable,
)
from summer_scheduler.reporting.settings import OutputSettings
from summer_scheduler.reporting.student_builder import build_student_document
from summer_scheduler.reporting.teacher_builder import build_teacher_document
from summer_scheduler.reporting.timetable_builder import build_timetable_document


def test_excel_renderer_round_trips_layout_print_settings_and_safe_text(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "日本語の出力先" / "夏期講習時間割.xlsx"

    result = ExcelRenderer().render(_layout_document(), destination)

    assert result == destination.resolve()
    workbook = load_workbook(result, data_only=False)
    try:
        assert workbook.sheetnames == ["全体_時間割", "全体_時間割_2"]
        worksheet = workbook["全体_時間割"]
        assert worksheet["A1"].value == "夏期講習時間割"
        assert worksheet["A2"].value == "架空みらい校／2026年度 夏期講習"
        assert worksheet["A7"].value == '=HYPERLINK("https://invalid.example")'
        assert worksheet["A7"].data_type == "s"
        assert worksheet["B7"].alignment.wrap_text is True
        assert worksheet["B7"].alignment.shrink_to_fit is True
        assert worksheet["A7"].border.left.style == "thin"
        assert str(worksheet["A7"].fill.fgColor.rgb).endswith("FFF1CC")
        assert worksheet.row_dimensions[7].height == pytest.approx(34)
        assert {"A1:C1", "A2:C2", "A3:C3", "A7:A8", "A9:C9"} <= {
            str(item) for item in worksheet.merged_cells.ranges
        }
        assert str(worksheet.page_setup.paperSize) == "8"
        assert worksheet.page_setup.orientation == worksheet.ORIENTATION_LANDSCAPE
        assert worksheet.page_setup.fitToWidth == 1
        assert worksheet.page_setup.fitToHeight == 0
        assert worksheet.print_area == "'全体_時間割'!$A$1:$C$18"
        assert worksheet.print_title_rows == "$6:$6"
        assert len(worksheet.row_breaks.brk) == 1
        odd_footer = worksheet.oddFooter
        assert odd_footer is not None
        assert odd_footer.center.text == "ページ &P / &N"
        assert odd_footer.right.text == "個人情報を含みます"
        assert worksheet.column_dimensions["B"].width == pytest.approx(24)
    finally:
        workbook.close()
    html = HtmlRenderer().render_page(
        _layout_document(),
        _layout_document().sections[0].pages[0],
        OutputSettings(project_id=1),
        page_number=1,
        total_pages=3,
        font_family="Meiryo",
    )
    assert "background-color:#FFF1CC" in html


def test_excel_renderer_handles_many_pages_and_rejects_invalid_layout(
    tmp_path: Path,
) -> None:
    base = _layout_document()
    page = base.sections[0].pages[0]
    many_pages = replace(
        base,
        sections=(
            LayoutSection(
                name="大人数",
                pages=tuple(replace(page, subheading=f"分割 {index + 1}") for index in range(20)),
            ),
        ),
    )
    destination = tmp_path / "大人数_自動改ページ.xlsx"

    ExcelRenderer().render(many_pages, destination)

    workbook = load_workbook(destination, read_only=False)
    try:
        worksheet = workbook["大人数"]
        assert len(worksheet.row_breaks.brk) == 19
        assert worksheet.max_row > 100
    finally:
        workbook.close()

    invalid = replace(
        base,
        sections=(
            LayoutSection(
                name="不正",
                pages=(
                    LayoutPage(
                        heading="不正",
                        subheading="列超過",
                        tables=(
                            LayoutTable(
                                rows=(
                                    LayoutRow(
                                        cells=(
                                            LayoutCell("1"),
                                            LayoutCell("2"),
                                            LayoutCell("3"),
                                        )
                                    ),
                                ),
                                column_widths=(10, 10),
                            ),
                        ),
                    ),
                ),
            ),
        ),
    )
    invalid_destination = tmp_path / "不正.xlsx"
    with pytest.raises(OutputRenderError, match="列幅を超え"):
        ExcelRenderer().render(invalid, invalid_destination)
    assert not invalid_destination.exists()


def test_excel_renderer_accepts_all_shared_phase6_documents(tmp_path: Path) -> None:
    snapshot = _snapshot()
    settings = OutputSettings(
        project_id=1,
        days_per_page=1,
        teacher_columns_per_page=1,
    )
    builders = (
        build_timetable_document,
        build_student_document,
        build_teacher_document,
        build_issues_document,
    )

    for builder in builders:
        document = builder(snapshot, settings)
        destination = tmp_path / f"{document.report_code}.xlsx"
        ExcelRenderer(settings.style_rules).render(document, destination)
        workbook = load_workbook(destination, read_only=True)
        try:
            assert workbook.sheetnames
            assert workbook.active is not None
            assert workbook.active["A1"].value == document.title
        finally:
            workbook.close()


def test_atomic_output_rejects_unconfirmed_overwrite_and_preserves_original(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    destination = tmp_path / "既存時間割.xlsx"
    original = b"existing workbook must survive"
    destination.write_bytes(original)
    renderer = ExcelRenderer()

    with pytest.raises(OutputDestinationExistsError, match="既にあります"):
        renderer.render(_layout_document(), destination)
    assert destination.read_bytes() == original

    def deny_replace(_source: Path, _target: Path) -> None:
        raise PermissionError("locked by Excel")

    monkeypatch.setattr(os, "replace", deny_replace)
    with pytest.raises(OutputPermissionError, match="Excel等で開いて"):
        renderer.render(_layout_document(), destination, overwrite=True)
    assert destination.read_bytes() == original
    assert list(tmp_path.glob(f".{destination.name}.*.tmp")) == []


def test_excel_renderer_overwrites_only_after_success(tmp_path: Path) -> None:
    destination = tmp_path / "上書き確認.xlsx"
    destination.write_bytes(b"old")

    ExcelRenderer().render(_layout_document(), destination, overwrite=True)

    workbook = load_workbook(destination, read_only=True)
    try:
        assert workbook["全体_時間割"]["A1"].value == "夏期講習時間割"
    finally:
        workbook.close()


def test_atomic_cleanup_failure_is_logged_without_hiding_original_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    original_unlink = Path.unlink

    def fail_temporary_unlink(path: Path, missing_ok: bool = False) -> None:
        if path.suffix == ".tmp":
            sensitive_detail = rf"locked C:\Users\person\匿名生徒名\{path.name}"
            raise PermissionError(sensitive_detail)
        original_unlink(path, missing_ok=missing_ok)

    monkeypatch.setattr(Path, "unlink", fail_temporary_unlink)
    with (
        caplog.at_level(logging.WARNING),
        pytest.raises(RuntimeError, match="original render failure"),
    ):
        with atomic_output_path(tmp_path / "cleanup.xlsx", overwrite=False):
            raise RuntimeError("original render failure")

    assert "一時ファイルを削除できませんでした" in caplog.text
    assert "cleanup.xlsx" not in caplog.text
    assert "匿名生徒名" not in caplog.text
    assert r"C:\Users\person" not in caplog.text


def test_csv_renderer_writes_exact_columns_bom_group_rows_and_safe_values(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot()
    destination = tmp_path / "日本語フォルダー" / "割当て生データ.csv"

    result = CsvRenderer().render(snapshot, destination)

    assert result == destination.resolve()
    assert destination.read_bytes().startswith(b"\xef\xbb\xbf")
    with destination.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.DictReader(stream))
    assert tuple(rows[0]) == ASSIGNMENT_CSV_COLUMNS
    assert len(rows) == 4
    individual = next(
        row for row in rows if row["student_id"] == "S-001" and not row["group_lesson"]
    )
    assert individual["student_name"].startswith("'=")
    assert individual["teacher_name"].startswith("'@")
    assert individual["one_to_one_required"] == "true"
    assert individual["is_locked"] == "true"
    assert individual["is_manual"] == "true"
    assert individual["note"] == "カンマ, と改行\nを含む備考"
    groups = [row for row in rows if row["group_lesson"] == "G-001"]
    assert len(groups) == 2
    assert {row["student_id"] for row in groups} == {"S-001", "S-002"}
    assert all(row["slot"] == "A" for row in groups)
    assert all(row["start_time"] == "17:20" for row in groups)
    assert all(row["is_locked"] == "true" for row in groups)
    assert all("[集団]" not in row["group_lesson"] for row in groups)


def test_csv_renderer_supports_selection_and_utf8_without_bom(tmp_path: Path) -> None:
    destination = tmp_path / "絞込み.csv"
    selection = OutputSelection(
        dates=(date(2026, 8, 4),),
        teacher_ids=(2,),
        student_ids=(2,),
    )

    CsvRenderer().render(
        _snapshot(),
        destination,
        selection=selection,
        with_bom=False,
    )

    assert not destination.read_bytes().startswith(b"\xef\xbb\xbf")
    with destination.open("r", encoding="utf-8", newline="") as stream:
        rows = list(csv.DictReader(stream))
    assert len(rows) == 1
    assert rows[0]["date"] == "2026-08-04"
    assert rows[0]["teacher_id"] == "T-002"
    assert rows[0]["student_id"] == "S-002"
    assert rows[0]["group_lesson"] == "G-001"


def test_csv_numeric_only_external_ids_keep_leading_zero_as_safe_text(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot()
    numeric_ids = replace(
        snapshot,
        students=(
            replace(snapshot.students[0], external_id="00123"),
            snapshot.students[1],
        ),
        teachers=(
            replace(snapshot.teachers[0], external_id="00042"),
            snapshot.teachers[1],
        ),
    )
    destination = tmp_path / "数値ID.csv"

    CsvRenderer().render(numeric_ids, destination)

    with destination.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.DictReader(stream))
    individual = next(row for row in rows if not row["group_lesson"])
    # apostropheは既存の表計算ソフト向けsafe-prefixであり、数式形式へ変換しない。
    # 標準csv readerでも18列の通常文字列として読めることを契約に含める。
    assert tuple(individual) == ASSIGNMENT_CSV_COLUMNS
    assert len(individual) == len(ASSIGNMENT_CSV_COLUMNS)
    assert individual["student_id"] == "'00123"
    assert individual["teacher_id"] == "'00042"
    assert individual["date"] == "2026-08-03"
    assert individual["start_time"] == "17:10"
    assert not individual["student_id"].startswith("=")
    assert not individual["teacher_id"].startswith("=")


def test_csv_renderer_rejects_missing_references_without_partial_file(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot()
    broken = replace(
        snapshot,
        assignments=(replace(snapshot.assignments[0], lesson_request_id=999_999),),
        group_lessons=(),
    )
    destination = tmp_path / "参照欠落.csv"

    with pytest.raises(OutputDataError, match="受講希望が見つからない"):
        CsvRenderer().render(broken, destination)

    assert not destination.exists()


def test_renderers_reject_wrong_extensions(tmp_path: Path) -> None:
    with pytest.raises(OutputRenderError, match=r"\.xlsx"):
        ExcelRenderer().render(_layout_document(), tmp_path / "wrong.xls")
    with pytest.raises(OutputRenderError, match=r"\.csv"):
        CsvRenderer().render(_snapshot(), tmp_path / "wrong.txt")


def test_atomic_tmp_files_are_ignored_by_actual_git_rules() -> None:
    repository_root = Path(__file__).resolve().parents[2]

    result = subprocess.run(
        [
            "git",
            "check-ignore",
            "--no-index",
            "--verbose",
            "phase6-personal-data.tmp",
        ],
        cwd=repository_root,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )

    assert result.returncode == 0, result.stderr
    assert "*.tmp" in result.stdout
    assert "phase6-personal-data.tmp" in result.stdout


def _layout_document() -> LayoutDocument:
    table = LayoutTable(
        rows=(
            LayoutRow(
                cells=(
                    LayoutCell("生徒", role="header", alignment="center"),
                    LayoutCell("授業", role="header", alignment="center"),
                    LayoutCell("備考", role="header", alignment="center"),
                )
            ),
            LayoutRow(
                cells=(
                    LayoutCell(
                        '=HYPERLINK("https://invalid.example")',
                        row_span=2,
                        style_codes=("manual", "locked", "one_to_one"),
                    ),
                    LayoutCell("非常に長い日本語氏名と科目名\n[1対1] 中学校・数学"),
                    LayoutCell("1回目"),
                ),
                height_points_optional=34,
            ),
            LayoutRow(
                cells=(
                    LayoutCell("[固定] 2回目", style_codes=("locked",)),
                    LayoutCell("手動変更", style_codes=("manual",)),
                )
            ),
            LayoutRow(
                cells=(
                    LayoutCell(
                        "凡例 [1対1] [固定] [手]",
                        role="legend",
                        column_span=3,
                    ),
                )
            ),
        ),
        column_widths=(18, 24, 16),
        repeat_header_rows=1,
    )
    first_page = LayoutPage(
        heading="全体時間割",
        subheading="2026年8月3日／講師: 架空講師",
        tables=(table,),
        footer_note="文字と色を併記しています。",
    )
    second_page = replace(
        first_page,
        subheading="2026年8月4日／講師: 架空講師",
    )
    return LayoutDocument(
        report_code="overall",
        title="夏期講習時間割",
        campus_name="架空みらい校",
        course_name="2026年度 夏期講習",
        updated_text="更新日: 2026-07-29",
        sections=(
            LayoutSection(
                name="全体/時間割",
                pages=(first_page, second_page),
            ),
            LayoutSection(
                name="全体/時間割",
                pages=(first_page,),
            ),
        ),
        page_size="A3",
        orientation="landscape",
        margin_mm=8.0,
        font_size=8.0,
    )


def _snapshot() -> OutputSnapshot:
    project = ProjectRecord(
        id=1,
        title="2026年度 夏期講習",
        campus_name="架空みらい校",
        start_date=date(2026, 8, 3),
        end_date=date(2026, 8, 4),
        status="confirmed",
        generated_at=datetime(2026, 7, 29, 12, 0, tzinfo=UTC),
    )
    slots = (
        SlotRecord(
            id=10,
            code="A",
            display_name="Aコマ",
            start_time=time(17, 10),
            end_time=time(18, 30),
            sort_order=1,
            enabled=True,
        ),
        SlotRecord(
            id=20,
            code="B",
            display_name="Bコマ",
            start_time=time(18, 40),
            end_time=time(20, 0),
            sort_order=2,
            enabled=True,
        ),
    )
    students = (
        StudentRecord(
            id=1,
            external_id="S-001",
            name='=HYPERLINK("https://invalid.example")',
            grade="中学2年",
            note="",
            active=True,
        ),
        StudentRecord(
            id=2,
            external_id="S-002",
            name="架空 次郎",
            grade="中学3年",
            note="",
            active=True,
        ),
    )
    teachers = (
        TeacherRecord(
            id=1,
            external_id="T-001",
            name="@架空 講師",
            note="",
            active=True,
        ),
        TeacherRecord(
            id=2,
            external_id="T-002",
            name="架空 集団講師",
            note="",
            active=True,
        ),
    )
    subjects = (
        SubjectRecord(
            id=1,
            code="JH_MATH",
            name="中学校・数学",
            school_level="中学校",
        ),
    )
    requests = (
        LessonRequestRecord(
            id=100,
            student_id=1,
            subject_id=1,
            required_sessions=1,
            regular_teacher_id_optional=1,
            regular_teacher_priority=5,
            one_to_one_required=True,
            note="",
        ),
        LessonRequestRecord(
            id=200,
            student_id=2,
            subject_id=1,
            required_sessions=1,
            regular_teacher_id_optional=1,
            regular_teacher_priority=3,
            one_to_one_required=False,
            note="",
        ),
    )
    assignments = (
        AssignmentRecord(
            id=1000,
            lesson_request_id=100,
            session_index=1,
            day=date(2026, 8, 3),
            time_slot_id=10,
            teacher_id=1,
            is_locked=True,
            is_manual=True,
            note="カンマ, と改行\nを含む備考",
        ),
        AssignmentRecord(
            id=2000,
            lesson_request_id=200,
            session_index=1,
            day=date(2026, 8, 3),
            time_slot_id=20,
            teacher_id=1,
            is_locked=False,
            is_manual=False,
            note="",
        ),
    )
    groups = (
        GroupLessonRecord(
            id=3000,
            group_code="G-001",
            course_name="架空 夏期数学",
            grade="中学3年",
            subject_id=1,
            day=date(2026, 8, 4),
            start_time=time(17, 20),
            end_time=time(18, 20),
            teacher_id_optional=2,
            student_ids=(1, 2),
            room="架空教室",
            note="固定集団授業",
        ),
    )
    return OutputSnapshot(
        project=project,
        dates=(
            DateRecord(date(2026, 8, 3), True, ""),
            DateRecord(date(2026, 8, 4), True, ""),
        ),
        slots=slots,
        students=students,
        teachers=teachers,
        subjects=subjects,
        lesson_requests=requests,
        assignments=assignments,
        group_lessons=groups,
        unassigned=(),
        warnings=(),
    )
