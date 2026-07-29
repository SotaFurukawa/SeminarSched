"""Phase 3のアンケート・集団授業テンプレート生成テスト。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

from datetime import date, time
from pathlib import Path

from openpyxl import load_workbook

from summer_scheduler.infrastructure.importing import (
    GROUP_LESSON_SHEET,
    GROUP_PARTICIPANT_SHEET,
    INSTRUCTIONS_SHEET,
    STUDENT_AVAILABILITY_SHEET,
    TEACHER_AVAILABILITY_SHEET,
    group_lesson_schema,
    group_participant_schema,
    map_table,
    read_group_workbook,
    write_group_lessons_template,
    write_student_availability_template,
    write_teacher_availability_template,
)


def test_student_template_supports_dynamic_slots_rows_and_japanese_path(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "日本語フォルダー" / "生徒アンケート.xlsx"

    write_student_availability_template(
        destination,
        ("Y", "夜間"),
        rows=(
            {
                "student_id": "S001",
                "student_name": "架空 青空",
                "subject_code": "JH_MATH",
                "date": date(2026, 8, 4),
                "slot:Y": 2,
                "slot:夜間": 1,
                "note": "確認用",
            },
        ),
    )

    workbook = load_workbook(destination, data_only=True)
    try:
        assert workbook.sheetnames == [STUDENT_AVAILABILITY_SHEET, INSTRUCTIONS_SHEET]
        worksheet = workbook[STUDENT_AVAILABILITY_SHEET]
        headers = tuple(cell.value for cell in worksheet[1])
        assert headers == (
            "例示行",
            "生徒ID",
            "生徒名",
            "科目コード",
            "日付",
            "Y",
            "夜間",
            "第1希望講師ID",
            "第2希望講師ID",
            "第3希望講師ID",
            "備考",
        )
        assert worksheet["A2"].value == "はい"
        assert worksheet["A3"].value == "いいえ"
        assert worksheet["C3"].value == "架空 青空"
        validations = tuple(worksheet.data_validations.dataValidation)
        assert len(validations) == 2
        assert {validation.formula1 for validation in validations} == {'"0,1,2"'}
        assert all(
            "2:" in str(validation.sqref) and str(validation.sqref).endswith("10000")
            for validation in validations
        )
    finally:
        workbook.close()


def test_teacher_template_has_availability_validation(tmp_path: Path) -> None:
    destination = tmp_path / "講師希望.xlsx"

    write_teacher_availability_template(destination, ("Y", "Z", "A"))

    workbook = load_workbook(destination, data_only=True)
    try:
        worksheet = workbook[TEACHER_AVAILABILITY_SHEET]
        assert tuple(cell.value for cell in worksheet[1]) == (
            "例示行",
            "講師ID",
            "講師名",
            "日付",
            "Y",
            "Z",
            "A",
            "備考",
        )
        assert len(tuple(worksheet.data_validations.dataValidation)) == 3
        assert "0 = 不可、1 = 可能、2 = 希望" in workbook[INSTRUCTIONS_SHEET]["B3"].value
    finally:
        workbook.close()


def test_group_template_two_data_sheets_are_readable_and_examples_are_skipped(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "集団授業取込み.xlsx"
    write_group_lessons_template(
        destination,
        lessons=(
            {
                "group_lesson_id": "G001",
                "grade": "中学3年",
                "subject_code": "JH_MATH",
                "course_name": "架空講座",
                "date": date(2026, 8, 5),
                "start_time": time(17, 20),
                "end_time": time(18, 20),
                "teacher_id": "T001",
            },
        ),
        participants=(
            {
                "group_lesson_id": "G001",
                "student_id": "S001",
            },
        ),
    )

    lessons, participants = read_group_workbook(destination)
    assert lessons.sheet_name == GROUP_LESSON_SHEET
    assert participants.sheet_name == GROUP_PARTICIPANT_SHEET

    lesson_result = map_table(lessons, group_lesson_schema())
    participant_result = map_table(participants, group_participant_schema())

    assert lesson_result.skipped_example_rows == (2,)
    assert participant_result.skipped_example_rows == (2,)
    assert not lesson_result.has_errors
    assert not participant_result.has_errors
    assert lesson_result.rows[0].values["group_lesson_id"] == "G001"
    assert lesson_result.rows[0].values["start_time"] == time(17, 20)
    assert participant_result.rows[0].values["student_id"] == "S001"
