"""master_data.xlsxの構造検証と行単位の正規化。"""

# mypy: disable-error-code=import-untyped

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from summer_scheduler.infrastructure.excel.contracts import ImportIssue, IssueSeverity
from summer_scheduler.infrastructure.excel.schema import (
    MASTER_DATA_SHEETS,
    OPTIONAL_HELPER_HEADERS,
    CellValueError,
    SheetSpec,
    normalize_cell_value,
)


@dataclass(frozen=True, slots=True)
class ParsedRow:
    """構造検証を通過した1行。"""

    sheet_name: str
    row_number: int
    values: dict[str, object]


@dataclass(frozen=True, slots=True)
class ReadResult:
    """ブック読取り結果。"""

    rows: tuple[ParsedRow, ...]
    issues: tuple[ImportIssue, ...]


def read_master_data_workbook(path: Path) -> ReadResult:
    """必須5シートを読み、セル単位の検証結果を返す。"""
    source = path.expanduser()
    if not source.is_file():
        return ReadResult(
            rows=(),
            issues=(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "Excelファイルが見つかりません。",
                    code="file_not_found",
                ),
            ),
        )

    try:
        workbook = load_workbook(source, data_only=True, read_only=False)
    except Exception as exc:
        return ReadResult(
            rows=(),
            issues=(
                ImportIssue(
                    IssueSeverity.ERROR,
                    f"Excelファイルを読み込めません: {exc}",
                    code="workbook_open",
                ),
            ),
        )

    rows: list[ParsedRow] = []
    issues: list[ImportIssue] = []
    try:
        expected_names = {sheet.name for sheet in MASTER_DATA_SHEETS}
        for actual_name in workbook.sheetnames:
            if actual_name not in expected_names:
                issues.append(
                    ImportIssue(
                        IssueSeverity.WARNING,
                        "未定義のシートは取込み対象外です。",
                        sheet_name=actual_name,
                        code="unexpected_sheet",
                    ),
                )

        for sheet_spec in MASTER_DATA_SHEETS:
            if sheet_spec.name not in workbook.sheetnames:
                issues.append(
                    ImportIssue(
                        IssueSeverity.ERROR,
                        "必須シートがありません。",
                        sheet_name=sheet_spec.name,
                        code="missing_sheet",
                    ),
                )
                continue
            parsed_rows, sheet_issues = _read_sheet(workbook[sheet_spec.name], sheet_spec)
            rows.extend(parsed_rows)
            issues.extend(sheet_issues)
    finally:
        workbook.close()

    return ReadResult(tuple(rows), tuple(issues))


def _read_sheet(
    worksheet: Any,
    sheet_spec: SheetSpec,
) -> tuple[list[ParsedRow], list[ImportIssue]]:
    issues: list[ImportIssue] = []
    header_positions = _read_header_positions(worksheet, sheet_spec, issues)
    if any(issue.severity is IssueSeverity.ERROR and issue.row_number == 1 for issue in issues):
        return [], issues

    rows: list[ParsedRow] = []
    seen_keys: dict[tuple[object, ...], int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        raw_values = {
            column.key: worksheet.cell(
                row=row_number,
                column=header_positions[column.header],
            ).value
            for column in sheet_spec.columns
        }
        if _is_empty_row(raw_values.values()):
            continue

        example_column = sheet_spec.columns[0]
        try:
            example_marker = normalize_cell_value(
                raw_values[example_column.key],
                example_column,
            )
        except CellValueError as exc:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    str(exc),
                    sheet_name=sheet_spec.name,
                    row_number=row_number,
                    column_name=example_column.header,
                    code="invalid_type",
                ),
            )
            example_marker = False
        if example_marker is True:
            continue

        values: dict[str, object] = {"is_example": False}
        row_has_error = False
        for column in sheet_spec.columns[1:]:
            try:
                values[column.key] = normalize_cell_value(raw_values[column.key], column)
            except CellValueError as exc:
                issues.append(
                    ImportIssue(
                        IssueSeverity.ERROR,
                        str(exc),
                        sheet_name=sheet_spec.name,
                        row_number=row_number,
                        column_name=column.header,
                        code="invalid_type",
                    ),
                )
                row_has_error = True

        if row_has_error:
            continue

        unique_key = tuple(values[key] for key in sheet_spec.unique_keys)
        previous_row = seen_keys.get(unique_key)
        if previous_row is not None:
            duplicate_column = ", ".join(
                _header_for_key(sheet_spec, key) for key in sheet_spec.unique_keys
            )
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    f"{previous_row}行目と同じキーが重複しています。",
                    sheet_name=sheet_spec.name,
                    row_number=row_number,
                    column_name=duplicate_column,
                    code="duplicate_row",
                ),
            )
        else:
            seen_keys[unique_key] = row_number
        rows.append(ParsedRow(sheet_spec.name, row_number, values))

    return rows, issues


def _read_header_positions(
    worksheet: Any,
    sheet_spec: SheetSpec,
    issues: list[ImportIssue],
) -> dict[str, int]:
    actual_positions: dict[str, int] = {}
    for column_number in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_number).value
        if value is None:
            continue
        if not isinstance(value, str):
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "ヘッダーは文字列で入力してください。",
                    sheet_name=sheet_spec.name,
                    row_number=1,
                    column_name=getattr(
                        worksheet.cell(row=1, column=column_number),
                        "coordinate",
                        None,
                    ),
                    code="invalid_header",
                ),
            )
            continue
        header = value.strip()
        if header in actual_positions:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "同じヘッダーが重複しています。",
                    sheet_name=sheet_spec.name,
                    row_number=1,
                    column_name=header,
                    code="duplicate_header",
                ),
            )
            continue
        actual_positions[header] = column_number

    positions: dict[str, int] = {}
    for column in sheet_spec.columns:
        matching_headers = tuple(
            header
            for header in dict.fromkeys((column.header, column.template_header))
            if header in actual_positions
        )
        if len(matching_headers) > 1:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "旧形式と新形式の同じ列が重複しています。",
                    sheet_name=sheet_spec.name,
                    row_number=1,
                    column_name=column.header,
                    code="duplicate_header",
                ),
            )
        elif matching_headers:
            positions[column.header] = actual_positions[matching_headers[0]]
        else:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "必須列がありません。",
                    sheet_name=sheet_spec.name,
                    row_number=1,
                    column_name=column.header,
                    code="missing_column",
                ),
            )

    expected_headers = {
        header
        for column in sheet_spec.columns
        for header in (column.header, column.template_header)
    }
    optional_helpers = OPTIONAL_HELPER_HEADERS.get(sheet_spec.name, frozenset())
    for actual_header in actual_positions:
        if actual_header not in expected_headers and actual_header not in optional_helpers:
            issues.append(
                ImportIssue(
                    IssueSeverity.WARNING,
                    "未定義の列は取込み対象外です。",
                    sheet_name=sheet_spec.name,
                    row_number=1,
                    column_name=actual_header,
                    code="unexpected_column",
                ),
            )

    return positions


def _is_empty_row(values: Iterable[object]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _header_for_key(sheet_spec: SheetSpec, key: str) -> str:
    return next(column.header for column in sheet_spec.columns if column.key == key)
